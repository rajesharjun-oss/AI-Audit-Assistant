from __future__ import annotations

from collections import Counter
from io import BytesIO
import re

import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from export_utils import clean_name_consistency_rows


def metric_lines(value: object, empty: str) -> list[str]:
    text = str(value or "").strip()
    if not text or text == empty:
        return []
    return [line.strip() for line in text.splitlines() if line.strip()]


def build_excel_export(result) -> bytes:
    output = BytesIO()
    summary_rows = [
        {"Metric": "Checks performed", "Value": result.metrics.get("checks_performed_count", 0)},
        {"Metric": "Checks passed", "Value": result.metrics.get("checks_passed_count", 0)},
        {"Metric": "Checks skipped", "Value": result.metrics.get("checks_skipped_count", 0)},
        {"Metric": "Findings", "Value": result.metrics.get("findings", 0)},
        {"Metric": "Review prompts not elevated", "Value": result.metrics.get("review_prompts_not_elevated_count", 0)},
        {"Metric": "High findings", "Value": result.metrics.get("high", 0)},
        {"Metric": "Medium findings", "Value": result.metrics.get("medium", 0)},
        {"Metric": "Low findings", "Value": result.metrics.get("low", 0)},
        {"Metric": "OCR text coverage", "Value": result.metrics.get("ocr_text_coverage", result.metrics.get("extraction_coverage", "0%"))},
        {"Metric": "OCR table candidates", "Value": result.metrics.get("ocr_tables", 0)},
        {"Metric": "Statement structure confidence", "Value": result.metrics.get("statement_structure_confidence", "0%")},
        {"Metric": "Note structure confidence", "Value": result.metrics.get("note_structure_confidence", "0%")},
        {"Metric": "Table arithmetic confidence", "Value": table_arithmetic_display(result)},
        {"Metric": "notes_section_start_page", "Value": result.metrics.get("notes_section_start_page", "Not detected")},
        {"Metric": "notes_heading_snippet", "Value": result.metrics.get("notes_heading_snippet", "No reliable notes heading detected.")},
        {"Metric": "cautious_note_validation_enabled", "Value": result.metrics.get("cautious_note_validation_enabled", False)},
        {"Metric": "note_validation_mode", "Value": result.metrics.get("note_validation_mode", "skipped")},
        {"Metric": "note_reference_rows_detected", "Value": result.metrics.get("note_reference_rows_detected", 0)},
        {"Metric": "note_headings_detected", "Value": result.metrics.get("note_headings_detected", 0)},
        {"Metric": "note_reference_findings", "Value": result.metrics.get("note_reference_findings", 0)},
        {"Metric": "AI review status", "Value": result.metrics.get("ai_review_status", "Not started")},
        {"Metric": "AI review mode", "Value": result.metrics.get("ai_review_mode", "standard")},
        {"Metric": "AI policy review status", "Value": result.metrics.get("ai_policy_review_status", "disabled")},
        {"Metric": "AI policy review message", "Value": result.metrics.get("ai_policy_review_message", "")},
        {"Metric": "AI policy review summary", "Value": result.metrics.get("ai_policy_review_summary", "")},
        {"Metric": "AI full review status", "Value": result.metrics.get("ai_full_review_status", "disabled")},
        {"Metric": "AI full review message", "Value": result.metrics.get("ai_full_review_message", "")},
        {"Metric": "AI full review summary", "Value": result.metrics.get("ai_full_review_summary", "")},
        {"Metric": "AI finding review status", "Value": result.metrics.get("ai_finding_review_status", "disabled")},
        {"Metric": "AI finding review message", "Value": result.metrics.get("ai_finding_review_message", "")},
        {"Metric": "AI finding review summary", "Value": result.metrics.get("ai_finding_review_summary", "")},
        {"Metric": "AI finding review count", "Value": result.metrics.get("ai_finding_reviewed", 0)},
        {"Metric": "AI finding review suppressed", "Value": result.metrics.get("ai_finding_suppressed", 0)},
    ]
    summary_rows.extend(ai_enhanced_summary_rows(result))
    checks_performed = [{"Check performed": item} for item in metric_lines(result.metrics.get("checks_performed"), "No deterministic checks completed.")]
    checks_skipped = checks_skipped_rows(result) or [{"Check area": "None", "Reason skipped": "No major checks skipped."}]
    check_results = result.metrics.get("check_results", [])
    if not isinstance(check_results, list) or not check_results:
        check_results = [{"Check": "No deterministic checks completed.", "Result": "Skipped", "Severity": "", "Evidence": ""}]
    detected_profile = result.metrics.get("detected_profile", {})
    profile_rows = [{"Field": key, "Detected value": value} for key, value in detected_profile.items()] if isinstance(detected_profile, dict) else []
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(profile_rows).to_excel(writer, sheet_name="Detected profile", index=False)
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)
        finding_summary_rows = finding_summary_rows_for_result(result) or [{"Severity": "", "Category": "", "Page reference": "", "Issue": "No automated findings were identified.", "Recommendation": ""}]
        pd.DataFrame(finding_summary_rows).to_excel(writer, sheet_name="Findings summary", index=False)
        ai_finding_status = str(result.metrics.get("ai_finding_review_status", "disabled") or "disabled")
        ai_finding_message = str(result.metrics.get("ai_finding_review_message", "") or "").strip()
        ai_finding_default_title = {
            "disabled": "AI finding review not enabled.",
            "unavailable": "AI finding review was enabled but is unavailable in this environment.",
            "skipped": "AI finding review was enabled but no weak deterministic findings were eligible.",
            "error": "AI finding review was enabled but failed during execution.",
            "deferred": "AI finding review was deferred due to API availability or rate limiting.",
            "completed": "AI finding review completed but returned no adjudication rows.",
        }.get(ai_finding_status, "AI finding review returned no rows.")
        ai_finding_rows = result.metrics.get("ai_finding_export", []) or [{"Finding ID": "", "Issue": ai_finding_default_title, "AI status": ai_finding_status, "Reason": ai_finding_message}]
        pd.DataFrame(ai_finding_rows).to_excel(writer, sheet_name="AI finding review", index=False)
        review_comment_rows = review_comment_rows_for_result(result) or [{"S/N": "", "Section / Statement / Note": "", "Page number": "", "Account / line item": "", "Current wording / amount / reference": "", "Issue identified": "No automated findings were identified.", "Expected correction / recommendation": "", "Category": "", "Priority": "", "Status": "Noted", "Reviewer comments": ""}]
        pd.DataFrame(review_comment_rows).to_excel(writer, sheet_name="Review comments", index=False)
        review_prompt_rows = review_prompts_not_elevated_rows(result) or [{"Issue": "No low-confidence review prompts were withheld from the exception register."}]
        pd.DataFrame(review_prompt_rows).to_excel(writer, sheet_name="Review prompts not elevated", index=False)
        ai_status = str(result.metrics.get("ai_policy_review_status", "disabled") or "disabled")
        ai_message = str(result.metrics.get("ai_policy_review_message", "") or "").strip()
        ai_default_title = {
            "disabled": "AI policy review not enabled.",
            "unavailable": "AI policy review was enabled but is unavailable in this environment.",
            "skipped": "AI policy review was enabled but no suitable policy/disclosure context was detected.",
            "error": "AI policy review was enabled but failed during execution.",
            "completed": "AI policy review completed but returned no observation rows.",
        }.get(ai_status, "AI policy review returned no rows.")
        ai_policy_rows = result.metrics.get("ai_policy_export", []) or [{"Title": ai_default_title, "Status": ai_status, "Message": ai_message}]
        pd.DataFrame(ai_policy_rows).to_excel(writer, sheet_name="AI policy judgement", index=False)
        ai_full_status = str(result.metrics.get("ai_full_review_status", "disabled") or "disabled")
        ai_full_message = str(result.metrics.get("ai_full_review_message", "") or "").strip()
        ai_full_default_title = {
            "disabled": "AI full review not enabled.",
            "unavailable": "AI full review was enabled but is unavailable in this environment.",
            "skipped": "AI full review was enabled but no suitable extracted context was detected.",
            "error": "AI full review was enabled but failed during execution.",
            "deferred": "AI full review was deferred due to API availability or rate limiting.",
            "completed": "AI full review completed but returned no observation rows.",
        }.get(ai_full_status, "AI full review returned no rows.")
        ai_full_rows = result.metrics.get("ai_full_export", []) or []
        if not ai_full_rows and ai_full_status == "completed" and result.metrics.get("ai_review_comment_rows"):
            ai_full_rows = [
                normalize_ai_review_comment_row(row, index, result)
                for index, row in enumerate(result.metrics.get("ai_review_comment_rows", []) or [], start=1)
                if isinstance(row, dict)
            ]
        if not ai_full_rows:
            ai_full_rows = [{"Title": ai_full_default_title, "Status": ai_full_status, "Message": ai_full_message}]
        pd.DataFrame(ai_full_rows).to_excel(writer, sheet_name="AI full review", index=False)
        ai_suppressed_rows = result.metrics.get("ai_suppressed_findings", []) or [{"Finding ID": "", "Issue": "No AI-suppressed false positives.", "AI status": ai_finding_status, "Reason": ai_finding_message}]
        pd.DataFrame(ai_suppressed_rows).to_excel(writer, sheet_name="AI suppressed findings", index=False)
        ai_evidence_rows = result.metrics.get("ai_evidence_packs", []) or [{"Evidence type": "None", "AI role": "AI review was not run or no evidence packs were eligible."}]
        pd.DataFrame(ai_evidence_rows).to_excel(writer, sheet_name="AI evidence packs", index=False)
        ai_error_rows = result.metrics.get("ai_error_log", []) or [{"Attempt": "", "Error category": "None", "Error message": "No AI provider errors recorded."}]
        pd.DataFrame(ai_error_rows).to_excel(writer, sheet_name="AI error log", index=False)
        exception_rows = finding_rows(result) or [
            {
                "ID": "",
                "Status": "Noted",
                "Severity": "",
                "Category": "",
                "Check type": "",
                "Confidence": "",
                "Page reference": "",
                "Note reference": "",
                "Statement": "",
                "Line item": "",
                "Referenced note": "",
                "Suggested note": "",
                "Match confidence": "",
                "Reason": "",
                "Current year amount found?": "",
                "Prior year amount found?": "",
                "Amount found in note": "",
                "Alternative note found": "",
                "Amount match confidence": "",
                "AI review status": "",
                "AI review confidence": "",
                "AI review reason": "",
                "Location": "",
                "Issue": "No automated findings were identified.",
                "Evidence": "",
                "Recommendation": "",
                "Reviewer comment": "",
                "Prepared by": "",
                "Reviewed by": "",
                "Date cleared": "",
            }
        ]
        pd.DataFrame(exception_rows).to_excel(writer, sheet_name="Exception register", index=False)

        note_agreement_rows = note_agreement_result_rows(result)
        primary_rows = note_agreement_rows or [{"Statement": "No lines detected"}]
        no_notes_rows = [r for r in note_agreement_rows if r.get("Has note?") == "No"] or [{"Statement": "None found"}]
        linked_rows = [r for r in note_agreement_rows if r.get("Has note?") == "Yes"] or [{"Statement": "None found"}]
        pd.DataFrame(primary_rows).to_excel(writer, sheet_name="Primary statement line items", index=False)
        pd.DataFrame(no_notes_rows).to_excel(writer, sheet_name="Items without notes summary", index=False)
        pd.DataFrame(linked_rows).to_excel(writer, sheet_name="Note-linked review", index=False)

        notes_detected_rows = note_heading_rows(result) or [{"Note": "None found"}]
        notes_heading_candidate_rows = notes_heading_candidate_rows_for_result(result)
        ocr_statement_rows = ocr_statement_row_rows(result)
        pd.DataFrame(notes_detected_rows).to_excel(writer, sheet_name="Notes detected", index=False)
        pd.DataFrame(notes_heading_candidate_rows).to_excel(writer, sheet_name="Notes heading candidates", index=False)
        pd.DataFrame(ocr_statement_rows).to_excel(writer, sheet_name="OCR statement rows", index=False)

        policy_rows = result.metrics.get("policy_export", []) or [{"Paragraph reviewed": "None found"}]
        pd.DataFrame(policy_rows).to_excel(writer, sheet_name="Notes 1 and 2 policy review", index=False)
        unref_rows = result.metrics.get("unreferenced_notes", []) or [{"Note": "None", "Heading": "None found", "Comment": "All notes referenced or filtered"}]
        pd.DataFrame(unref_rows).to_excel(writer, sheet_name="Unreferenced notes", index=False)

        cross_export = result.metrics.get("cross_page_export", {})
        amount_rows = [
            translate_row_page_fields(row, result, ("Pages checked", "Page", "Context", "Issue"))
            for row in (cross_export.get("key_amounts", []) or [{"Metric": "None found"}])
        ]
        name_rows = [
            translate_row_page_fields(row, result, ("Page 1", "Page 2"))
            for row in (clean_name_consistency_rows(cross_export.get("names", [])) or [{"Name variant 1": "None found"}])
        ]
        date_rows = [
            translate_row_page_fields(row, result, ("Page", "Comment"))
            for row in (cross_export.get("dates", []) or [{"Date found": "None found"}])
        ]
        grammar_rows = [
            translate_row_page_fields(row, result, ("Page", "Context"))
            for row in (cross_export.get("grammar", []) or [{"Page": "None found"}])
        ]
        pd.DataFrame(amount_rows).to_excel(writer, sheet_name="Key amount consistency", index=False)
        pd.DataFrame(name_rows).to_excel(writer, sheet_name="Name consistency", index=False)
        pd.DataFrame(date_rows).to_excel(writer, sheet_name="Date consistency", index=False)
        pd.DataFrame(grammar_rows).to_excel(writer, sheet_name="Grammar review", index=False)

        canonical_check_rows = result.metrics.get("canonical_recalculation_checks", []) or [{"Check": "None", "Status": "Not tested", "Recommendation": "No canonical recalculation checks were available."}]
        canonical_audit_rows = result.metrics.get("canonical_extraction_audit", []) or [{"Page": "None", "Reason": "No canonical statement facts were parsed."}]
        deterministic_section_rows = result.metrics.get("deterministic_section_map", []) or [{"Page": "None", "Section": "No section map available"}]
        deterministic_table_rows = result.metrics.get("deterministic_table_classification", []) or [{"Page": "None", "Table type": "No table classification available"}]
        contents_agreement_rows = result.metrics.get("contents_agreement", []) or [{"Statement": "None", "Status": "Not tested", "Reason": "No contents-page statement references were detected."}]
        pd.DataFrame(canonical_check_rows).to_excel(writer, sheet_name="Canonical recalculation checks", index=False)
        pd.DataFrame(canonical_audit_rows).to_excel(writer, sheet_name="Canonical extraction audit", index=False)
        pd.DataFrame(deterministic_section_rows).to_excel(writer, sheet_name="Deterministic section map", index=False)
        pd.DataFrame(deterministic_table_rows).to_excel(writer, sheet_name="Table classification", index=False)
        pd.DataFrame(contents_agreement_rows).to_excel(writer, sheet_name="Contents agreement", index=False)

        pd.DataFrame(check_results).to_excel(writer, sheet_name="Checks results", index=False)
        pd.DataFrame(checks_performed).to_excel(writer, sheet_name="Checks performed", index=False)
        pd.DataFrame(checks_skipped).to_excel(writer, sheet_name="Checks skipped", index=False)
        skipped_summary_rows = result.metrics.get("skipped_table_summary", []) or [{"Skipped check group": "None", "Reason skipped": "No table-specific skips recorded."}]
        pd.DataFrame(skipped_summary_rows).to_excel(writer, sheet_name="Skipped checks summary", index=False)
        skipped_table_rows = result.metrics.get("skipped_table_details", []) or [{"Page": "None", "Reason skipped": "No table-specific skips recorded."}]
        pd.DataFrame(skipped_table_rows).to_excel(writer, sheet_name="Skipped table details", index=False)

        for worksheet in writer.book.worksheets:
            worksheet.freeze_panes = "A2"
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F2937")
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            for column_cells in worksheet.columns:
                max_length = max((len(str(cell.value or "")) for cell in column_cells), default=10)
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 70)

        format_exception_register_sheet(writer.book["Exception register"])
        format_excel_table_sheet(writer.book["Findings summary"], "FindingsSummary")
        format_excel_table_sheet(writer.book["Review comments"], "ReviewComments")
        format_excel_table_sheet(writer.book["Review prompts not elevated"], "ReviewPromptsNotElevated")
        format_excel_table_sheet(writer.book["Primary statement line items"], "PrimaryLineItems")
        format_excel_table_sheet(writer.book["Items without notes summary"], "NoNotesSummary")
        format_excel_table_sheet(writer.book["Note-linked review"], "NoteLinkedReview")
        format_excel_table_sheet(writer.book["Notes detected"], "NotesDetected")
        format_excel_table_sheet(writer.book["Notes heading candidates"], "NotesHeadingCandidates")
        format_excel_table_sheet(writer.book["OCR statement rows"], "OCRStatementRows")
        format_excel_table_sheet(writer.book["Notes 1 and 2 policy review"], "PolicyReview")
        format_excel_table_sheet(writer.book["AI policy judgement"], "AIPolicyJudgement")
        format_excel_table_sheet(writer.book["AI full review"], "AIFullReview")
        format_excel_table_sheet(writer.book["AI finding review"], "AIFindingReview")
        format_excel_table_sheet(writer.book["AI suppressed findings"], "AISuppressedFindings")
        format_excel_table_sheet(writer.book["AI evidence packs"], "AIEvidencePacks")
        format_excel_table_sheet(writer.book["AI error log"], "AIErrorLog")
        format_excel_table_sheet(writer.book["Unreferenced notes"], "UnreferencedNotes")
        format_excel_table_sheet(writer.book["Key amount consistency"], "AmountConsistency")
        format_excel_table_sheet(writer.book["Name consistency"], "NameConsistency")
        format_excel_table_sheet(writer.book["Date consistency"], "DateConsistency")
        format_excel_table_sheet(writer.book["Grammar review"], "GrammarReview")
        format_excel_table_sheet(writer.book["Canonical recalculation checks"], "CanonicalRecalculationChecks")
        format_excel_table_sheet(writer.book["Canonical extraction audit"], "CanonicalExtractionAudit")
        format_excel_table_sheet(writer.book["Deterministic section map"], "DeterministicSectionMap")
        format_excel_table_sheet(writer.book["Table classification"], "TableClassification")
        format_excel_table_sheet(writer.book["Contents agreement"], "ContentsAgreement")
        format_excel_table_sheet(writer.book["Checks results"], "ChecksResults")
        format_excel_table_sheet(writer.book["Skipped checks summary"], "SkippedChecksSummary")
        format_excel_table_sheet(writer.book["Skipped table details"], "SkippedTableDetails")
    return output.getvalue()


def checks_skipped_rows(result) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for item in metric_lines(result.metrics.get("checks_skipped"), "No major checks skipped."):
        row = parse_skipped_check(item)
        rows.append(
            translate_row_page_fields(
                row,
                result,
                (
                    "Page reference",
                    "Affected statement/note",
                    "Reason skipped",
                    "Reviewer action",
                    "Automation requirement",
                    "Original message",
                ),
            )
        )
    return rows


def review_prompts_not_elevated_rows(result) -> list[dict[str, str]]:
    rows = result.metrics.get("review_prompts_not_elevated", [])
    if not isinstance(rows, list):
        return []
    cleaned: list[dict[str, str]] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        # These rows are built after reviewer-page translation, so translating the
        # page reference again would shift printed page numbers incorrectly.
        translated = dict(row)
        for field in ("Evidence", "Issue"):
            if field in translated:
                translated[field] = translate_page_tokens(translated.get(field, ""), result)
        cleaned.append(translated)
    return cleaned



def ai_enhanced_summary_rows(result) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    priority_counts = Counter(str(finding.severity or "") for finding in result.findings)
    category_counts = Counter(review_comment_category_for_finding(finding) for finding in result.findings)
    rows.extend(
        [
            {"Metric": "Priority - High", "Value": priority_counts.get("High", 0)},
            {"Metric": "Priority - Medium", "Value": priority_counts.get("Medium", 0)},
            {"Metric": "Priority - Low", "Value": priority_counts.get("Low", 0)},
        ]
    )
    for category in REVIEW_COMMENT_CATEGORIES:
        rows.append({"Metric": f"Category - {category}", "Value": category_counts.get(category, 0)})
    high_findings = [finding for finding in result.findings if finding.severity == "High"]
    if high_findings:
        rows.append({"Metric": "Key high-priority findings", "Value": "\n".join(f"{page_reference_for_finding(finding, result) or finding.location}: {finding.issue}" for finding in high_findings[:5])})
    else:
        rows.append({"Metric": "Key high-priority findings", "Value": "No high-priority findings identified."})
    ai_summary = result.metrics.get("ai_summary_fields", {}) if isinstance(result.metrics.get("ai_summary_fields", {}), dict) else {}
    default_conclusion = "Not ready for final sign-off until open High and Medium findings are cleared." if (priority_counts.get("High", 0) or priority_counts.get("Medium", 0)) else "No high- or medium-priority automated findings; reviewer should still complete professional review procedures."
    rows.extend(
        [
            {"Metric": "Overall conclusion on final sign-off", "Value": ai_summary.get("Overall sign-off conclusion") or default_conclusion},
            {"Metric": "Recommended immediate action points", "Value": ai_summary.get("Recommended immediate action points") or immediate_action_points(result)},
            {"Metric": "Cash flow correctness note", "Value": ai_summary.get("Cash flow correctness note") or cash_flow_summary_note(result)},
            {"Metric": "Regulatory-reference note", "Value": ai_summary.get("Regulatory-reference note") or regulatory_summary_note(result)},
            {"Metric": "Casting and cross-casting note", "Value": ai_summary.get("Casting and cross-casting note") or casting_summary_note(result)},
        ]
    )
    return rows


REVIEW_COMMENT_CATEGORIES = (
    "Spelling / Grammar",
    "Regulatory Reference",
    "Note Cross-reference",
    "Casting",
    "Cross-casting",
    "Cash Flow",
    "Disclosure",
    "Presentation",
    "Internal Consistency",
)


def review_comment_rows_for_result(result) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()
    for finding in result.findings:
        metadata = finding.metadata or {}
        row = {
            "S/N": str(len(rows) + 1),
            "Section / Statement / Note": metadata.get("statement") or note_reference_for_finding(finding, result) or finding.category,
            "Page number": page_reference_for_finding(finding, result),
            "Account / line item": metadata.get("line_item", ""),
            "Current wording / amount / reference": translate_page_tokens(finding.evidence, result),
            "Issue identified": finding.issue,
            "Expected correction / recommendation": finding.recommendation,
            "Category": review_comment_category_for_finding(finding),
            "Priority": finding.severity,
            "Status": "Open",
            "Reviewer comments": metadata.get("ai_review_reason", ""),
        }
        key = (row["Page number"], row["Category"], row["Issue identified"])
        if key not in seen:
            seen.add(key)
            rows.append(row)
    for raw_row in result.metrics.get("ai_review_comment_rows", []) or []:
        if not isinstance(raw_row, dict):
            continue
        row = normalize_ai_review_comment_row(raw_row, len(rows) + 1, result)
        key = (row["Page number"], row["Category"], row["Issue identified"])
        if row["Issue identified"] and key not in seen:
            seen.add(key)
            rows.append(row)
    for index, row in enumerate(rows, start=1):
        row["S/N"] = str(index)
    return rows


def normalize_ai_review_comment_row(row: dict[str, object], index: int, result) -> dict[str, str]:
    return {
        "S/N": str(row.get("S/N") or row.get("s_n") or row.get("sn") or index),
        "Section / Statement / Note": str(row.get("Section / Statement / Note") or row.get("section_or_statement_or_note") or row.get("section") or "").strip(),
        "Page number": translate_page_tokens(row.get("Page number") or row.get("page_number") or row.get("page_reference") or "", result),
        "Account / line item": str(row.get("Account / line item") or row.get("account_or_line_item") or row.get("line_item") or "").strip(),
        "Current wording / amount / reference": str(row.get("Current wording / amount / reference") or row.get("current_wording_amount_reference") or row.get("current_reference") or row.get("evidence") or "").strip(),
        "Issue identified": str(row.get("Issue identified") or row.get("issue_identified") or row.get("issue") or "").strip(),
        "Expected correction / recommendation": str(row.get("Expected correction / recommendation") or row.get("expected_correction_recommendation") or row.get("recommendation") or "").strip(),
        "Category": normalize_review_comment_category(row.get("Category") or row.get("category") or row.get("issue_identified") or ""),
        "Priority": normalize_priority(row.get("Priority") or row.get("priority") or row.get("severity") or "Low"),
        "Status": str(row.get("Status") or row.get("status") or "Open").strip() or "Open",
        "Reviewer comments": str(row.get("Reviewer comments") or row.get("reviewer_comments") or row.get("rationale") or "").strip(),
    }


def review_comment_category_for_finding(finding) -> str:
    text = " ".join(str(part or "") for part in (finding.category, finding.issue, finding.evidence, finding.recommendation))
    return normalize_review_comment_category(text)


def normalize_review_comment_category(value: object) -> str:
    lower = str(value or "").lower()
    if "spell" in lower or "grammar" in lower or "typograph" in lower or "wording" in lower:
        return "Spelling / Grammar"
    if "regulat" in lower or "cama" in lower or "frc" in lower or "icfr" in lower or "securities" in lower or "tax law" in lower:
        return "Regulatory Reference"
    if "note" in lower and ("reference" in lower or "cross" in lower or "agreement" in lower):
        return "Note Cross-reference"
    if "cross" in lower and ("cast" in lower or "tie" in lower or "agreement" in lower):
        return "Cross-casting"
    if "cash" in lower or "ias 7" in lower:
        return "Cash Flow"
    if "cast" in lower or "total" in lower or "subtotal" in lower or "arithmetic" in lower or "rounding" in lower:
        return "Casting"
    if "disclosure" in lower or "missing" in lower:
        return "Disclosure"
    if "present" in lower or "format" in lower or "caption" in lower:
        return "Presentation"
    return "Internal Consistency"


def normalize_priority(value: object) -> str:
    text = str(value or "").strip().title()
    return text if text in {"High", "Medium", "Low"} else "Low"


def immediate_action_points(result) -> str:
    if result.metrics.get("high", 0):
        return "Clear high-priority findings first, then rerun the review on the updated draft."
    if result.metrics.get("medium", 0):
        return "Review medium-priority drafting, reference, casting, and disclosure findings before final sign-off."
    return "Review low-priority formatting/drafting points and complete professional sign-off procedures."


def cash_flow_summary_note(result) -> str:
    related = [finding.issue for finding in result.findings if review_comment_category_for_finding(finding) == "Cash Flow"]
    if related:
        return "Cash-flow review identified: " + "; ".join(related[:3])
    performed = str(result.metrics.get("checks_performed", ""))
    if "cash flow" in performed.lower():
        return "Automated cash-flow checks were performed where rows were confidently extracted; no elevated cash-flow finding was identified."
    return "Cash-flow correctness requires manual review where extraction did not confidently parse the statement."


def regulatory_summary_note(result) -> str:
    related = [finding.issue for finding in result.findings if review_comment_category_for_finding(finding) == "Regulatory Reference"]
    return "Regulatory-reference review identified: " + "; ".join(related[:3]) if related else "No elevated regulatory-reference finding was identified from extracted evidence."


def casting_summary_note(result) -> str:
    related = [finding.issue for finding in result.findings if review_comment_category_for_finding(finding) in {"Casting", "Cross-casting"}]
    return "Casting/cross-casting review identified: " + "; ".join(related[:3]) if related else "No elevated casting or cross-casting issue was identified from automated checks."



def finding_rows(result) -> list[dict[str, str]]:
    rows = []
    for index, finding in enumerate(result.findings, start=1):
        metadata = finding.metadata or {}
        row = {
            "ID": f"EX-{index:03d}",
            "Status": "Open",
            "Severity": finding.severity,
            "Category": finding.category,
            "Check type": finding.category,
            "Confidence": finding_confidence(finding, result),
            "Page reference": page_reference_for_finding(finding, result),
            "Note reference": note_reference_for_finding(finding, result),
            "Statement": metadata.get("statement", ""),
            "Line item": metadata.get("line_item", ""),
            "Referenced note": metadata.get("referenced_note", ""),
            "Suggested note": metadata.get("suggested_note", ""),
            "Match confidence": metadata.get("match_confidence", finding.severity),
            "Reason": metadata.get("reason", ""),
            "Current year amount found?": metadata.get("current_year_amount_found", ""),
            "Prior year amount found?": metadata.get("prior_year_amount_found", ""),
            "Amount found in note": metadata.get("amount_found_in_note", ""),
            "Alternative note found": metadata.get("alternative_note_found", ""),
            "Amount match confidence": metadata.get("amount_match_confidence", ""),
            "AI review status": metadata.get("ai_review_status", ""),
            "AI review confidence": metadata.get("ai_review_confidence", ""),
            "AI review reason": metadata.get("ai_review_reason", ""),
            "Location": translate_page_tokens(finding.location, result),
            "Issue": finding.issue,
            "Evidence": translate_page_tokens(finding.evidence, result),
            "Recommendation": finding.recommendation,
            "Reviewer comment": "",
            "Prepared by": "",
            "Reviewed by": "",
            "Date cleared": "",
        }
        rows.append(row)
    return rows


def finding_summary_rows_for_result(result) -> list[dict[str, str]]:
    return [
        {
            "Severity": finding.severity,
            "Category": finding.category,
            "Page reference": page_reference_for_finding(finding, result),
            "Note reference": note_reference_for_finding(finding, result),
            "Issue": finding.issue,
            "Recommendation": finding.recommendation,
        }
        for finding in result.findings
    ]


def table_arithmetic_display(result) -> str:
    skipped_details = result.metrics.get("skipped_table_details", [])
    if isinstance(skipped_details, list) and skipped_details:
        return "0% (Skipped)"
    return result.metrics.get("table_arithmetic_confidence", "0%")


def exported_file_stem(result) -> str:
    detected = result.metrics.get("detected_profile", {})
    company_name = ""
    if isinstance(detected, dict):
        company_name = str(detected.get("Company name", "") or "").strip()
    if not company_name:
        company_name = "financial_statement"
    stem = re.sub(r"[^A-Za-z0-9]+", "_", company_name).strip("_").lower()
    return stem[:90] or "financial_statement"


def note_heading_rows(result) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for line in metric_lines(result.metrics.get("note_headings"), "No note headings detected."):
        parts = [part.strip() for part in line.split("|")]
        if len(parts) >= 4:
            rows.append(
                {
                    "Note": parts[0].replace("Note ", "", 1).strip(),
                    "Page": parts[1].replace("Page ", "", 1).strip(),
                    "Page range": parts[2].strip(),
                    "Heading": parts[3].strip(),
                    "Confidence": parts[4].replace("Confidence:", "", 1).strip() if len(parts) >= 5 else "",
                    "Source snippet": parts[5].replace("Source:", "", 1).strip() if len(parts) >= 6 else "",
                }
            )
        elif len(parts) >= 3:
            rows.append(
                {
                    "Note": parts[0].replace("Note ", "", 1).strip(),
                    "Page": parts[1].replace("Page ", "", 1).strip(),
                    "Page range": parts[1].strip(),
                    "Heading": " | ".join(parts[2:]).strip(),
                    "Confidence": "",
                    "Source snippet": "",
                }
            )
    return rows


def notes_heading_candidate_rows_for_result(result) -> list[dict[str, str]]:
    rows = result.metrics.get("notes_heading_candidates", [])
    if isinstance(rows, list) and rows:
        return [row for row in rows if isinstance(row, dict)]
    return [{"Page": "", "Raw OCR snippet": "No possible notes heading candidates detected.", "Normalized snippet": "", "Similarity score": "", "Accepted": "No", "Reason": ""}]


def note_agreement_result_rows(result) -> list[dict[str, str]]:
    rows = result.metrics.get("note_agreement_results", [])
    if isinstance(rows, list):
        return [row for row in rows if isinstance(row, dict)]
    return []


def ocr_statement_row_rows(result) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for line in metric_lines(result.metrics.get("ocr_statement_rows"), "No OCR primary statement rows detected."):
        parts = [part.strip() for part in line.split("|")]
        if len(parts) >= 10:
            rows.append(
                {
                    "Statement": parts[0],
                    "Page": parts[1].replace("Page ", "", 1).strip(),
                    "Line item": parts[2],
                    "Note detected": parts[3],
                    "Current year amount": parts[4],
                    "Prior year amount": parts[5],
                    "Raw OCR line": parts[6],
                    "Parse confidence": parts[7],
                    "Correction applied?": parts[8],
                    "Correction reason": parts[9],
                }
            )
    return rows or [{"Statement": "", "Page": "", "Line item": "No OCR primary statement rows detected.", "Note detected": "", "Current year amount": "", "Prior year amount": "", "Raw OCR line": "", "Parse confidence": "", "Correction applied?": "", "Correction reason": ""}]


def printed_page_map(result) -> dict[int, int]:
    raw_map = result.metrics.get("printed_page_map", {})
    if not isinstance(raw_map, dict):
        return {}
    mapped: dict[int, int] = {}
    for key, value in raw_map.items():
        try:
            mapped[int(key)] = int(value)
        except (TypeError, ValueError):
            continue
    return mapped


def reviewer_page_number(result, page_number: int) -> int:
    return printed_page_map(result).get(page_number, page_number)


def translate_page_tokens(text: object, result) -> str:
    value = str(text or "").strip()
    if not value:
        return ""

    def replace_page_phrase(match: re.Match) -> str:
        prefix = match.group(1)
        body = match.group(2)
        translated: list[str] = []
        for token in re.split(r"(\D+)", body):
            if token.isdigit():
                translated.append(str(reviewer_page_number(result, int(token))))
            else:
                translated.append(token)
        return prefix + "".join(translated)

    return re.sub(r"\b(Pages?\s+)([\d,\-\sand]+)", replace_page_phrase, value, flags=re.I)


def translate_row_page_fields(row: dict[str, object], result, fields: tuple[str, ...]) -> dict[str, object]:
    translated = dict(row)
    for field in fields:
        if field not in translated:
            continue
        value = translated.get(field, "")
        if field.lower() in {"page", "page 1", "page 2"}:
            translated[field] = translate_bare_page_value(value, result)
        else:
            translated[field] = translate_page_tokens(value, result)
    return translated


def translate_bare_page_value(value: object, result) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if not re.fullmatch(r"[0-9,\sand\-]+", text, flags=re.I):
        return translate_page_tokens(text, result)
    translated: list[str] = []
    for token in re.split(r"(\D+)", text):
        if token.isdigit():
            translated.append(str(reviewer_page_number(result, int(token))))
        else:
            translated.append(token)
    return "".join(translated).strip()


def skipped_check_affected_area(item: str, check_area: str) -> str:
    lower = str(item or "").lower()
    note_match = re.search(r"\bNote\s+(\d+[A-Z]?)\b", item, flags=re.I)
    if note_match:
        return f"Note {note_match.group(1).upper()}"
    statement_aliases = (
        ("statement of cash flows", "Statement of cash flows"),
        ("cash flow", "Statement of cash flows"),
        ("statement of financial position", "Statement of financial position"),
        ("financial position", "Statement of financial position"),
        ("profit or loss", "Statement of profit or loss"),
        ("income statement", "Statement of profit or loss"),
        ("income and expenditure", "Statement of income and expenditure"),
        ("changes in equity", "Statement of changes in equity"),
        ("accumulated fund", "Statement of changes in accumulated fund"),
    )
    for marker, label in statement_aliases:
        if marker in lower:
            return label
    if "income-to-equity" in lower or "income or equity" in lower:
        return "Profit or loss and equity movement"
    if "generic table arithmetic" in lower:
        return "Skipped table details"
    if "note" in lower or "table extraction confidence" in lower:
        return "Notes to the financial statements"
    if "limited-scope" in lower or "full afs" in lower:
        return "Upload scope / full financial statements"
    if "pdf extraction" in lower or "extraction" in lower:
        return "Extraction quality"
    return check_area


def skipped_check_manual_priority(item: str, can_fix: str) -> str:
    lower = str(item or "").lower()
    if any(marker in lower for marker in ("cash flow", "financial position", "profit or loss", "income-to-equity", "changes in equity")):
        return "High"
    if any(marker in lower for marker in ("note", "table extraction", "generic table arithmetic", "not confidently parsed")):
        return "Medium"
    if str(can_fix or "").lower().startswith("not applicable"):
        return "Low"
    return "Medium"


def skipped_check_automation_requirement(item: str, can_fix: str) -> str:
    lower = str(item or "").lower()
    if "limited-scope" in lower or "complete afs" in lower or "full afs" in lower:
        return "Upload the complete financial statements, including all primary statements and notes."
    if "notes section start was not detected" in lower:
        return "Reliable notes-section heading detection from extracted text or OCR."
    if "table extraction confidence is below threshold" in lower:
        return "Relevant note tables must reach the detailed-agreement confidence threshold."
    if "generic table arithmetic" in lower:
        return "Confident table classification and numeric row/column structure; supplementary tables remain intentionally excluded."
    if "income-to-equity" in lower or "income or equity" in lower:
        return "Confident extraction of both the income result and the equity/accumulated-fund movement row."
    if any(marker in lower for marker in ("cash flow", "financial position", "profit or loss", "changes in equity", "not confidently parsed")):
        return "Reliable line/table extraction from the affected primary statement page."
    if "pdf extraction is unreliable" in lower or "extraction" in lower:
        return "Readable text/OCR coverage and stable statement row extraction."
    if str(can_fix or "").lower().startswith("not applicable"):
        return "No automation change planned; this skip is intentional for the table type."
    return "Improve extraction confidence or provide clearer table/statement structure."


def parse_skipped_check(item: str) -> dict[str, str]:
    check_area = item
    page_reference = ""
    reason = item
    can_fix = "Partially"
    reviewer_action = "Review the referenced page or supporting detail sheet, then rerun after improving extraction if needed."
    statement_match = re.match(r"(.+?):\s+Page\s+(\d+)\s+skipped because\s+(.+)", item, flags=re.I)
    if statement_match:
        check_area = statement_match.group(1).strip()
        page_reference = f"Page {statement_match.group(2)}"
        reason = statement_match.group(3).rstrip(".")
        reviewer_action = "Open the page and inspect the statement manually; automated casting is withheld because extraction did not produce reliable rows/columns."
    elif "cross-source income-to-equity linkage skipped" in item.lower():
        check_area = "Cross-source income-to-equity linkage"
        page_tokens = re.findall(r"\bPage\s+\d+", item, flags=re.I)
        page_reference = ", ".join(dict.fromkeys(page_tokens))
        reason = "Only one of income statement result or equity movement reference was confidently parsed."
        if "Available evidence:" in item:
            reason += " " + item.split("Available evidence:", 1)[1].strip()
        reviewer_action = "Review the income statement and statement of changes in equity/accumulated fund pages; rerun after improving row extraction if material."
    elif item.lower().startswith("generic table arithmetic skipped"):
        check_area = "Generic table arithmetic"
        page_reference = "See Skipped table details"
        reason = "Low-confidence, note, supplementary, or non-standard tables are listed separately."
        can_fix = "Partially; some table exclusions are intentional"
        reviewer_action = "Use Skipped checks summary and Skipped table details. Value-added statements and multi-year summaries are intentionally excluded; low-confidence and notes tables require manual inspection or better extraction."
    elif "notes section start was not detected" in item.lower():
        check_area = "OCR note-reference validation"
        reason = "Notes section start was not detected reliably."
        can_fix = "Yes, if OCR/notes heading extraction improves"
        reviewer_action = "Review Notes heading candidates and confirm the page where notes begin."
    elif "table extraction confidence is below threshold" in item.lower():
        check_area = "Detailed note agreement"
        reason = "Table extraction confidence is below the safe threshold."
        reviewer_action = "Use Note-linked review and Note agreement results for cautious review prompts; rerun detailed agreement when table extraction improves."
    elif "limited-scope statement extract" in item.lower():
        check_area = "Full AFS completeness and note agreement"
        reason = "The upload appears to be a limited-scope statement extract."
        can_fix = "Yes, with complete AFS upload"
        reviewer_action = "Upload the complete financial statements to run full checklist, policy, and note-agreement checks."
    elif "pdf extraction is unreliable" in item.lower():
        check_area = "Primary statement checks"
        reason = "PDF extraction is unreliable."
        reviewer_action = "Review extraction confidence, OCR statement rows, and source pages before relying on automated checks."
    affected_area = skipped_check_affected_area(item, check_area)
    manual_priority = skipped_check_manual_priority(item, can_fix)
    automation_requirement = skipped_check_automation_requirement(item, can_fix)
    return {
        "Check area": check_area,
        "Affected statement/note": affected_area,
        "Page reference": page_reference,
        "Reason skipped": reason,
        "Requires manual review?": "Optional" if str(can_fix).lower().startswith("not applicable") else "Yes",
        "Manual review priority": manual_priority,
        "Can automated check be fixed?": can_fix,
        "Automation requirement": automation_requirement,
        "Reviewer action": reviewer_action,
        "Original message": item,
    }


def finding_confidence(finding, result) -> str:
    if finding.metadata and finding.metadata.get("ai_review_confidence"):
        return f"AI-reviewed / {finding.metadata['ai_review_confidence']}"
    if finding.metadata and finding.metadata.get("match_confidence"):
        return f"Review prompt / {finding.metadata['match_confidence']}"
    if finding.category == "Extraction quality":
        if finding.location in {"PDF extraction", "Table extraction", "Notes agreement"}:
            return (
                f"OCR text {result.metrics.get('ocr_text_coverage', result.metrics.get('extraction_coverage', '0%'))} / "
                f"Statement {result.metrics.get('statement_structure_confidence', '0%')} / "
                f"Table arithmetic {table_arithmetic_display(result)}"
            )
        return finding.severity
    if finding.category == "Totals and rounding":
        return f"Table arithmetic {table_arithmetic_display(result)}"
    return finding.severity


def page_reference(location: str, evidence: str = "", result=None) -> str:
    text = f"{location}\n{evidence}"
    pages: set[int] = set()
    for match in re.finditer(r"\bpages?\s*:?\s+([0-9,\sand\-]+)", text, flags=re.I):
        for token in re.findall(r"\d+(?:\s*-\s*\d+)?", match.group(1)):
            _add_page_reference_token(pages, token, result)
    for match in re.findall(r"\bPage\s+(\d+)\b", text, flags=re.I):
        _add_page_reference_token(pages, match, result)
    return _format_page_reference_set(pages)


def _add_page_reference_token(pages: set[int], token: str, result=None) -> None:
    numbers = [int(number) for number in re.findall(r"\d+", token)]
    if not numbers:
        return
    if len(numbers) >= 2 and "-" in token:
        start, end = numbers[0], numbers[1]
        if start <= end and end - start <= 50:
            for page_number in range(start, end + 1):
                pages.add(reviewer_page_number(result, page_number) if result is not None else page_number)
            return
    page_number = numbers[0]
    pages.add(reviewer_page_number(result, page_number) if result is not None else page_number)


def _format_page_reference_set(pages: set[int]) -> str:
    if not pages:
        return ""
    ordered = sorted(pages)
    ranges: list[str] = []
    start = previous = ordered[0]
    for page in ordered[1:]:
        if page == previous + 1:
            previous = page
            continue
        ranges.append(str(start) if start == previous else f"{start}-{previous}")
        start = previous = page
    ranges.append(str(start) if start == previous else f"{start}-{previous}")
    return f"Page {ranges[0]}" if len(ordered) == 1 else "Pages " + ", ".join(ranges)


def page_reference_for_finding(finding, result) -> str:
    direct = page_reference(finding.location, finding.evidence, result)
    metadata = finding.metadata or {}
    if finding.category == "Notes agreement":
        pages: set[int] = set(int(match) for match in re.findall(r"\bPage\s+(\d+)\b", direct or ""))
        note_pages = note_page_reference_map(result)
        for key in ("referenced_note", "suggested_note", "alternative_note_found"):
            note_ref = str(metadata.get(key, "") or "").upper().strip()
            if not note_ref:
                continue
            page_text = note_pages.get(note_ref) or note_pages.get(re.sub(r"[A-Z]+$", "", note_ref))
            if page_text:
                pages.update(int(match) for match in re.findall(r"\d+", page_text))
        if pages:
            ordered = sorted(pages)
            return f"Page {ordered[0]}" if len(ordered) == 1 else "Pages " + ", ".join(str(page) for page in ordered)
    if direct:
        return direct
    return str(finding.location or "").strip() or "Document-wide"


def note_reference_for_finding(finding, result) -> str:
    metadata = finding.metadata or {}
    note_candidates: list[str] = []
    for key in ("referenced_note", "suggested_note", "alternative_note_found", "amount_found_in_note"):
        value = str(metadata.get(key, "") or "").strip().upper()
        if value:
            note_candidates.append(value if value.startswith("NOTE ") else f"Note {value}")
    if note_candidates:
        return ", ".join(dict.fromkeys(note_candidates))
    text = "\n".join(str(part or "") for part in (finding.location, finding.issue, finding.evidence, metadata.get("reason", "")))
    matches = re.findall(r"\bNote\s+(\d+[A-Z]?)\b", text, flags=re.I)
    if matches:
        return ", ".join(f"Note {match.upper()}" for match in dict.fromkeys(matches))
    return ""


def note_page_reference_map(result) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for row in note_heading_rows(result):
        note = str(row.get("Note", "")).upper().strip()
        page_range = str(row.get("Page range", "") or row.get("Page", "")).strip()
        if note and page_range:
            mapping[note] = page_range
    for row in note_agreement_result_rows(result):
        note = str(row.get("Note number", "")).upper().strip()
        page_range = str(row.get("Note section page range", "")).strip()
        if note and page_range:
            mapping[note] = page_range
    return mapping


def format_excel_table_sheet(worksheet, table_name: str) -> None:
    max_row = max(worksheet.max_row, 2)
    max_col = worksheet.max_column
    table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    worksheet.add_table(table)


def format_exception_register_sheet(worksheet) -> None:
    max_row = max(worksheet.max_row, 2)
    max_col = worksheet.max_column
    table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName="ExceptionRegister", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    worksheet.add_table(table)
    headers = {str(cell.value): cell.column for cell in worksheet[1]}
    status_col = headers.get("Status")
    severity_col = headers.get("Severity")
    if status_col:
        status_letter = get_column_letter(status_col)
        validation = DataValidation(type="list", formula1='"Open,Cleared,False Positive,Noted"', allow_blank=True)
        worksheet.add_data_validation(validation)
        validation.add(f"{status_letter}2:{status_letter}500")
    if severity_col:
        severity_letter = get_column_letter(severity_col)
        color_map = {"High": ("FEE2E2", "991B1B"), "Medium": ("FEF3C7", "92400E"), "Low": ("DBEAFE", "1E40AF")}
        for severity, (fill, font) in color_map.items():
            worksheet.conditional_formatting.add(
                f"{severity_letter}2:{severity_letter}{max_row}",
                FormulaRule(formula=[f'${severity_letter}2="{severity}"'], fill=PatternFill("solid", fgColor=fill), font=Font(color=font, bold=True)),
            )
    for header in ("Evidence", "Recommendation", "Issue", "Reviewer comment"):
        column = headers.get(header)
        if column:
            letter = get_column_letter(column)
            worksheet.column_dimensions[letter].width = 55 if header in {"Evidence", "Recommendation", "Issue"} else 32
            for row in range(2, max_row + 1):
                worksheet[f"{letter}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.auto_filter.ref = table_ref

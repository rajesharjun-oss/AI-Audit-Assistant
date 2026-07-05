from __future__ import annotations

import re
from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from canonical_models import ReconciliationCheckResult
from models import Finding

REVIEW_HEADERS = [
    "S/N",
    "Section / Statement / Note",
    "Page number",
    "Account / line item",
    "Current wording / amount / reference",
    "Issue identified",
    "Expected correction / recommendation",
    "Category",
    "Priority",
    "Status",
    "Reviewer comments",
    "Reported amount",
    "Expected amount",
    "Difference",
    "Formula",
    "Confidence",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HIGH_FILL = PatternFill("solid", fgColor="FFC7CE")
MEDIUM_FILL = PatternFill("solid", fgColor="FFEB9C")
LOW_FILL = PatternFill("solid", fgColor="E2F0D9")


def write_canonical_review_workbook(
    *,
    findings: list[Finding],
    check_results: list[ReconciliationCheckResult],
    audit_rows: list[dict[str, object]],
    output_path: str | Path,
    template_path: str | Path | None = None,
) -> Path:
    """Append canonical QC findings to an existing workbook or create a new one."""
    output = Path(output_path)
    if template_path and Path(template_path).exists():
        wb = load_workbook(template_path)
    else:
        wb = Workbook()
        wb.active.title = "Review Comments"
    review_ws = _review_sheet(wb)
    _ensure_review_headers(review_ws)
    _append_findings(review_ws, findings)
    _replace_sheet(wb, "Summary")
    _write_summary(wb["Summary"], findings, check_results)
    _replace_sheet(wb, "Recalculation Checks")
    _write_rows(wb["Recalculation Checks"], [result.to_row() for result in check_results])
    _replace_sheet(wb, "Extraction Audit")
    _write_rows(wb["Extraction Audit"], audit_rows)
    _format_workbook(wb)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def _review_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().lower() in {"review comments", "review comment", "sheet1", "comments", "exception register"}:
            return wb[name]
    return wb[wb.sheetnames[0]]


def _ensure_review_headers(ws) -> None:
    existing = [str(cell.value or "").strip() for cell in ws[1]]
    if not any(existing):
        for col, header in enumerate(REVIEW_HEADERS, start=1):
            ws.cell(1, col, header)
        return
    lower_existing = {header.lower(): idx for idx, header in enumerate(existing, start=1) if header}
    next_col = len(existing) + 1
    for header in REVIEW_HEADERS:
        if header.lower() not in lower_existing:
            ws.cell(1, next_col, header)
            next_col += 1


def _append_findings(ws, findings: list[Finding]) -> None:
    header_map = {str(cell.value or "").strip().lower(): idx for idx, cell in enumerate(ws[1], start=1)}
    next_row = ws.max_row + 1 if ws.max_row >= 1 else 2
    existing_serials = [ws.cell(row, header_map.get("s/n", 1)).value for row in range(2, ws.max_row + 1)] if ws.max_row >= 2 else []
    serial = _next_serial(existing_serials)
    for finding in findings:
        metadata = finding.metadata or {}
        values = {
            "s/n": serial,
            "section / statement / note": metadata.get("statement", finding.location),
            "page number": _page_number_from_location(finding.location),
            "account / line item": metadata.get("line_item", ""),
            "current wording / amount / reference": finding.evidence,
            "issue identified": finding.issue,
            "expected correction / recommendation": finding.recommendation,
            "category": _normalise_category(finding.category),
            "priority": _normalise_priority(finding.severity),
            "status": "Open",
            "reviewer comments": "",
            "reported amount": metadata.get("reported_amount", ""),
            "expected amount": metadata.get("expected_amount", ""),
            "difference": metadata.get("difference", ""),
            "formula": metadata.get("formula", ""),
            "confidence": metadata.get("match_confidence", metadata.get("confidence", "")),
        }
        for key, value in values.items():
            col = header_map.get(key)
            if col:
                ws.cell(next_row, col, value)
        next_row += 1
        serial += 1


def _next_serial(values: Iterable[object]) -> int:
    numbers: list[int] = []
    for value in values:
        try:
            numbers.append(int(value))
        except (TypeError, ValueError):
            continue
    return max(numbers, default=0) + 1


def _page_number_from_location(location: str) -> str:
    pages = re.findall(r"\bPage[s]?\s+([0-9, -]+)", str(location or ""), flags=re.I)
    return pages[0] if pages else ""


def _normalise_priority(value: str) -> str:
    lower = str(value or "").strip().lower()
    if lower.startswith("high"):
        return "High"
    if lower.startswith("low") or lower == "passed":
        return "Low"
    return "Medium"


def _normalise_category(value: str) -> str:
    mapping = {
        "totals and rounding": "Casting",
        "notes agreement": "Note Cross-reference",
        "consistency": "Internal Consistency",
        "narrative consistency": "Internal Consistency",
    }
    lower = str(value or "").strip().lower()
    return mapping.get(lower, str(value or "").strip() or "Internal Consistency")


def _replace_sheet(wb, name: str) -> None:
    if name in wb.sheetnames:
        del wb[name]
    wb.create_sheet(name)


def _write_summary(ws, findings: list[Finding], check_results: list[ReconciliationCheckResult]) -> None:
    priorities = Counter(_normalise_priority(finding.severity) for finding in findings)
    categories = Counter(_normalise_category(finding.category) for finding in findings)
    failed_checks = [result for result in check_results if result.status == "Fail"]
    passed_checks = [result for result in check_results if result.status == "Pass"]
    summary_rows: list[tuple[str, object]] = [
        ("Total findings", len(findings)),
        ("High-priority findings", priorities.get("High", 0)),
        ("Medium-priority findings", priorities.get("Medium", 0)),
        ("Low-priority findings", priorities.get("Low", 0)),
        ("Deterministic checks passed", len(passed_checks)),
        ("Deterministic checks failed", len(failed_checks)),
        ("Overall sign-off readiness", "Not ready for final sign-off" if priorities.get("High", 0) or failed_checks else "Manual review required before sign-off"),
        ("Cash flow conclusion", _cash_flow_conclusion(check_results)),
        ("Casting and cross-casting conclusion", _casting_conclusion(check_results)),
        ("Regulatory-reference conclusion", "Run the existing AI/regulatory review layer or the firm regulatory checklist for legal-reference wording; this sheet focuses on canonical deterministic checks."),
    ]
    row = 1
    ws.cell(row, 1, "QC Review Summary")
    ws.cell(row, 1).font = Font(bold=True, size=14)
    row += 2
    for metric, value in summary_rows:
        ws.cell(row, 1, metric)
        ws.cell(row, 2, value)
        row += 1
    row += 1
    ws.cell(row, 1, "Findings by category")
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    for category, count in sorted(categories.items()):
        ws.cell(row, 1, category)
        ws.cell(row, 2, count)
        row += 1
    row += 1
    ws.cell(row, 1, "Key high-priority findings")
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    high_findings = [finding for finding in findings if _normalise_priority(finding.severity) == "High"][:10]
    if not high_findings:
        ws.cell(row, 1, "No high-priority canonical findings generated.")
    else:
        for finding in high_findings:
            ws.cell(row, 1, finding.location)
            ws.cell(row, 2, finding.issue)
            ws.cell(row, 3, finding.recommendation)
            row += 1


def _cash_flow_conclusion(check_results: list[ReconciliationCheckResult]) -> str:
    cash_flow = [result for result in check_results if result.category == "Cash Flow"]
    failed = [result for result in cash_flow if result.status == "Fail"]
    if not cash_flow:
        return "Cash-flow checks were not tested because required cash-flow facts were not parsed."
    if failed:
        return f"Cash-flow review identified {len(failed)} failed check(s); inspect Recalculation Checks."
    return "Cash-flow arithmetic checks passed for parsed facts."


def _casting_conclusion(check_results: list[ReconciliationCheckResult]) -> str:
    relevant = [result for result in check_results if result.category in {"Casting", "Cross-casting"}]
    failed = [result for result in relevant if result.status == "Fail"]
    if not relevant:
        return "Casting/cross-casting checks were not tested because required facts were not parsed."
    if failed:
        return f"Casting/cross-casting review identified {len(failed)} failed check(s); inspect Recalculation Checks."
    return "Casting/cross-casting checks passed for parsed facts."


def _write_rows(ws, rows: list[dict[str, object]]) -> None:
    if not rows:
        ws.cell(1, 1, "No rows generated")
        return
    headers = list(dict.fromkeys(header for row in rows for header in row.keys()))
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)
    for row_index, row in enumerate(rows, start=2):
        for col, header in enumerate(headers, start=1):
            ws.cell(row_index, col, row.get(header, ""))


def _format_workbook(wb) -> None:
    for ws in wb.worksheets:
        max_col = max(ws.max_column, 1)
        max_row = max(ws.max_row, 1)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
        for row in ws.iter_rows(min_row=2, max_row=max_row):
            priority = str(row[8].value if len(row) >= 9 else "").lower()
            fill = HIGH_FILL if priority == "high" else MEDIUM_FILL if priority == "medium" else LOW_FILL if priority == "low" else None
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if fill and ws.title.lower().startswith("review"):
                    cell.fill = fill
        for col in range(1, max_col + 1):
            header = str(ws.cell(1, col).value or "")
            width = min(max(12, len(header) + 2), 45)
            if header.lower() in {"issue identified", "expected correction / recommendation", "current wording / amount / reference", "source rows", "raw row"}:
                width = 60
            ws.column_dimensions[get_column_letter(col)].width = width
        if max_row >= 2 and max_col >= 2:
            ref = f"A1:{get_column_letter(max_col)}{max_row}"
            try:
                table = Table(displayName=_table_name(ws.title), ref=ref)
                table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                ws.add_table(table)
            except ValueError:
                pass


def _table_name(title: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", title).strip("_") or "Table"
    return (cleaned[:24] + "Tbl") if len(cleaned) < 28 else cleaned[:28]

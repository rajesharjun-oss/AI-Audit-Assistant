from __future__ import annotations

from io import BytesIO
import hashlib
import os
import re
import tempfile
from pathlib import Path
from xml.sax.saxutils import escape as xml_escape
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
import streamlit as st
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, PatternFill

from export_utils import clean_name_consistency_rows
from models import CompanyProfile, DEFAULT_AI_MODEL, ReviewOptions
from report_exports import build_excel_export, exported_file_stem
from reviewer import build_ai_review_memo, extract_review_document, findings_to_markdown, normalize_reporting_currency, rerun_ai_review_from_cached_result, review_document

st.set_page_config(page_title="AI Audit Assistant", page_icon="AI", layout="wide", initial_sidebar_state="expanded")

def _friendly_review_failure_message(exc: Exception) -> str:
    text = str(exc or "").strip()
    lower = text.lower()
    if any(marker in lower for marker in ("429", "rate limit", "rate exceeded", "too many requests", "ai service busy", "timed out", "timeout")):
        return (
            "The AI service or review worker is temporarily busy, so this upload could not finish cleanly. "
            "No exception register was produced for this run. Please wait about 20 seconds and try again. "
            "If it happens again, temporarily turn off AI policy judgement to generate the deterministic review immediately."
        )
    return f"The review could not be completed: {text or type(exc).__name__}."

def _metric_lines(value: object, empty: str) -> list[str]:
    text = str(value or "").strip()
    if not text or text == empty:
        return []
    return [line.strip() for line in text.splitlines() if line.strip()]


def _checks_skipped_rows(result) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for item in _metric_lines(result.metrics.get("checks_skipped"), "No major checks skipped."):
        row = _parse_skipped_check(item)
        rows.append(_translate_row_page_fields(row, result, ("Page reference", "Original message", "Reason skipped")))
    return rows


def _printed_page_map(result) -> dict[int, int]:
    raw_map = result.metrics.get("printed_page_map", {})
    if not isinstance(raw_map, dict):
        return {}
    mapped: dict[int, int] = {}
    for key, value in raw_map.items():
        try:
            mapped[int(key)] = int(value)
        except (TypeError, ValueError):
            continue
    return mapped


def _reviewer_page_number(result, page_number: int) -> int:
    return _printed_page_map(result).get(page_number, page_number)


def _translate_page_tokens(text: object, result) -> str:
    value = str(text or "").strip()
    if not value:
        return ""

    def replace_page_phrase(match: re.Match) -> str:
        prefix = match.group(1)
        body = match.group(2)
        translated: list[str] = []
        for token in re.split(r"(\D+)", body):
            if token.isdigit():
                translated.append(str(_reviewer_page_number(result, int(token))))
            else:
                translated.append(token)
        return prefix + "".join(translated)

    value = re.sub(r"\b(Pages?\s+)([\d,\-\sand]+)", replace_page_phrase, value, flags=re.I)
    return value


def _translate_row_page_fields(row: dict[str, object], result, fields: tuple[str, ...]) -> dict[str, object]:
    translated = dict(row)
    for field in fields:
        if field in translated:
            translated[field] = _translate_page_tokens(translated.get(field, ""), result)
    return translated


def _parse_skipped_check(item: str) -> dict[str, str]:
    check_area = item
    page_reference = ""
    reason = item
    can_fix = "Partially"
    reviewer_action = "Review the referenced page or supporting detail sheet, then rerun after improving extraction if needed."

    statement_match = re.match(r"(.+?):\s+Page\s+(\d+)\s+skipped because\s+(.+)", item, flags=re.I)
    if statement_match:
        check_area = statement_match.group(1).strip()
        page_reference = f"Page {statement_match.group(2)}"
        reason = statement_match.group(3).rstrip(".")
        can_fix = "Partially"
        reviewer_action = "Open the page and inspect the statement manually; automated casting is withheld because extraction did not produce reliable rows/columns."
    elif item.lower().startswith("generic table arithmetic skipped"):
        check_area = "Generic table arithmetic"
        page_reference = "See Skipped table details"
        reason = "Low-confidence or non-standard tables are listed separately."
        can_fix = "Partially"
        reviewer_action = "Use Skipped checks summary and Skipped table details to inspect the affected note tables manually."
    elif "notes section start was not detected" in item.lower():
        check_area = "OCR note-reference validation"
        reason = "Notes section start was not detected reliably."
        can_fix = "Yes, if OCR/notes heading extraction improves"
        reviewer_action = "Review Notes heading candidates and confirm the page where notes begin."
    elif "table extraction confidence is below threshold" in item.lower():
        check_area = "Detailed note agreement"
        reason = "Table extraction confidence is below the safe threshold."
        can_fix = "Partially"
        reviewer_action = "Use Note-linked review and Note agreement results for cautious review prompts; rerun detailed agreement when table extraction improves."
    elif "limited-scope statement extract" in item.lower():
        check_area = "Full AFS completeness and note agreement"
        reason = "The upload appears to be a limited-scope statement extract."
        can_fix = "Yes, with complete AFS upload"
        reviewer_action = "Upload the complete financial statements to run full checklist, policy, and note-agreement checks."
    elif "pdf extraction is unreliable" in item.lower():
        check_area = "Primary statement checks"
        reason = "PDF extraction is unreliable."
        can_fix = "Partially"
        reviewer_action = "Review extraction confidence, OCR statement rows, and source pages before relying on automated checks."

    return {
        "Check area": check_area,
        "Page reference": page_reference,
        "Reason skipped": reason,
        "Can automated check be fixed?": can_fix,
        "Reviewer action": reviewer_action,
        "Original message": item,
    }


def _finding_rows(result) -> list[dict[str, str]]:
    rows = []
    for index, finding in enumerate(result.findings, start=1):
        metadata = finding.metadata or {}
        if finding.category == "Notes agreement" and metadata.get("match_confidence") == "Low":
            pass  # We now filter this in reviewer.py before the result is built.
        row = {
            "ID": f"EX-{index:03d}",
            "Status": "Open",
            "Severity": finding.severity,
            "Category": finding.category,
            "Check type": finding.category,
            "Confidence": _finding_confidence(finding, result),
            "Page reference": _page_reference_for_finding(finding, result),
            "Note reference": _note_reference_for_finding(finding, result),
            "Statement": metadata.get("statement", ""),
            "Line item": metadata.get("line_item", ""),
            "Referenced note": metadata.get("referenced_note", ""),
            "Suggested note": metadata.get("suggested_note", ""),
            "Match confidence": metadata.get("match_confidence", finding.severity),
            "Reason": metadata.get("reason", ""),
            "Current year amount found?": metadata.get("current_year_amount_found", ""),
            "Prior year amount found?": metadata.get("prior_year_amount_found", ""),
            "Amount found in note": metadata.get("amount_found_in_note", ""),
            "Alternative note found": metadata.get("alternative_note_found", ""),
            "Amount match confidence": metadata.get("amount_match_confidence", ""),
            "Location": _translate_page_tokens(finding.location, result),
            "Issue": finding.issue,
            "Evidence": _translate_page_tokens(finding.evidence, result),
            "Recommendation": finding.recommendation,
            "Reviewer comment": "",
            "Prepared by": "",
            "Reviewed by": "",
            "Date cleared": "",
            }
        rows.append(row)
    return rows


def _finding_summary_rows(result) -> list[dict[str, str]]:
    rows = []
    for finding in result.findings:
        rows.append(
            {
                "Severity": finding.severity,
                "Category": finding.category,
                "Page reference": _page_reference_for_finding(finding, result),
                "Note reference": _note_reference_for_finding(finding, result),
                "Issue": finding.issue,
                "Recommendation": finding.recommendation,
            }
        )
    return rows


def _finding_confidence(finding, result) -> str:
    if finding.metadata and finding.metadata.get("match_confidence"):
        return f"Review prompt / {finding.metadata['match_confidence']}"
    if finding.category == "Extraction quality":
        if finding.location in {"PDF extraction", "Table extraction", "Notes agreement"}:
            return (
                f"OCR text {result.metrics.get('ocr_text_coverage', result.metrics.get('extraction_coverage', '0%'))} / "
                f"Statement {result.metrics.get('statement_structure_confidence', '0%')} / "
                f"Table arithmetic {_table_arithmetic_display(result)}"
            )
        return finding.severity
    if finding.category == "Totals and rounding":
        return f"Table arithmetic {_table_arithmetic_display(result)}"
    if finding.severity == "High":
        return "High"
    if finding.severity == "Medium":
        return "Medium"
    return "Low"


def _table_arithmetic_display(result) -> str:
    skipped_details = result.metrics.get("skipped_table_details", [])
    if isinstance(skipped_details, list) and skipped_details:
        return "0% (Skipped)"
    return result.metrics.get("table_arithmetic_confidence", "0%")


def _page_reference(location: str, evidence: str = "", result=None) -> str:
    text = f"{location}\n{evidence}"
    pages = set()
    for match in re.finditer(r"\bpages?\s*:?\s+([0-9,\sand]+)", text, flags=re.I):
        for number in re.findall(r"\d+", match.group(1)):
            page_number = int(number)
            pages.add(_reviewer_page_number(result, page_number) if result is not None else page_number)
    for match in re.findall(r"\bPage\s+(\d+)\b", text, flags=re.I):
        page_number = int(match)
        pages.add(_reviewer_page_number(result, page_number) if result is not None else page_number)
    pages = sorted(pages)
    if not pages:
        return ""
    if len(pages) == 1:
        return f"Page {pages[0]}"
    return "Pages " + ", ".join(str(page) for page in pages)


def _page_reference_for_finding(finding, result) -> str:
    direct = _page_reference(finding.location, finding.evidence, result)
    metadata = finding.metadata or {}
    if finding.category == "Notes agreement":
        pages: set[int] = set(int(match) for match in re.findall(r"\bPage\s+(\d+)\b", direct or ""))
        note_pages = _note_page_reference_map(result)
        for key in ("referenced_note", "suggested_note", "alternative_note_found"):
            note_ref = str(metadata.get(key, "") or "").upper().strip()
            if not note_ref:
                continue
            page_text = note_pages.get(note_ref) or note_pages.get(re.sub(r"[A-Z]+$", "", note_ref))
            if page_text:
                pages.update(int(match) for match in re.findall(r"\d+", page_text))
        if pages:
            ordered = sorted(pages)
            return f"Page {ordered[0]}" if len(ordered) == 1 else "Pages " + ", ".join(str(page) for page in ordered)
    if direct:
        return direct
    note_match = re.search(r"\bNote\s+(\d+[A-Z]?)\b", f"{finding.location}\n{finding.issue}\n{finding.evidence}", flags=re.I)
    if note_match:
        note_pages = _note_page_reference_map(result)
        note_ref = note_match.group(1).upper()
        page_text = note_pages.get(note_ref) or note_pages.get(re.sub(r"[A-Z]+$", "", note_ref))
        if page_text:
            return page_text
    location = str(finding.location or "").strip()
    if location:
        return location
    return "Document-wide"


def _note_reference_for_finding(finding, result) -> str:
    metadata = finding.metadata or {}
    note_candidates: list[str] = []
    for key in ("referenced_note", "suggested_note", "alternative_note_found", "amount_found_in_note"):
        value = str(metadata.get(key, "") or "").strip().upper()
        if value:
            if not value.startswith("NOTE "):
                value = f"Note {value}"
            note_candidates.append(value)
    if note_candidates:
        deduped = list(dict.fromkeys(note_candidates))
        return ", ".join(deduped)

    text = "\n".join(
        str(part or "")
        for part in (
            finding.location,
            finding.issue,
            finding.evidence,
            metadata.get("reason", ""),
        )
    )
    matches = re.findall(r"\bNote\s+(\d+[A-Z]?)\b", text, flags=re.I)
    if matches:
        return ", ".join(f"Note {match.upper()}" for match in dict.fromkeys(matches))

    page_ref = _page_reference_for_finding(finding, result)
    if finding.category in {"Notes agreement", "Totals and rounding"} and "Page " in page_ref:
        note_pages = _note_page_reference_map(result)
        pages = set(re.findall(r"\d+", page_ref))
        matched_notes = []
        for note, note_page_range in note_pages.items():
            note_pages_in_range = set(re.findall(r"\d+", note_page_range))
            if pages & note_pages_in_range:
                matched_notes.append(f"Note {note}")
        if matched_notes:
            return ", ".join(dict.fromkeys(matched_notes))
    return ""


def _note_page_reference_map(result) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for row in _note_heading_rows(result):
        note = str(row.get("Note", "")).upper().strip()
        page_range = str(row.get("Page range", "") or row.get("Page", "")).strip()
        if note and page_range:
            mapping[note] = page_range
    for row in _note_agreement_result_rows(result):
        note = str(row.get("Note number", "")).upper().strip()
        page_range = str(row.get("Note section page range", "")).strip()
        if note and page_range:
            mapping[note] = page_range
    return mapping


def _export_file_stem(result) -> str:
    detected = result.metrics.get("detected_profile", {})
    company_name = ""
    if isinstance(detected, dict):
        company_name = str(detected.get("Company name", "") or "").strip()
    if not company_name:
        company_name = "financial_statement"
    stem = re.sub(r"[^A-Za-z0-9]+", "_", company_name).strip("_").lower()
    return stem[:90] or "financial_statement"


def _note_heading_rows(result) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for line in _metric_lines(result.metrics.get("note_headings"), "No note headings detected."):
        parts = [part.strip() for part in line.split("|")]
        if len(parts) >= 4:
            rows.append(
                {
                    "Note": parts[0].replace("Note ", "", 1).strip(),
                    "Page": parts[1].replace("Page ", "", 1).strip(),
                    "Page range": parts[2].strip(),
                    "Heading": parts[3].strip(),
                    "Confidence": parts[4].replace("Confidence:", "", 1).strip() if len(parts) >= 5 else "",
                    "Source snippet": parts[5].replace("Source:", "", 1).strip() if len(parts) >= 6 else "",
                }
            )
        elif len(parts) >= 3:
            rows.append(
                {
                    "Note": parts[0].replace("Note ", "", 1).strip(),
                    "Page": parts[1].replace("Page ", "", 1).strip(),
                    "Page range": parts[1].strip(),
                    "Heading": " | ".join(parts[2:]).strip(),
                    "Confidence": "",
                    "Source snippet": "",
                }
            )
    return rows


def _notes_heading_candidate_rows(result) -> list[dict[str, str]]:
    rows = result.metrics.get("notes_heading_candidates", [])
    if isinstance(rows, list) and rows:
        return [row for row in rows if isinstance(row, dict)]
    return [
        {
            "Page": "",
            "Raw OCR snippet": "No possible notes heading candidates detected.",
            "Normalized snippet": "",
            "Similarity score": "",
            "Accepted": "No",
            "Reason": "",
        }
    ]


def _note_agreement_result_rows(result) -> list[dict[str, str]]:
    rows = result.metrics.get("note_agreement_results", [])
    if isinstance(rows, list):
        return [row for row in rows if isinstance(row, dict)]
    return []


def _ocr_statement_row_rows(result) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for line in _metric_lines(result.metrics.get("ocr_statement_rows"), "No OCR primary statement rows detected."):
        parts = [part.strip() for part in line.split("|")]
        if len(parts) >= 10:
            rows.append(
                {
                    "Statement": parts[0],
                    "Page": parts[1].replace("Page ", "", 1).strip(),
                    "Line item": parts[2],
                    "Note detected": parts[3],
                    "Current year amount": parts[4],
                    "Prior year amount": parts[5],
                    "Raw OCR line": parts[6],
                    "Parse confidence": parts[7],
                    "Correction applied?": parts[8],
                    "Correction reason": parts[9],
                }
            )
        elif len(parts) >= 4:
            has_confidence = len(parts) >= 7
            rows.append(
                {
                    "Statement": parts[0],
                    "Page": parts[1].replace("Page ", "", 1).strip(),
                    "Line item": parts[2],
                    "Note detected": "",
                    "Current year amount": "",
                    "Prior year amount": "",
                    "Amounts": " | ".join(parts[3:-3]) if has_confidence else (" | ".join(parts[3:-1]) if len(parts) >= 5 else parts[3]),
                    "Raw OCR line": parts[-3] if has_confidence else (parts[-1] if len(parts) >= 5 else ""),
                    "Parse confidence": parts[-2] if has_confidence else "",
                    "Correction applied?": "",
                    "Correction reason": parts[-1] if has_confidence else "",
                }
            )
    return rows or [{"Statement": "", "Page": "", "Line item": "No OCR primary statement rows detected.", "Note detected": "", "Current year amount": "", "Prior year amount": "", "Raw OCR line": "", "Parse confidence": "", "Correction applied?": "", "Correction reason": ""}]


def _build_excel_export(result) -> bytes:
    output = BytesIO()
    summary_rows = [
        {"Metric": "Checks performed", "Value": result.metrics.get("checks_performed_count", 0)},
        {"Metric": "Checks passed", "Value": result.metrics.get("checks_passed_count", 0)},
        {"Metric": "Checks skipped", "Value": result.metrics.get("checks_skipped_count", 0)},
        {"Metric": "Findings", "Value": result.metrics.get("findings", 0)},
        {"Metric": "High findings", "Value": result.metrics.get("high", 0)},
        {"Metric": "Medium findings", "Value": result.metrics.get("medium", 0)},
        {"Metric": "Low findings", "Value": result.metrics.get("low", 0)},
        {"Metric": "OCR text coverage", "Value": result.metrics.get("ocr_text_coverage", result.metrics.get("extraction_coverage", "0%"))},
        {"Metric": "OCR table candidates", "Value": result.metrics.get("ocr_tables", 0)},
        {"Metric": "Statement structure confidence", "Value": result.metrics.get("statement_structure_confidence", "0%")},
        {"Metric": "Note structure confidence", "Value": result.metrics.get("note_structure_confidence", "0%")},
        {"Metric": "Table arithmetic confidence", "Value": _table_arithmetic_display(result)},
        {"Metric": "notes_section_start_page", "Value": result.metrics.get("notes_section_start_page", "Not detected")},
        {"Metric": "notes_heading_snippet", "Value": result.metrics.get("notes_heading_snippet", "No reliable notes heading detected.")},
        {"Metric": "cautious_note_validation_enabled", "Value": result.metrics.get("cautious_note_validation_enabled", False)},
        {"Metric": "note_validation_mode", "Value": result.metrics.get("note_validation_mode", "skipped")},
        {"Metric": "note_reference_rows_detected", "Value": result.metrics.get("note_reference_rows_detected", 0)},
        {"Metric": "note_headings_detected", "Value": result.metrics.get("note_headings_detected", 0)},
        {"Metric": "note_reference_findings", "Value": result.metrics.get("note_reference_findings", 0)},
        {"Metric": "AI policy review status", "Value": result.metrics.get("ai_policy_review_status", "disabled")},
        {"Metric": "AI policy review message", "Value": result.metrics.get("ai_policy_review_message", "")},
        {"Metric": "AI policy review summary", "Value": result.metrics.get("ai_policy_review_summary", "")},
    ]
    checks_performed = [{"Check performed": item} for item in _metric_lines(result.metrics.get("checks_performed"), "No deterministic checks completed.")]
    checks_skipped = _checks_skipped_rows(result) or [{"Check area": "None", "Reason skipped": "No major checks skipped."}]
    check_results = result.metrics.get("check_results", [])
    if not isinstance(check_results, list) or not check_results:
        check_results = [{"Check": "No deterministic checks completed.", "Result": "Skipped", "Severity": "", "Evidence": ""}]
    detected_profile = result.metrics.get("detected_profile", {})
    profile_rows = [{"Field": key, "Detected value": value} for key, value in detected_profile.items()] if isinstance(detected_profile, dict) else []
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(profile_rows).to_excel(writer, sheet_name="Detected profile", index=False)
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)
        finding_summary_rows = _finding_summary_rows(result) or [{"Severity": "", "Category": "", "Page reference": "", "Issue": "No automated findings were identified.", "Recommendation": ""}]
        pd.DataFrame(finding_summary_rows).to_excel(writer, sheet_name="Findings summary", index=False)
        ai_status = str(result.metrics.get("ai_policy_review_status", "disabled") or "disabled")
        ai_message = str(result.metrics.get("ai_policy_review_message", "") or "").strip()
        ai_default_title = {
            "disabled": "AI policy review not enabled.",
            "unavailable": "AI policy review was enabled but is unavailable in this environment.",
            "skipped": "AI policy review was enabled but no suitable policy/disclosure context was detected.",
            "error": "AI policy review was enabled but failed during execution.",
            "completed": "AI policy review completed but returned no observation rows.",
        }.get(ai_status, "AI policy review returned no rows.")
        ai_policy_rows = result.metrics.get("ai_policy_export", []) or [{"Title": ai_default_title, "Status": ai_status, "Message": ai_message}]
        pd.DataFrame(ai_policy_rows).to_excel(writer, sheet_name="AI policy judgement", index=False)
        ai_finding_status = str(result.metrics.get("ai_finding_review_status", "disabled") or "disabled")
        ai_finding_message = str(result.metrics.get("ai_finding_review_message", "") or "").strip()
        ai_finding_default_title = {
            "disabled": "AI finding review not enabled.",
            "unavailable": "AI finding review was enabled but is unavailable in this environment.",
            "skipped": "AI finding review was enabled but no weak deterministic findings were eligible.",
            "error": "AI finding review was enabled but failed during execution.",
            "deferred": "AI finding review was deferred due to API availability or rate limiting.",
            "completed": "AI finding review completed but returned no adjudication rows.",
        }.get(ai_finding_status, "AI finding review returned no rows.")
        ai_finding_rows = result.metrics.get("ai_finding_export", []) or [
            {"Finding ID": "", "Issue": ai_finding_default_title, "AI status": ai_finding_status, "Reason": ai_finding_message}
        ]
        pd.DataFrame(ai_finding_rows).to_excel(writer, sheet_name="AI finding review", index=False)
        ai_evidence_rows = result.metrics.get("ai_evidence_packs", []) or [
            {"Evidence type": "None", "AI role": "AI review was not run or no evidence packs were eligible."}
        ]
        pd.DataFrame(ai_evidence_rows).to_excel(writer, sheet_name="AI evidence packs", index=False)
        exception_rows = _finding_rows(result) or [
            {
                "ID": "",
                "Status": "Noted",
                "Severity": "",
                "Category": "",
                "Check type": "",
                "Confidence": "",
                "Page reference": "",
                "Note reference": "",
                "Statement": "",
                "Line item": "",
                "Referenced note": "",
                "Suggested note": "",
                "Match confidence": "",
                "Reason": "",
                "Current year amount found?": "",
                "Prior year amount found?": "",
                "Amount found in note": "",
                "Alternative note found": "",
                "Amount match confidence": "",
                "Location": "",
                "Issue": "No automated findings were identified.",
                "Evidence": "",
                "Recommendation": "",
                "Reviewer comment": "",
                "Prepared by": "",
                "Reviewed by": "",
                "Date cleared": "",
            }
        ]
        pd.DataFrame(exception_rows).to_excel(writer, sheet_name="Exception register", index=False)
        
        note_agreement_rows = _note_agreement_result_rows(result)
        primary_rows = note_agreement_rows or [{"Statement": "No lines detected"}]
        no_notes_rows = [r for r in note_agreement_rows if r.get("Has note?") == "No"] or [{"Statement": "None found"}]
        linked_rows = [r for r in note_agreement_rows if r.get("Has note?") == "Yes"] or [{"Statement": "None found"}]
        
        pd.DataFrame(primary_rows).to_excel(writer, sheet_name="Primary statement line items", index=False)
        pd.DataFrame(no_notes_rows).to_excel(writer, sheet_name="Items without notes summary", index=False)
        pd.DataFrame(linked_rows).to_excel(writer, sheet_name="Note-linked review", index=False)

        notes_detected_rows = _note_heading_rows(result) or [{"Note": "None found"}]
        notes_heading_candidate_rows = _notes_heading_candidate_rows(result)
        ocr_statement_rows = _ocr_statement_row_rows(result)
        pd.DataFrame(notes_detected_rows).to_excel(writer, sheet_name="Notes detected", index=False)
        pd.DataFrame(notes_heading_candidate_rows).to_excel(writer, sheet_name="Notes heading candidates", index=False)
        pd.DataFrame(ocr_statement_rows).to_excel(writer, sheet_name="OCR statement rows", index=False)
        
        policy_rows = result.metrics.get("policy_export", []) or [{"Paragraph reviewed": "None found"}]
        pd.DataFrame(policy_rows).to_excel(writer, sheet_name="Notes 1 and 2 policy review", index=False)
        unref_rows = result.metrics.get("unreferenced_notes", []) or [{"Note": "None", "Heading": "None found", "Comment": "All notes referenced or filtered"}]
        pd.DataFrame(unref_rows).to_excel(writer, sheet_name="Unreferenced notes", index=False)
        
        cross_export = result.metrics.get("cross_page_export", {})
        amount_rows = [
            _translate_row_page_fields(row, result, ("Pages checked", "Context", "Issue"))
            for row in (cross_export.get("key_amounts", []) or [{"Metric": "None found"}])
        ]
        name_rows = [
            _translate_row_page_fields(row, result, ("Page 1", "Page 2"))
            for row in (clean_name_consistency_rows(cross_export.get("names", [])) or [{"Name variant 1": "None found"}])
        ]
        date_rows = [
            _translate_row_page_fields(row, result, ("Page", "Comment"))
            for row in (cross_export.get("dates", []) or [{"Date found": "None found"}])
        ]
        grammar_rows = [
            _translate_row_page_fields(row, result, ("Page", "Context"))
            for row in (cross_export.get("grammar", []) or [{"Page": "None found"}])
        ]
        
        pd.DataFrame(amount_rows).to_excel(writer, sheet_name="Key amount consistency", index=False)
        pd.DataFrame(name_rows).to_excel(writer, sheet_name="Name consistency", index=False)
        pd.DataFrame(date_rows).to_excel(writer, sheet_name="Date consistency", index=False)
        pd.DataFrame(grammar_rows).to_excel(writer, sheet_name="Grammar review", index=False)
        
        pd.DataFrame(check_results).to_excel(writer, sheet_name="Checks results", index=False)
        pd.DataFrame(checks_performed).to_excel(writer, sheet_name="Checks performed", index=False)
        pd.DataFrame(checks_skipped).to_excel(writer, sheet_name="Checks skipped", index=False)
        skipped_summary_rows = result.metrics.get("skipped_table_summary", []) or [{"Skipped check group": "None", "Reason skipped": "No table-specific skips recorded."}]
        pd.DataFrame(skipped_summary_rows).to_excel(writer, sheet_name="Skipped checks summary", index=False)
        skipped_table_rows = result.metrics.get("skipped_table_details", []) or [{"Page": "None", "Reason skipped": "No table-specific skips recorded."}]
        pd.DataFrame(skipped_table_rows).to_excel(writer, sheet_name="Skipped table details", index=False)
        for worksheet in writer.book.worksheets:
            worksheet.freeze_panes = "A2"
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F2937")
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            for column_cells in worksheet.columns:
                max_length = max((len(str(cell.value or "")) for cell in column_cells), default=10)
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 70)
        _format_exception_register_sheet(writer.book["Exception register"])
        _format_excel_table_sheet(writer.book["Findings summary"], "FindingsSummary")
        _format_excel_table_sheet(writer.book["Primary statement line items"], "PrimaryLineItems")
        _format_excel_table_sheet(writer.book["Items without notes summary"], "NoNotesSummary")
        _format_excel_table_sheet(writer.book["Note-linked review"], "NoteLinkedReview")
        _format_excel_table_sheet(writer.book["Notes detected"], "NotesDetected")
        _format_excel_table_sheet(writer.book["Notes heading candidates"], "NotesHeadingCandidates")
        _format_excel_table_sheet(writer.book["OCR statement rows"], "OCRStatementRows")
        _format_excel_table_sheet(writer.book["Notes 1 and 2 policy review"], "PolicyReview")
        _format_excel_table_sheet(writer.book["AI policy judgement"], "AIPolicyJudgement")
        _format_excel_table_sheet(writer.book["AI finding review"], "AIFindingReview")
        _format_excel_table_sheet(writer.book["AI evidence packs"], "AIEvidencePacks")
        _format_excel_table_sheet(writer.book["Unreferenced notes"], "UnreferencedNotes")
        _format_excel_table_sheet(writer.book["Key amount consistency"], "AmountConsistency")
        _format_excel_table_sheet(writer.book["Name consistency"], "NameConsistency")
        _format_excel_table_sheet(writer.book["Date consistency"], "DateConsistency")
        _format_excel_table_sheet(writer.book["Grammar review"], "GrammarReview")
        _format_excel_table_sheet(writer.book["Checks results"], "ChecksResults")
        _format_excel_table_sheet(writer.book["Skipped checks summary"], "SkippedChecksSummary")
        _format_excel_table_sheet(writer.book["Skipped table details"], "SkippedTableDetails")
    return output.getvalue()


def _format_excel_table_sheet(worksheet, table_name: str) -> None:
    max_row = max(worksheet.max_row, 2)
    max_col = worksheet.max_column
    table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _format_exception_register_sheet(worksheet) -> None:
    max_row = max(worksheet.max_row, 2)
    max_col = worksheet.max_column
    table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName="ExceptionRegister", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    headers = {str(cell.value): cell.column for cell in worksheet[1]}
    status_col = headers.get("Status")
    severity_col = headers.get("Severity")
    if status_col:
        status_letter = get_column_letter(status_col)
        validation = DataValidation(type="list", formula1='"Open,Cleared,False Positive,Noted"', allow_blank=True)
        worksheet.add_data_validation(validation)
        validation.add(f"{status_letter}2:{status_letter}500")
    if severity_col:
        severity_letter = get_column_letter(severity_col)
        color_map = {
            "High": ("FEE2E2", "991B1B"),
            "Medium": ("FEF3C7", "92400E"),
            "Low": ("DBEAFE", "1E40AF"),
        }
        for severity, (fill, font) in color_map.items():
            worksheet.conditional_formatting.add(
                f"{severity_letter}2:{severity_letter}{max_row}",
                FormulaRule(
                    formula=[f'${severity_letter}2="{severity}"'],
                    fill=PatternFill("solid", fgColor=fill),
                    font=Font(color=font, bold=True),
                ),
            )
    for header in ("Evidence", "Recommendation", "Issue", "Reviewer comment"):
        column = headers.get(header)
        if column:
            letter = get_column_letter(column)
            worksheet.column_dimensions[letter].width = 55 if header in {"Evidence", "Recommendation", "Issue"} else 32
            for row in range(2, max_row + 1):
                worksheet[f"{letter}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.auto_filter.ref = table_ref


def _docx_paragraph(text: str, style: str = "") -> str:
    style_xml = f'<w:pPr><w:pStyle w:val="{style}"/></w:pPr>' if style else ""
    return f"<w:p>{style_xml}<w:r><w:t>{xml_escape(text)}</w:t></w:r></w:p>"


def _docx_table(rows: list[list[object]]) -> str:
    table_rows = []
    for row_index, row in enumerate(rows):
        cells = []
        for value in row:
            display_value = "" if value is None else str(value)
            shading = '<w:shd w:fill="1F2937"/>' if row_index == 0 else ""
            color = '<w:color w:val="FFFFFF"/>' if row_index == 0 else ""
            bold = "<w:b/>" if row_index == 0 else ""
            cells.append(
                "<w:tc>"
                f"<w:tcPr>{shading}</w:tcPr>"
                f"<w:p><w:r><w:rPr>{bold}{color}</w:rPr><w:t>{xml_escape(display_value)}</w:t></w:r></w:p>"
                "</w:tc>"
            )
        table_rows.append(f"<w:tr>{''.join(cells)}</w:tr>")
    return (
        "<w:tbl>"
        '<w:tblPr><w:tblStyle w:val="TableGrid"/><w:tblW w:w="0" w:type="auto"/></w:tblPr>'
        f"{''.join(table_rows)}"
        "</w:tbl>"
    )


def _detected_profile_rows(result) -> list[list[str]]:
    detected = result.metrics.get("detected_profile", {})
    if not isinstance(detected, dict):
        detected = {}
    wanted = ("Company name", "Year end", "Currency", "Framework", "Entity type", "Extraction confidence")
    return [["Field", "Detected value"], *[[field, str(detected.get(field, "Not detected"))] for field in wanted]]


def _build_word_memo_export(result) -> bytes:
    memo = build_ai_review_memo(result)
    assurance = str(result.metrics.get("positive_assurance", ""))
    ai_summary = str(result.metrics.get("ai_policy_review_summary", "") or "").strip()
    ai_status = str(result.metrics.get("ai_policy_review_status", "disabled") or "disabled")
    disclaimer = (
        "This automated review supports financial statement review procedures but does not replace professional "
        "judgement, firm methodology, or the auditor's responsibility to evaluate the report and underlying evidence."
    )
    body_parts = [
        _docx_paragraph("AI Audit Assistant Review Memo", "Title"),
        _docx_paragraph("Detected Company Profile", "Heading1"),
        _docx_table(_detected_profile_rows(result)),
        _docx_paragraph("Executive Memo", "Heading1"),
        _docx_paragraph(memo),
        _docx_paragraph(disclaimer),
    ]
    if ai_summary or ai_status != "disabled":
        body_parts.append(_docx_paragraph("AI Policy Judgement", "Heading1"))
        if ai_summary:
            body_parts.append(_docx_paragraph(ai_summary))
        else:
            body_parts.append(_docx_paragraph(f"Status: {ai_status}"))
    if assurance:
        body_parts.append(_docx_paragraph(assurance))
    body_parts.extend(
        [
            _docx_paragraph("Dashboard", "Heading1"),
            _docx_table(
                [
                    ["Metric", "Value"],
                    ["Checks performed", result.metrics.get("checks_performed_count", 0)],
                    ["Checks passed", result.metrics.get("checks_passed_count", 0)],
                    ["Checks skipped", result.metrics.get("checks_skipped_count", 0)],
                    ["Findings", result.metrics.get("findings", 0)],
                    ["High findings", result.metrics.get("high", 0)],
                    ["Medium findings", result.metrics.get("medium", 0)],
                    ["Low findings", result.metrics.get("low", 0)],
                    ["OCR text coverage", result.metrics.get("ocr_text_coverage", result.metrics.get("extraction_coverage", "0%"))],
                    ["Statement structure confidence", result.metrics.get("statement_structure_confidence", "0%")],
                    ["Note structure confidence", result.metrics.get("note_structure_confidence", "0%")],
                    ["Table arithmetic confidence", _table_arithmetic_display(result)],
                ]
            ),
            _docx_paragraph("Checks Performed", "Heading1"),
        ]
    )
    body_parts.extend(_docx_paragraph(item) for item in _metric_lines(result.metrics.get("checks_performed"), "No deterministic checks completed."))
    body_parts.append(_docx_paragraph("Checks Skipped", "Heading1"))
    skipped = _metric_lines(result.metrics.get("checks_skipped"), "No major checks skipped.")
    body_parts.extend(_docx_paragraph(item) for item in skipped) if skipped else body_parts.append(_docx_paragraph("No major checks skipped."))
    body_parts.append(_docx_paragraph("Findings Summary", "Heading1"))
    if result.findings:
        grouped: dict[tuple[str, str], list[object]] = {}
        for finding in result.findings:
            grouped.setdefault((finding.severity, finding.category), []).append(finding)
        severity_order = {"High": 0, "Medium": 1, "Low": 2}
        for (severity, category), findings in sorted(grouped.items(), key=lambda item: (severity_order.get(item[0][0], 9), item[0][1])):
            body_parts.append(_docx_paragraph(f"{severity} | {category}", "Heading2"))
            body_parts.append(
                _docx_table(
                    [
                        ["Location", "Issue", "Evidence", "Recommendation"],
                        *[
                            [finding.location, finding.issue, finding.evidence, finding.recommendation]
                            for finding in findings
                        ],
                    ]
                )
            )
    else:
        body_parts.append(_docx_paragraph("No automated findings were identified."))
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body>{''.join(body_parts)}<w:sectPr><w:pgSz w:w=\"12240\" w:h=\"15840\"/><w:pgMar w:top=\"1440\" w:right=\"1440\" w:bottom=\"1440\" w:left=\"1440\"/></w:sectPr></w:body>"
        "</w:document>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        "</Types>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
        "</Relationships>"
    )
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as docx:
        docx.writestr("[Content_Types].xml", content_types)
        docx.writestr("_rels/.rels", rels)
        docx.writestr("word/document.xml", document_xml)
    return output.getvalue()


st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

    :root {
        --bg-color: #0B1120;
        --panel-bg: rgba(30, 41, 59, 0.7);
        --text-primary: #F8FAFC;
        --text-muted: #94A3B8;
        --accent-blue: #3B82F6;
        --accent-purple: #8B5CF6;
        --border-color: rgba(255, 255, 255, 0.1);
        --gold: #F59E0B;
        --green: #10B981;
        --red: #EF4444;
    }

    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif !important;
        color: var(--text-primary);
    }

    .stApp {
        background: var(--bg-color);
        background-image: 
            radial-gradient(at 0% 0%, rgba(59, 130, 246, 0.15) 0px, transparent 50%),
            radial-gradient(at 100% 0%, rgba(139, 92, 246, 0.15) 0px, transparent 50%);
        background-attachment: fixed;
    }

    [data-testid="stHeader"] {
        background: transparent !important;
    }

    .block-container {
        max-width: 1440px;
        padding-top: 3rem;
        padding-bottom: 4rem;
    }

    h1, h2, h3, h4, h5, h6 {
        letter-spacing: -0.02em;
        color: var(--text-primary);
        font-weight: 600;
    }

    /* Premium Hero Section */
    .premium-hero {
        background: linear-gradient(135deg, rgba(15, 23, 42, 0.8), rgba(30, 41, 59, 0.9));
        border: 1px solid var(--border-color);
        border-radius: 16px;
        padding: 40px 48px;
        box-shadow: 0 20px 40px -10px rgba(0,0,0,0.5);
        margin-bottom: 32px;
        backdrop-filter: blur(12px);
        -webkit-backdrop-filter: blur(12px);
        position: relative;
        overflow: hidden;
    }
    
    .premium-hero::before {
        content: '';
        position: absolute;
        top: 0; left: 0; right: 0; height: 1px;
        background: linear-gradient(90deg, transparent, var(--accent-blue), var(--accent-purple), transparent);
        opacity: 0.5;
    }

    .eyebrow {
        color: var(--accent-blue);
        font-size: 13px;
        font-weight: 700;
        letter-spacing: 0.15em;
        text-transform: uppercase;
        margin-bottom: 12px;
    }

    .premium-title {
        font-size: 42px;
        font-weight: 800;
        letter-spacing: -0.03em;
        margin-bottom: 16px;
        background: linear-gradient(135deg, #FFFFFF, #94A3B8);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        line-height: 1.1;
    }

    .premium-subtitle {
        font-size: 18px;
        color: var(--text-muted);
        max-width: 800px;
        line-height: 1.6;
    }

    @media (max-width: 1100px) {
        .module-grid {
            grid-template-columns: repeat(2, minmax(0, 1fr));
        }
        .premium-title {
            font-size: 34px;
        }
    }

    @media (max-width: 680px) {
        .module-grid {
            grid-template-columns: 1fr;
        }
        .premium-hero {
            padding: 26px 22px;
        }
        .premium-title {
            font-size: 30px;
        }
    }

    /* Module Grid */
    .module-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
        gap: 16px;
        margin-bottom: 24px;
    }
    .module-card {
        background: var(--panel-bg);
        border: 1px solid var(--border-color);
        backdrop-filter: blur(12px);
        -webkit-backdrop-filter: blur(12px);
        border-radius: 12px;
        padding: 20px;
        transition: transform 0.2s ease, box-shadow 0.2s ease, border-color 0.2s ease;
    }
    .module-card:hover {
        transform: translateY(-4px);
        box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.2), 0 0 15px rgba(59, 130, 246, 0.15);
        border-color: rgba(59, 130, 246, 0.4);
    }
    .module-index {
        color: var(--accent-blue);
        font-size: 12px;
        font-weight: 700;
        letter-spacing: 0.1em;
        text-transform: uppercase;
        margin-bottom: 8px;
    }
    .module-title {
        color: var(--text-primary);
        font-size: 16px;
        font-weight: 700;
        margin-bottom: 8px;
    }
    .module-copy {
        color: var(--text-muted);
        font-size: 14px;
        line-height: 1.5;
    }

    /* Additional Text Classes */
    .profile-heading {
        font-size: 16px;
        font-weight: 600;
        color: var(--text-primary);
        margin-bottom: 4px;
    }
    .profile-copy {
        font-size: 14px;
        color: var(--text-muted);
        margin-bottom: 12px;
    }
    .control-label {
        font-size: 14px;
        font-weight: 600;
        color: var(--text-primary);
        margin-bottom: 8px;
    }
    
    /* Memo specific styles */
    .memo-panel {
        background: rgba(15, 23, 42, 0.6);
        border-radius: 12px;
        border: 1px solid var(--border-color);
        padding: 24px;
        margin-bottom: 16px;
    }
    .memo-kicker {
        font-size: 12px;
        font-weight: 700;
        color: var(--accent-blue);
        text-transform: uppercase;
        letter-spacing: 0.1em;
        margin-bottom: 8px;
    }
    .memo-text {
        font-size: 15px;
        color: var(--text-primary);
        line-height: 1.6;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <section class="premium-hero">
        <div class="eyebrow">Financial Statement Intelligence</div>
        <div class="premium-title">AI Audit Assistant</div>
        <p class="premium-subtitle">
            A high-assurance review workspace for prepared financial statements, built to surface arithmetic,
            presentation, note-agreement, policy, and standards-checklist exceptions before final sign-off.
        </p>
    </section>
    """,
    unsafe_allow_html=True,
)

with st.container(border=True):
    st.markdown(
        """
        <div class="profile-heading">Review mode</div>
        <div class="profile-copy">Quick Review only needs the PDF. Advanced Review lets you override the detected engagement profile.</div>
        """,
        unsafe_allow_html=True,
    )
    review_mode = st.radio("Mode", ["Quick Review", "Advanced Review"], horizontal=True, index=1)
    company_name = ""
    industry = ""
    reporting_currency = "Not specified"
    presentation_standard = "IFRS"
    expected_policies_text = ""
    significant_transactions_text = ""
    checklist_areas_text = ""
    if review_mode == "Advanced Review":
        profile_cols = st.columns([1.2, 1, 0.8, 0.8])
        company_name = profile_cols[0].text_input("Company name")
        industry = profile_cols[1].text_input("Industry")
        currency_choice = profile_cols[2].selectbox(
            "Reporting currency",
            ["Not specified", "NGN", "USD", "GBP", "EUR", "ZAR", "GHS", "KES", "Other"],
            index=0,
        )
        reporting_currency = currency_choice
        if currency_choice == "Other":
            reporting_currency = profile_cols[2].text_input("Custom currency", placeholder="Example: NGN")
        presentation_standard = profile_cols[3].selectbox("Presentation standard", ["IFRS", "Local GAAP"], index=0)
        detail_cols = st.columns(3)
        expected_policies_text = detail_cols[0].text_area(
            "Expected policies",
            placeholder="Example: revenue, financial instruments, tax",
            help="Comma-separated policy areas that are expected even if balances are not obvious in the extracted PDF text.",
        )
        significant_transactions_text = detail_cols[1].text_area(
            "Significant transactions",
            placeholder="Example: leases, share-based payments, foreign currency loans",
            help="Comma-separated transactions or balances that should have tailored accounting policy coverage.",
        )
        checklist_areas_text = detail_cols[2].text_area(
            "Force checklist areas",
            placeholder="Example: IFRS 15, IFRS 16, revenue, leases, EPS",
            help="Comma-separated standards or areas to check even if the PDF text does not clearly trigger them.",
        )
    ocr_cols = st.columns([1, 1, 1])
    ocr_cols[0].markdown('<div class="control-label">OCR scanned PDFs</div>', unsafe_allow_html=True)
    use_ocr = ocr_cols[0].toggle(
        "Enable OCR for scanned PDFs",
        value=True,
        label_visibility="collapsed",
        help="When text coverage is low, render pages in memory and run local Tesseract OCR. No OCR output or images are saved.",
    )
    ocr_max_pages = ocr_cols[1].number_input(
        "OCR page limit",
        min_value=1,
        max_value=300,
        value=60,
        step=5,
        help="Limits OCR work for very large PDFs. Increase if the full report is scanned.",
    )
    ocr_dpi = ocr_cols[2].select_slider(
        "OCR quality",
        options=[150, 200, 250, 300],
        value=300,
        help="Higher DPI can improve OCR accuracy but takes longer.",
    )
    ai_review_mode = "Standard review"
    with st.expander("Advanced settings", expanded=review_mode == "Advanced Review"):
        cautious_note_agreement = st.checkbox(
            "Run cautious note-reference validation anyway",
            value=True,
            key="run_cautious_note_reference_validation_anyway",
            help=(
                "When note/table confidence is below 80%, detailed note-reference validation is normally skipped. "
                "Enable this to run a heading-based review-prompt check using detected note headings and line-extracted primary statement rows. "
                "Detailed amount agreement remains skipped until table confidence improves."
            ),
        )
        use_ai_policy_review = st.checkbox(
            "Enable AI policy judgement and review assist",
            value=True,
            help=(
                "Runs an optional AI review over accounting policies and related disclosures to assess "
                "policy relevance, standards context, disclosure completeness, and industry fit. "
                "Requires OPENAI_API_KEY on the server."
            ),
        )
        ai_review_mode = st.selectbox(
            "AI review mode",
            ["Quick review", "Standard review", "Deep review"],
            index=1,
            help=(
                "Quick uses the smallest compact evidence package. Standard reviews primary statements, Notes 1 and 2, and the draft register. "
                "Deep includes broader context and may use more tokens."
            ),
        )
        use_ai_full_review = st.checkbox(
            "Run full AI financial statement review",
            value=True,
            help=(
                "Runs a ChatGPT-style quality-control pass over extracted statement text and note context. "
                "Outputs concise findings, evidence snippets, page/note references, reviewer rationale, and recommended actions."
            ),
        )
    if review_mode != "Advanced Review":
        openai_key = os.getenv("OPENAI_API_KEY", "").strip()
        if not openai_key:
            use_ai_policy_review = False
            use_ai_full_review = False

st.markdown('<div class="section-label">Audit review modules</div>', unsafe_allow_html=True)
st.markdown(
    """
    <div class="module-grid">
        <div class="module-card">
            <div class="module-index">Module 01</div>
            <div class="module-title">Totals and rounding</div>
            <div class="module-copy">Totals, subtotals, cross-footings, duplicate totals, and $000s / millions labels.</div>
        </div>
        <div class="module-card">
            <div class="module-index">Module 02</div>
            <div class="module-title">Formatting</div>
            <div class="module-copy">Number styles, brackets for negatives, currency markers, comparatives, and headings.</div>
        </div>
        <div class="module-card">
            <div class="module-index">Module 03</div>
            <div class="module-title">Notes agreement</div>
            <div class="module-copy">Face statement references, segment totals, EPS, tax, depreciation, and note totals.</div>
        </div>
        <div class="module-card">
            <div class="module-index">Module 04</div>
            <div class="module-title">Accounting policies</div>
            <div class="module-copy">Irrelevant policies, boilerplate wording, missing policies, and superseded standards.</div>
        </div>
        <div class="module-card">
            <div class="module-index">Module 05</div>
            <div class="module-title">Standards checklist</div>
            <div class="module-copy">Triggered IFRS disclosure checks for presentation and significant transaction areas.</div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown('<div class="upload-shell">', unsafe_allow_html=True)
uploaded = st.file_uploader("Upload prepared financial statement PDF", type=["pdf"])
st.markdown("</div>", unsafe_allow_html=True)

if not uploaded:
    st.info("Upload a PDF to start the review.")
    st.stop()

expected_policies = tuple(item.strip() for item in expected_policies_text.split(",") if item.strip())
significant_transactions = tuple(item.strip() for item in significant_transactions_text.split(",") if item.strip())
checklist_areas = tuple(item.strip() for item in checklist_areas_text.split(",") if item.strip())
normalised_currency = "" if reporting_currency == "Not specified" else normalize_reporting_currency(reporting_currency)
if reporting_currency != "Not specified" and not normalised_currency:
    st.warning("Enter a valid reporting currency before running the review. Example: NGN, USD, GBP, EUR, ZAR, GHS, or KES.")
    st.stop()
profile = CompanyProfile(
    company_name=company_name.strip(),
    industry=industry.strip(),
    reporting_currency=normalised_currency,
    expected_policies=expected_policies,
    significant_transactions=significant_transactions,
    presentation_standard=presentation_standard,
    checklist_areas=checklist_areas,
)

review_options = ReviewOptions(
    use_ocr=use_ocr,
    ocr_max_pages=int(ocr_max_pages),
    ocr_dpi=int(ocr_dpi),
    run_cautious_note_agreement=cautious_note_agreement,
    use_ai_policy_review=use_ai_policy_review,
    use_ai_full_review=use_ai_full_review,
    ai_model=DEFAULT_AI_MODEL,
    ai_review_mode=ai_review_mode,
)
deterministic_review_options = ReviewOptions(
    use_ocr=use_ocr,
    ocr_max_pages=int(ocr_max_pages),
    ocr_dpi=int(ocr_dpi),
    run_cautious_note_agreement=cautious_note_agreement,
    use_ai_policy_review=False,
    use_ai_full_review=False,
    ai_model=DEFAULT_AI_MODEL,
    ai_review_mode=ai_review_mode,
)
uploaded_bytes = uploaded.getvalue()
file_hash = hashlib.sha256(uploaded_bytes).hexdigest()
settings_key = repr(
    (
        review_mode,
        company_name.strip(),
        industry.strip(),
        normalised_currency,
        presentation_standard,
        expected_policies,
        significant_transactions,
        checklist_areas,
        bool(use_ocr),
        int(ocr_max_pages),
        int(ocr_dpi),
        bool(cautious_note_agreement),
        bool(use_ai_policy_review),
        bool(use_ai_full_review),
        DEFAULT_AI_MODEL,
        ai_review_mode,
    )
)
cache = st.session_state.get("review_cache", {})
same_cached_file = cache.get("file_hash") == file_hash and cache.get("settings_key") == settings_key
cached_result = cache.get("result") if same_cached_file else None
ai_retry_available = bool(
    cached_result
    and str(cached_result.metrics.get("ai_review_status", "")).lower().startswith("failed")
    and cache.get("document") is not None
)
retry_ai_requested = False
if ai_retry_available:
    retry_ai_requested = st.button(
        "Retry AI Review",
        type="primary",
        help="Reuse the cached extraction and deterministic review context, then rerun the AI review layer.",
    )
run_ai_requested = bool(st.session_state.pop("run_ai_review_requested", False))

if cached_result is not None and not retry_ai_requested and not run_ai_requested:
    result = cached_result
else:
    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_file:
            temp_file.write(uploaded_bytes)
            temp_path = Path(temp_file.name)
        if (retry_ai_requested or run_ai_requested) and same_cached_file and cache.get("document") is not None and cache.get("result") is not None:
            with st.spinner("Running AI review using cached extraction and deterministic findings..."):
                document = cache["document"]
                result = rerun_ai_review_from_cached_result(document, temp_path, profile, review_options, cache["result"])
        else:
            with st.spinner("Extracting PDF text, running OCR if needed, and performing deterministic review checks..."):
                document = extract_review_document(temp_path, deterministic_review_options)
                result = review_document(document, temp_path, profile, deterministic_review_options)
        st.session_state["review_cache"] = {
            "file_hash": file_hash,
            "settings_key": settings_key,
            "uploaded_bytes": uploaded_bytes,
            "document": document,
            "result": result,
        }
        cache = st.session_state["review_cache"]
    except Exception as exc:
        st.error(_friendly_review_failure_message(exc))
        st.stop()
    finally:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)

ai_layer_not_started = str(result.metrics.get("ai_review_status", "Not started") or "Not started") == "Not started"
ai_layer_enabled = bool(use_ai_policy_review or use_ai_full_review)
if ai_layer_enabled and ai_layer_not_started and cache.get("document") is not None:
    st.info("Deterministic review is ready. Run the AI layer separately to avoid blocking the base exception register during extraction.")
    if st.button("Run AI Review", type="primary", help="Use the cached extraction and deterministic findings; no PDF re-upload or OCR rerun is needed."):
        st.session_state["run_ai_review_requested"] = True
        st.rerun()

st.markdown('<div class="section-label">Review dashboard</div>', unsafe_allow_html=True)
review_cols = st.columns(5)
review_cols[0].metric("Checks performed", result.metrics.get("checks_performed_count", 0))
review_cols[1].metric("Checks passed", result.metrics.get("checks_passed_count", 0))
review_cols[2].metric("Checks skipped", result.metrics.get("checks_skipped_count", 0))
review_cols[3].metric("Findings", result.metrics["findings"])
review_cols[4].metric("OCR text coverage", result.metrics.get("ocr_text_coverage", result.metrics.get("extraction_coverage", "0%")))

risk_cols = st.columns(8)
risk_cols[0].metric("High", result.metrics["high"])
risk_cols[1].metric("Medium", result.metrics["medium"])
risk_cols[2].metric("Low", result.metrics["low"])
risk_cols[3].metric("Pages", result.metrics["pages"])
risk_cols[4].metric("OCR table candidates", result.metrics.get("ocr_tables", 0))
risk_cols[5].metric("Statement structure", result.metrics.get("statement_structure_confidence", "0%"))
risk_cols[6].metric("Note structure", result.metrics.get("note_structure_confidence", "0%"))
risk_cols[7].metric("Table arithmetic", _table_arithmetic_display(result))

ai_overall_status = str(result.metrics.get("ai_review_status", "Not started") or "Not started")
ai_stage_rows = result.metrics.get("ai_review_stage_status", [])
st.markdown('<div class="section-label">AI review status</div>', unsafe_allow_html=True)
if ai_overall_status == "Completed":
    st.success("AI review completed.")
elif ai_overall_status == "Not started":
    st.info("AI review not started.")
elif ai_overall_status == "Skipped":
    st.info("AI review was skipped because no suitable AI context was available.")
elif ai_overall_status.startswith("Failed"):
    st.warning("AI review failed after automatic retries. The deterministic review and exports are still available; use Retry AI Review to run the AI layer again.")
else:
    st.info(f"AI review status: {ai_overall_status}")
st.caption(f"AI review mode: {result.metrics.get('ai_review_mode', 'standard')}")
if isinstance(ai_stage_rows, list) and ai_stage_rows:
    st.dataframe(pd.DataFrame(ai_stage_rows), use_container_width=True, hide_index=True)
ai_error_rows = result.metrics.get("ai_error_log", [])
if isinstance(ai_error_rows, list) and ai_error_rows:
    with st.expander("AI debug details", expanded=False):
        st.dataframe(pd.DataFrame(ai_error_rows), use_container_width=True, hide_index=True)
detected_profile = result.metrics.get("detected_profile", {})
if isinstance(detected_profile, dict):
    st.markdown('<div class="section-label">Detected profile after upload</div>', unsafe_allow_html=True)
    profile_rows = [
        {"Field": key, "Detected value": value}
        for key, value in detected_profile.items()
    ]
    st.dataframe(pd.DataFrame(profile_rows), use_container_width=True, hide_index=True)

if result.metrics.get("ai_policy_review_status") == "completed":
    ai_summary = str(result.metrics.get("ai_policy_review_summary", "") or "").strip()
    if ai_summary:
        st.markdown('<div class="section-label">AI policy judgement</div>', unsafe_allow_html=True)
        st.info(ai_summary)
elif use_ai_policy_review:
    ai_status = str(result.metrics.get("ai_policy_review_status", "disabled") or "disabled")
    ai_message = str(result.metrics.get("ai_policy_review_message", "") or "").strip()
    st.markdown('<div class="section-label">AI policy judgement</div>', unsafe_allow_html=True)
    if ai_message:
        if ai_status == "deferred":
            st.info(ai_message)
        else:
            st.warning(f"Status: {ai_status}. {ai_message}")
    else:
        st.warning(f"Status: {ai_status}. No AI policy observations were returned.")

if result.metrics.get("ai_full_review_status") == "completed":
    ai_full_summary = str(result.metrics.get("ai_full_review_summary", "") or "").strip()
    if ai_full_summary:
        st.markdown('<div class="section-label">AI full review</div>', unsafe_allow_html=True)
        st.info(ai_full_summary)
elif use_ai_full_review:
    ai_full_status = str(result.metrics.get("ai_full_review_status", "disabled") or "disabled")
    ai_full_message = str(result.metrics.get("ai_full_review_message", "") or "").strip()
    if ai_full_status != "disabled":
        st.markdown('<div class="section-label">AI full review</div>', unsafe_allow_html=True)
        if ai_full_message:
            if ai_full_status == "deferred":
                st.info(ai_full_message)
            else:
                st.warning(f"Status: {ai_full_status}. {ai_full_message}")
        else:
            st.info(f"Status: {ai_full_status}.")

if result.metrics.get("ai_finding_review_status") == "completed":
    ai_finding_summary = str(result.metrics.get("ai_finding_review_summary", "") or "").strip()
    if ai_finding_summary:
        st.markdown('<div class="section-label">AI finding review</div>', unsafe_allow_html=True)
        st.info(ai_finding_summary)
elif use_ai_policy_review or use_ai_full_review:
    ai_finding_status = str(result.metrics.get("ai_finding_review_status", "disabled") or "disabled")
    ai_finding_message = str(result.metrics.get("ai_finding_review_message", "") or "").strip()
    if ai_finding_status != "disabled":
        st.markdown('<div class="section-label">AI finding review</div>', unsafe_allow_html=True)
        if ai_finding_message:
            if ai_finding_status == "deferred":
                st.info(ai_finding_message)
            else:
                st.warning(f"Status: {ai_finding_status}. {ai_finding_message}")
        else:
            st.info(f"Status: {ai_finding_status}.")

st.markdown(
    f"""
    <section class="memo-panel">
        <div class="memo-kicker">Executive review memo</div>
        <div class="memo-text">{build_ai_review_memo(result)}</div>
    </section>
    """,
    unsafe_allow_html=True,
)

export_stem = exported_file_stem(result)
download_cols = st.columns(2)
download_cols[0].download_button(
    "Download Excel Exception Register",
    build_excel_export(result),
    file_name=f"{export_stem}_exception_register.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
download_cols[1].download_button(
    "Download Word Review Memo",
    _build_word_memo_export(result),
    file_name=f"{export_stem}_review_memo.docx",
    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
)

status_cols = st.columns(2)
with status_cols[0]:
    with st.expander("Checks performed", expanded=True):
        assurance = str(result.metrics.get("positive_assurance", ""))
        if assurance:
            st.success(assurance)
        st.write(str(result.metrics.get("checks_performed", "No deterministic checks completed.")))
with status_cols[1]:
    with st.expander("Checks skipped", expanded=False):
        skipped_text = str(result.metrics.get("checks_skipped", "No major checks skipped."))
        if skipped_text.strip() == "No major checks skipped.":
            st.success(skipped_text)
        else:
            st.write(skipped_text)

if not result.findings:
    st.success("No issues were detected by the automated checks.")
    with st.expander("Developer and debug details", expanded=False):
        note_rows = _note_heading_rows(result)
        if note_rows:
            st.markdown("**Notes detected**")
            st.dataframe(pd.DataFrame(note_rows), use_container_width=True, hide_index=True)
        ocr_rows = _ocr_statement_row_rows(result)
        st.markdown("**OCR statement rows**")
        st.dataframe(pd.DataFrame(ocr_rows), use_container_width=True, hide_index=True)
        st.markdown("**Note validation debug**")
        st.json(
            {
                "cautious_note_validation_enabled": result.metrics.get("cautious_note_validation_enabled", False),
                "note_validation_mode": result.metrics.get("note_validation_mode", "skipped"),
                "note_reference_rows_detected": result.metrics.get("note_reference_rows_detected", 0),
                "note_headings_detected": result.metrics.get("note_headings_detected", 0),
                "notes_section_start_page": result.metrics.get("notes_section_start_page", "Not detected"),
                "note_reference_findings": result.metrics.get("note_reference_findings", 0),
                "ocr_text_coverage": result.metrics.get("ocr_text_coverage", result.metrics.get("extraction_coverage", "0%")),
                "statement_structure_confidence": result.metrics.get("statement_structure_confidence", "0%"),
                "note_structure_confidence": result.metrics.get("note_structure_confidence", "0%"),
                "table_arithmetic_confidence": result.metrics.get("table_arithmetic_confidence", "0%"),
                "primary_statement_pages": result.metrics.get("primary_statement_pages", "No primary statement pages detected."),
                "ai_review_mode": result.metrics.get("ai_review_mode", "standard"),
                "ai_error_log": result.metrics.get("ai_error_log", []),
            }
        )
        markdown_report = findings_to_markdown(result)
        st.download_button(
            "Download Markdown Developer Report",
            markdown_report,
            file_name=f"{export_stem}_review_debug.md",
            mime="text/markdown",
        )
    st.stop()

severity_order = ["High", "Medium", "Low"]
st.markdown('<div class="section-label">Findings</div>', unsafe_allow_html=True)
filter_cols = st.columns([1.4, 1])
category_filter = filter_cols[0].multiselect(
    "Category",
    sorted({finding.category for finding in result.findings}),
    placeholder="All categories",
)
severity_filter = filter_cols[1].multiselect("Severity", severity_order, default=severity_order)

filtered = [
    finding
    for finding in result.findings
    if (not category_filter or finding.category in category_filter)
    and (not severity_filter or finding.severity in severity_filter)
]
rows = [
    {
        "Severity": finding.severity,
        "Category": finding.category,
        "Page reference": _page_reference_for_finding(finding, result),
        "Note reference": _note_reference_for_finding(finding, result),
        "Location": _translate_page_tokens(finding.location, result),
        "Issue": finding.issue,
        "Evidence": _translate_page_tokens(finding.evidence, result),
        "Recommendation": finding.recommendation,
    }
    for finding in filtered
]
st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

with st.expander("Finding details", expanded=False):
    for finding in filtered:
        page_ref = _page_reference_for_finding(finding, result) or finding.location
        note_ref = _note_reference_for_finding(finding, result)
        with st.expander(f"{finding.severity}: {page_ref} - {finding.issue}", expanded=False):
            st.write(f"**Category:** {finding.category}")
            st.write(f"**Page reference:** {page_ref}")
            if note_ref:
                st.write(f"**Note reference:** {note_ref}")
            st.write(f"**Location:** {_translate_page_tokens(finding.location, result)}")
            st.write(f"**Evidence:** {_translate_page_tokens(finding.evidence, result)}")
            st.write(f"**Recommendation:** {finding.recommendation}")

with st.expander("Developer and debug details", expanded=False):
    note_rows = _note_heading_rows(result)
    if note_rows:
        st.markdown("**Notes detected**")
        st.dataframe(pd.DataFrame(note_rows), use_container_width=True, hide_index=True)
    else:
        st.markdown("**Notes detected**")
        st.write("No note headings detected.")

    st.markdown("**OCR statement rows**")
    ocr_rows = _ocr_statement_row_rows(result)
    st.dataframe(pd.DataFrame(ocr_rows), use_container_width=True, hide_index=True)

    st.markdown("**Note validation debug**")
    st.json(
        {
            "cautious_note_validation_enabled": result.metrics.get("cautious_note_validation_enabled", False),
            "note_validation_mode": result.metrics.get("note_validation_mode", "skipped"),
            "note_reference_rows_detected": result.metrics.get("note_reference_rows_detected", 0),
            "note_headings_detected": result.metrics.get("note_headings_detected", 0),
            "notes_section_start_page": result.metrics.get("notes_section_start_page", "Not detected"),
            "note_reference_findings": result.metrics.get("note_reference_findings", 0),
            "ocr_text_coverage": result.metrics.get("ocr_text_coverage", result.metrics.get("extraction_coverage", "0%")),
            "statement_structure_confidence": result.metrics.get("statement_structure_confidence", "0%"),
            "note_structure_confidence": result.metrics.get("note_structure_confidence", "0%"),
            "table_arithmetic_confidence": result.metrics.get("table_arithmetic_confidence", "0%"),
            "primary_statement_pages": result.metrics.get("primary_statement_pages", "No primary statement pages detected."),
        }
    )

    markdown_report = findings_to_markdown(result)
    st.download_button(
        "Download Markdown Developer Report",
        markdown_report,
        file_name=f"{export_stem}_review_debug.md",
        mime="text/markdown",
    )

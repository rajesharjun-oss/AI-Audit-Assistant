from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from models import CompanyProfile, ReviewOptions
from report_exports import build_excel_export
from reviewer import review_pdf


ROOT = Path(__file__).resolve().parents[1]
FIXTURE_DIR = ROOT / "local_test_fixtures" / "cases"


def _case_files() -> list[Path]:
    if not FIXTURE_DIR.exists():
        return []
    return sorted(FIXTURE_DIR.glob("*.json"))


def _load_case(case_path: Path) -> dict:
    return json.loads(case_path.read_text(encoding="utf-8"))


def _summary_map(workbook) -> dict[str, object]:
    sheet = workbook["Summary"]
    values: dict[str, object] = {}
    for metric, value in sheet.iter_rows(min_row=2, values_only=True):
        if metric:
            values[str(metric)] = value
    return values


def _column_index_map(sheet) -> dict[str, int]:
    headers = [cell.value for cell in sheet[1]]
    return {str(header): idx for idx, header in enumerate(headers) if header}


def _sheet_rows(sheet) -> list[dict[str, object]]:
    col_map = _column_index_map(sheet)
    rows: list[dict[str, object]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({header: row[idx] for header, idx in col_map.items()})
    return rows


def _build_profile(profile_data: dict | None) -> CompanyProfile:
    profile_data = profile_data or {}
    return CompanyProfile(
        company_name=profile_data.get("company_name", ""),
        industry=profile_data.get("industry", ""),
        reporting_currency=profile_data.get("reporting_currency", ""),
        expected_policies=tuple(profile_data.get("expected_policies", [])),
        significant_transactions=tuple(profile_data.get("significant_transactions", [])),
        presentation_standard=profile_data.get("presentation_standard", "IFRS"),
        checklist_areas=tuple(profile_data.get("checklist_areas", [])),
    )


def _build_options(options_data: dict | None) -> ReviewOptions:
    options_data = options_data or {}
    return ReviewOptions(
        use_ocr=bool(options_data.get("use_ocr", False)),
        ocr_max_pages=options_data.get("ocr_max_pages", 60),
        ocr_dpi=options_data.get("ocr_dpi", 300),
        run_cautious_note_agreement=bool(options_data.get("run_cautious_note_agreement", False)),
        use_ai_policy_review=bool(options_data.get("use_ai_policy_review", False)),
    )


@pytest.mark.parametrize("case_path", _case_files(), ids=lambda path: path.stem)
def test_local_pipeline_regression(case_path: Path, tmp_path: Path):
    case = _load_case(case_path)
    pdf_path = Path(case["pdf_path"])
    if not pdf_path.exists():
        pytest.skip(f"Fixture PDF not found on this machine: {pdf_path}")

    result = review_pdf(pdf_path, _build_profile(case.get("profile")), _build_options(case.get("options")))
    output_bytes = build_excel_export(result)
    output_path = tmp_path / f"{case.get('name', case_path.stem)}_exception_register.xlsx"
    output_path.write_bytes(output_bytes)

    assert output_path.exists(), "Expected Excel output file was not created."
    assert output_path.stat().st_size > 0, "Excel output file is empty."

    workbook = openpyxl.load_workbook(output_path, data_only=True)
    summary = _summary_map(workbook)
    findings_summary_rows = _sheet_rows(workbook["Findings summary"])
    exception_rows = _sheet_rows(workbook["Exception register"])
    checks_performed_rows = _sheet_rows(workbook["Checks performed"])
    checks_skipped_rows = _sheet_rows(workbook["Checks skipped"])

    assertions = case.get("assertions", {})
    required_sheets = assertions.get("required_sheets", [])
    for sheet_name in required_sheets:
        assert sheet_name in workbook.sheetnames, f"Required sheet missing: {sheet_name}"

    for metric_name, expected_value in assertions.get("expected_summary", {}).items():
        assert summary.get(metric_name) == expected_value, f"Summary metric mismatch for {metric_name}: expected {expected_value!r}, got {summary.get(metric_name)!r}"

    for metric_name, max_value in assertions.get("max_summary", {}).items():
        value = summary.get(metric_name)
        assert value is not None and value <= max_value, f"Summary metric {metric_name} expected <= {max_value}, got {value}"

    for metric_name, min_value in assertions.get("min_summary", {}).items():
        value = summary.get(metric_name)
        assert value is not None and value >= min_value, f"Summary metric {metric_name} expected >= {min_value}, got {value}"

    performed_text = "\n".join(str(row.get("Check performed", "")) for row in checks_performed_rows)
    skipped_text = "\n".join(" | ".join(str(value or "") for value in row.values()) for row in checks_skipped_rows)
    findings_text = "\n".join(
        " | ".join(str(row.get(key, "")) for key in ("Severity", "Category", "Page reference", "Note reference", "Issue"))
        for row in findings_summary_rows
    )
    exception_text = "\n".join(
        " | ".join(str(row.get(key, "")) for key in ("Severity", "Category", "Page reference", "Note reference", "Issue"))
        for row in exception_rows
    )

    for snippet in assertions.get("required_checks_performed", []):
        assert snippet in performed_text, f"Expected performed check text not found: {snippet}"

    for snippet in assertions.get("forbidden_checks_skipped", []):
        assert snippet not in skipped_text, f"Forbidden skipped-check text found: {snippet}"

    for snippet in assertions.get("required_findings", []):
        assert snippet in findings_text or snippet in exception_text, f"Expected finding text not found: {snippet}"

    for snippet in assertions.get("forbidden_findings", []):
        assert snippet not in findings_text and snippet not in exception_text, f"Forbidden finding text found: {snippet}"

    if assertions.get("require_exception_rows", False):
        meaningful_rows = [row for row in exception_rows if str(row.get("Issue", "")).strip() and "No automated findings" not in str(row.get("Issue", ""))]
        assert meaningful_rows, "Expected exception register rows were not generated."


def test_local_pipeline_regression_has_cases_or_skips_cleanly():
    if not _case_files():
        pytest.skip("No local regression case manifests found under local_test_fixtures/cases.")
    assert True

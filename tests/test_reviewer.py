from decimal import Decimal
from io import BytesIO
from pathlib import Path

import openpyxl
import extraction
import reviewer
import ai_combined_review
import ai_finding_review
import ai_policy_review
import ai_review_pipeline
from ai_finding_review import run_ai_finding_review
from ai_policy_review import _parse_response_json
from cross_page_consistency import _names_look_like_spelling_variants, check_cross_page_consistency
from extraction import _line_to_table_row, _reconstruct_ocr_tables, extract_pdf_with_ocr
from report_exports import build_excel_export, parse_skipped_check
from models import CompanyProfile, PdfDocument, PdfPage, ReviewOptions
from reviewer import (
    _note_headings,
    _note_headings_by_page,
    _amounts_from_statement_line,
    _simple_note_amounts_from_line,
    _line_amount_for_aliases,
    _parse_ocr_statement_row,
    check_formatting,
    check_extraction_quality,
    check_notes_agreement,
    check_primary_statement_consistency,
    check_policy_relevance,
    check_rounding_and_casting,
    check_standard_checklist,
    build_ai_review_memo,
    infer_detected_profile,
    normalize_reporting_currency,
    review_pdf,
)


def test_name_consistency_only_flags_typo_like_variants_not_joined_names():
    assert _names_look_like_spelling_variants("Mzer Michael Terungwa", "Mzer Micheal Terungwa")
    assert _names_look_like_spelling_variants("Lai Labode", "Lait Labode")

    assert not _names_look_like_spelling_variants("Aliyu Ibrahim Bala", "Aliyu Ibrahim Bala Edeh")
    assert not _names_look_like_spelling_variants("Edeh Anthony Uzodinma", "Anthony Uzodinma")
    assert not _names_look_like_spelling_variants("Adeoye Simileoluwa", "Simileoluwa Adeoye")
    assert not _names_look_like_spelling_variants("Lai Labode", "Lai Labode Stanley Emurotu")


def test_detected_profile_prefers_company_profile_context_over_unrelated_membership_language():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Powergen Interconnected Energy Limited\n"
                "Corporate information\n"
                "Nature of business and principal activities\n"
                "The company engages in the business of provision of distributable renewable energy generation solutions, "
                "buying and distribution of mini grids for energy generation in urban and rural households.\n",
                [],
            ),
            PdfPage(2, "Directors' report\nShare capital\nOrdinary shares\n", []),
            PdfPage(20, "Other note text mentioning members, associates, subscriptions and council governance in a different context.\n", []),
        ]
    )

    profile = infer_detected_profile(document)

    assert profile["Entity type"] == "Private company"
    assert "Professional membership body" not in profile["Principal activities"]
    assert "Renewable energy" in profile["Principal activities"]


def test_scalar_equation_uses_mismatch_wording_for_failures():
    findings: list[reviewer.Finding] = []

    reviewer._check_scalar_equation(
        findings,
        20,
        "Statement of cash flows",
        "Opening cash plus total movement agrees to closing cash. Column 1.",
        Decimal("100"),
        Decimal("90"),
        Decimal("1"),
    )

    assert findings
    assert "does not agree" in findings[0].issue.lower()
    assert "agrees to closing cash" not in findings[0].issue.lower()


def test_note_reference_alternative_for_current_tax_receivable_requires_current_tax_heading():
    item = reviewer.StatementNoteLine(
        statement_name="Statement of financial position",
        page_number=17,
        line="Current tax receivable 13 - 1,015",
        line_item="Current tax receivable",
        ref="13",
        amounts=(Decimal("1015"),),
        explicit_ref=True,
    )
    headings = {"13": "Current tax payable/(receivable)", "22": "Taxation"}
    sections = {
        "13": "13. Current tax payable/(receivable)\nCurrent tax receivable 1,015",
        "22": "22. Taxation\nIncome tax expense 1,015\nDeferred tax abatement 500",
    }

    alt = reviewer._alternative_note_for_missing_amounts(item, sections, headings, Decimal("1"))

    assert alt == ""


def test_revenue_consistency_skips_policy_heading_context():
    document = PdfDocument(
        [
            PdfPage(4, "Statement of profit or loss\nRevenue 349,890 249,069\n", []),
            PdfPage(
                14,
                "14. Revenue\nRevenue from contracts with customers\nElectricity sales 349,768 247,628\nConnection fees 122 1,441\nRevenue 349,890 249,069",
                [],
            ),
        ]
    )

    findings, export = check_cross_page_consistency(document)

    assert not any("revenue varies across pages" in finding.issue.lower() for finding in findings)
    assert any(row["Metric"] == "Revenue" for row in export["key_amounts"])


def test_profit_after_tax_consistency_ignores_statement_of_changes_rows():
    document = PdfDocument([
        PdfPage(4, "Directors report\nProfit for the year 760,902 (691,976)", []),
        PdfPage(13, "Statement of profit or loss\nProfit for the year 760,902 (691,976)", []),
        PdfPage(14, "Statement of changes in equity\nLoss for the year - (691,976) (691,976)\nProfit for the year - 760,902 760,902", []),
    ])

    findings, export = check_cross_page_consistency(document)

    assert not any(finding.category == "Consistency" and "Profit after tax" in finding.issue for finding in findings)
    row = next(row for row in export["key_amounts"] if row["Metric"] == "Profit after tax")
    assert row["Amount"] == "760,902"
    assert row["Pages checked"] == "Pages 4, 13"


def test_taxation_key_amount_allows_opposite_sign_rounding_difference():
    document = PdfDocument([
        PdfPage(3, "Directors report\nTaxation 19,185 -", []),
        PdfPage(12, "Statement of profit or loss\nTaxation 19 19,185 -", []),
        PdfPage(36, "19. Taxation\nCurrent tax 30,105 -\nDeferred tax (49,289) -\n(19,184) -", []),
    ])

    findings, export = check_cross_page_consistency(document)

    assert not any(finding.category == "Consistency" and "Taxation varies" in finding.issue for finding in findings)
    taxation_row = next(row for row in export["key_amounts"] if row["Metric"] == "Taxation")
    assert taxation_row["Issue"] == "Consistent"


def test_name_consistency_ignores_internal_control_headings():
    document = PdfDocument(
        [
            PdfPage(5, "Directors' report\nInternal Control\n", []),
            PdfPage(30, "Corporate governance\nInternal Controls\n", []),
        ]
    )

    findings, export = check_cross_page_consistency(document)

    assert not findings
    assert not export["names"]


def test_cash_flow_closing_check_uses_exchange_on_cash_line():
    document = PdfDocument(
        [
            PdfPage(
                19,
                "\n".join(
                    [
                        "Statement of Cash Flows",
                        "Net cash generated from/(used in) operating activities 341,897 816,212",
                        "Net cash used in investing activities (428,180) (427,963)",
                        "Total cash movement for the year (86,283) 388,249",
                        "Profit on foreign exchange on cash and cash equivalents 87,178 (385,317)",
                        "Cash and cash equivalents at the beginning of the year 56,041 53,109",
                        "Cash and cash equivalents at the end of the year 56,936 56,041",
                    ]
                ),
                [],
            )
        ]
    )

    findings, performed, skipped = check_primary_statement_consistency(document)

    assert any("opening plus movement checked to closing" in item for item in performed)
    assert not any("closing cash" in finding.issue.lower() for finding in findings)


def test_cash_flow_without_financing_section_does_not_log_false_skip():
    document = PdfDocument(
        [
            PdfPage(
                19,
                "\n".join(
                    [
                        "Statement of Cash Flows",
                        "Net cash generated from/(used in) operating activities 341,897 816,212",
                        "Net cash used in investing activities (428,180) (427,963)",
                        "Total cash movement for the year (86,283) 388,249",
                        "Cash and cash equivalents at the beginning of the year 56,041 53,109",
                        "Losses/(Gains) on foreign exchange on cash and cash equivalents 87,178 (385,317)",
                        "Cash and cash equivalents at the end of the year 56,936 56,041",
                    ]
                ),
                [],
            )
        ]
    )

    findings, performed, skipped = check_primary_statement_consistency(document)

    assert any("opening plus movement checked to closing" in item for item in performed)
    assert any("net cash increase checked" in item for item in performed)
    assert not any("operating/investing/financing/movement rows were not confidently parsed" in item for item in skipped)
    assert not findings


def test_current_tax_payable_note_matches_absolute_amount_sign():
    section = "11. Current tax payable\nCurrent tax (30,105) -"

    match = reviewer._amount_match_in_section(
        Decimal("30105"),
        section,
        Decimal("1"),
        allow_absolute=reviewer._note_heading_allows_signless_amount_match("Current tax payable", section),
    )

    assert match["found"] is True
    assert "absolute value" in str(match["method"])


def test_payable_receivable_note_matches_absolute_amount_sign():
    section = "13. Current tax payable/(receivable)\nCompany Income tax - current period - (1,015)"

    match = reviewer._amount_match_in_section(
        Decimal("1015"),
        section,
        Decimal("1"),
        allow_absolute=reviewer._note_heading_allows_signless_amount_match("Current tax payable/(receivable)", section),
    )

    assert match["found"] is True
    assert "absolute value" in str(match["method"])


def test_revenue_consistency_prefers_statement_and_note_totals_over_front_matter_outlier():
    document = PdfDocument(
        [
            PdfPage(3, "Directors' report\nRevenue 349,768 247,628", []),
            PdfPage(17, "Statement of profit or loss\nRevenue 14 349,890 249,069", []),
            PdfPage(33, "Notes to the financial statements\n14. Revenue\nRevenue from contracts with customers\nElectricity sales 349,768 247,628\nConnection fees 122 1,441\nRevenue 349,890 249,069", []),
            PdfPage(41, "Revenue 349,890 249,069", []),
            PdfPage(42, "Revenue 3A49,890 249,069 141,301 143,321 89,105", []),
        ]
    )

    findings, export = check_cross_page_consistency(document)

    assert not any("revenue varies across pages" in finding.issue.lower() for finding in findings)
    revenue_row = next(row for row in export["key_amounts"] if row["Metric"] == "Revenue")
    assert revenue_row["Amount"] == "349,890"


def test_month_year_references_are_not_flagged_as_bad_date_format():
    document = PdfDocument(
        [
            PdfPage(
                35,
                "22. Taxation\nThe Company was granted Pioneer Status in May 2025.",
                [],
            )
        ]
    )

    findings, export = check_cross_page_consistency(document)

    assert not any("preferred format" in finding.issue for finding in findings)


def test_name_consistency_suppresses_single_page_ocr_artifact_when_canonical_exists_same_page():
    document = PdfDocument(
        [
            PdfPage(2, "Directors Lai Labode", []),
            PdfPage(5, "Directors\nLait Labode Chief Executive Officer\nShareholding\nLai Labode 1,027,442", []),
            PdfPage(41, "Related parties\nLai Labode, a major shareholder and director", []),
        ]
    )

    findings, export = check_cross_page_consistency(document)

    assert not any("Lait Labode" in finding.issue for finding in findings)
    assert not any("Lait Labode" in str(row) for row in export["names"])


def test_note_narrative_contradiction_flags_nil_text_against_same_note_amount():
    document = PdfDocument(
        [
            PdfPage(
                15,
                "Notes to the Financial Statements\n"
                "3. Cash and cash equivalents\n"
                "Cash and cash equivalents at the end of the reporting period was nil.\n"
                "Cash on hand 1,000\n"
                "Bank balances 326,625\n"
                "Short-term deposits 2,588,842\n"
                "2,916,467",
                [],
            )
        ]
    )

    findings = reviewer._check_note_contradictions(document)

    assert any(finding.category == "Narrative consistency" and "nil" in finding.evidence.lower() for finding in findings)


def test_value_added_interest_expense_difference_is_review_prompt_not_casting_error():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Acme Limited\n"
                "Statement of profit or loss\n"
                "2025 2024\n"
                "Revenue 500 400\n"
                "Interest expense (737,539) -\n"
                "Profit before taxation 100 -\n"
                "Taxation (10) -\n"
                "Profit for the year 90 -",
                [],
            ),
            PdfPage(
                2,
                "Value Added Statement\n"
                "2025 2024\n"
                "Management fee income 1,219,571 -\n"
                "Interest expense (2,009,024) -\n"
                "Other operating income 1,461,391 -",
                [],
            ),
        ]
    )

    findings, performed, _skipped = check_primary_statement_consistency(document)

    assert any("Value Added Statement" in check for check in performed)
    assert any(finding.category == "Value Added Statement" and finding.severity == "Medium" for finding in findings)


def test_finance_cost_line_ignores_note_number_before_dash_amount():
    amounts, raw_line = _line_amount_for_aliases(
        "Finance costs 15 - (868)",
        ("interest expense", "finance cost", "finance costs"),
    )

    assert raw_line == "Finance costs 15 - (868)"
    assert amounts == [Decimal("0"), Decimal("-868")]


def test_simple_note_amounts_treat_leading_dash_as_zero_and_ignore_draft_footer():
    assert _simple_note_amounts_from_line("- 32,264") == [Decimal("0"), Decimal("32264")]
    assert _simple_note_amounts_from_line("94,703 472,784 27 DRAFT") == [Decimal("94703"), Decimal("472784")]


def test_ses_style_simple_note_sections_do_not_raise_false_total_findings():
    document = PdfDocument(
        [
            PdfPage(
                28,
                "\n".join(
                    [
                        "11.Other operating losses",
                        "Foreign exchange losses",
                        "Unrealised exchange gains/(losses) 1,348 (3,775)",
                        "Realised exchange gains/ (losses) (390) 13,739",
                        "958 9,964",
                        "12.Employee costs",
                        "Salaries - 26,679",
                        "Bonus - 3,905",
                        "Other short-term costs - 95",
                        "Pension costs - 1,585",
                        "- 32,264",
                        "Average number of persons employed during the year",
                        "Administration - 1",
                        "13.Depreciation expenses",
                        "Depreciation",
                        "Property, plant and equipment - 30,216",
                        "14.Operating expenses",
                        "Auditors remuneration 4,677 5,422",
                        "Bank charges - 28",
                        "Professional fees 40,985 49,806",
                        "Office expenses 352 21,487",
                        "Loss on disposal - 6,426",
                        "Restructuring expense 180 1,652",
                        "Fines and penalties * 10,113 -",
                        "Security - 503",
                        "Technical service fees ** 34,418 372,009",
                        "Telecommunication expenses 3,978 3,805",
                        "Transportation and travelling - 11,646",
                        "94,703 472,784",
                        "27",
                        "DRAFT",
                    ]
                ),
                [],
            )
        ]
    )

    findings = reviewer.check_totals_and_rounding(document)

    assert not any("Note 12 Employee costs" in finding.location and finding.severity != "Passed" for finding in findings)
    assert not any("Note 14 Operating expenses" in finding.location and finding.severity != "Passed" for finding in findings)


def test_ses_style_note_4_internal_total_is_not_flagged_from_heading_number_noise():
    findings: list[reviewer.Finding] = []
    section = "\n".join(
        [
            "4.Trade and other receivables",
            "Amount due from related parties (note 17) 260,213 225,062",
            "Prepayments 1,577 -",
            "Total trade and other receivables 261,790 225,062",
        ]
    )

    reviewer._check_note_internal_total(findings, "4", "trade and other receivables", section, Decimal("1"))

    assert not findings


def test_footnote_asterisks_are_not_unreadable_placeholders():
    document = PdfDocument(
        [
            PdfPage(
                25,
                "4. Other receivables\nPrepayments*** 55,331 -\n"
                "***This relates to payment made during the year in respect of rent.",
                [],
            )
        ]
    )

    assert document.unreadable_value_count == 0
    assert not [finding for finding in check_extraction_quality(document) if "Unreadable or placeholder" in finding.issue]




def test_signature_underscores_are_not_unreadable_placeholders():
    document = PdfDocument(
        [
            PdfPage(
                7,
                "Approved by the board\n________________________\nDirector\nStatement of financial position",
                [[["Approved by the board"], ["________________________"], ["Director"]]],
            )
        ]
    )

    assert document.unreadable_value_count == 0
    assert not [finding for finding in check_extraction_quality(document) if "Unreadable or placeholder" in finding.issue]


def test_combined_ai_package_includes_full_qc_scope_context():
    document = PdfDocument(
        [
            PdfPage(1, "Directors' Report\nThe directors present their report under CAMA 2020 and FRC requirements.", []),
            PdfPage(2, "Independent Auditor's Report\nBasis for opinion under ISA and Nigerian reporting requirements.", []),
            PdfPage(3, "Statement of cash flows\nNet cash generated from operating activities 100\nCash and cash equivalents 500", []),
            PdfPage(4, "Notes to the financial statements\n1. Significant accounting policies\nBasis of preparation under IFRS.", []),
            PdfPage(5, "Five-Year Financial Summary\nRevenue 100 90 80", []),
        ]
    )
    package = ai_combined_review._build_compact_review_package(
        document,
        CompanyProfile(company_name="Scope Test Limited", presentation_standard="IFRS"),
        {"1": "Significant accounting policies\nBasis of preparation under IFRS."},
        {"revenue": True},
        [],
        [],
        "standard",
    )
    prompt = ai_combined_review._build_prompt(package)

    assert any(row["section"] == "directors_report" for row in package["front_matter_and_other_sections"])
    assert any("cash equivalents" in row["snippet"].lower() for row in package["cash_flow_context"])
    assert any("cama" in row["snippet"].lower() for row in package["regulatory_reference_snippets"])
    assert "CAMA 2020" in prompt
    assert "review_comment_rows" in prompt
    assert "Spelling / Grammar|Regulatory Reference|Note Cross-reference|Casting|Cross-casting|Cash Flow|Disclosure|Presentation|Internal Consistency" in prompt


def test_excel_export_adds_review_comments_and_enhanced_summary(monkeypatch):
    document = PdfDocument([
        PdfPage(1, "Statement of cash flows\nNet increase in cash 100\nCash at end 500", []),
        PdfPage(2, "Notes to the financial statements\n1. Significant accounting policies", []),
    ])
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)
    monkeypatch.setattr(reviewer, "check_extraction_quality", lambda _document: [])
    monkeypatch.setattr(reviewer, "check_primary_statement_consistency", lambda _document: ([], [], []))
    monkeypatch.setattr(reviewer, "check_totals_and_rounding", lambda _document: [])
    monkeypatch.setattr(reviewer, "check_formatting", lambda _document, _profile: [])
    monkeypatch.setattr(reviewer, "check_notes_agreement", lambda _document, cautious_low_confidence=False: [])
    monkeypatch.setattr(reviewer, "_check_note_contradictions", lambda _document: [])
    monkeypatch.setattr(reviewer, "review_notes_1_and_2", lambda *args, **kwargs: ([], []))
    monkeypatch.setattr(reviewer, "check_cross_page_consistency", lambda _document: ([], {}))
    ai_finding = reviewer.Finding(
        category="AI full review",
        severity="High",
        location="Page 1 | Statement of cash flows",
        issue="Cash-flow subtotal does not agree to the reported net movement.",
        evidence="Reported net movement: 100; recomputed movement: 90; difference: 10.",
        recommendation="Correct the cash-flow subtotal or supporting line items before sign-off.",
        metadata={"page_reference": "Page 1", "statement": "Statement of cash flows", "line_item": "Net movement in cash"},
    )
    monkeypatch.setattr(
        ai_review_pipeline,
        "run_combined_ai_review",
        lambda *args, **kwargs: ai_combined_review.CombinedAiReviewResult(
            findings=list(args[4]) + [ai_finding],
            full_export=[{"Title": "Cash-flow subtotal", "Status": "exception", "Severity": "High", "Page reference": "Page 1"}],
            review_comment_rows=[
                {
                    "section_or_statement_or_note": "Statement of cash flows",
                    "page_number": "Page 1",
                    "account_or_line_item": "Net movement in cash",
                    "current_wording_amount_reference": "Reported 100; expected 90; difference 10",
                    "issue_identified": "Cash-flow subtotal does not agree to the reported net movement.",
                    "expected_correction_recommendation": "Correct the cash-flow subtotal or supporting line items.",
                    "category": "Cash Flow",
                    "priority": "High",
                    "status": "Open",
                }
            ],
            summary_fields={
                "Overall sign-off conclusion": "Not ready for final sign-off until cash-flow issue is cleared.",
                "Cash flow correctness note": "Cash-flow subtotal requires correction.",
            },
            status="completed",
            model="gpt-5-mini",
            reviewed_count=0,
            review_mode="standard",
        ),
    )

    result = review_pdf("unused.pdf", options=ReviewOptions(use_ai_full_review=True))
    workbook = openpyxl.load_workbook(BytesIO(build_excel_export(result)), data_only=True)

    assert "Review comments" in workbook.sheetnames
    header = [cell.value for cell in workbook["Review comments"][1]]
    assert header[:11] == [
        "S/N",
        "Section / Statement / Note",
        "Page number",
        "Account / line item",
        "Current wording / amount / reference",
        "Issue identified",
        "Expected correction / recommendation",
        "Category",
        "Priority",
        "Status",
        "Reviewer comments",
    ]
    rows = list(workbook["Review comments"].iter_rows(min_row=2, values_only=True))
    assert any("Cash-flow subtotal" in str(row[5] or "") and row[7] == "Cash Flow" and row[8] == "High" for row in rows)
    summary = {row[0].value: row[1].value for row in workbook["Summary"].iter_rows(min_row=2, max_col=2)}
    assert summary["Overall conclusion on final sign-off"] == "Not ready for final sign-off until cash-flow issue is cleared."
    assert summary["Category - Cash Flow"] >= 1

def test_optional_ai_policy_review_adds_findings_and_export(monkeypatch):
    filler = "Additional extracted policy context.\n" * 80
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\nCash and cash equivalents 100 90\nTotal assets 100 90\nEquity 100 90\nTotal equity and liabilities 100 90\n"
                + filler,
                [],
            ),
            PdfPage(
                2,
                "Notes to the financial statements\n1. Significant accounting policies\nRevenue from contracts with customers ...\n"
                "The entity also mentions IAS 17 as the current lease policy.\n"
                + filler,
                [],
            )
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)
    ai_finding = reviewer.Finding(
        category="AI policy judgement",
        severity="Medium",
        location="Page 1",
        issue="Lease policy appears to cite a superseded standard in current policy wording.",
        evidence="IAS 17 appears in the current lease accounting paragraph.",
        recommendation="Update the lease policy wording to IFRS 16 if that is the current basis.",
        metadata={"match_confidence": "Medium"},
    )
    monkeypatch.setattr(
        ai_review_pipeline,
        "run_combined_ai_review",
        lambda *args, **kwargs: ai_combined_review.CombinedAiReviewResult(
            findings=list(args[4]) + [ai_finding],
            policy_export=[{"Title": "Lease policy context", "Status": "review_prompt"}],
            summary="Lease policy wording may still reference a superseded standard.",
            status="completed",
            model="gpt-5-mini",
            evidence_rows=[{"Evidence type": "Combined AI compact review package", "Snippet": "IAS 17 current lease policy"}],
            review_mode="standard",
        ),
    )

    result = review_pdf("unused.pdf", options=ReviewOptions(use_ai_policy_review=True))

    assert any(f.category == "AI policy judgement" for f in result.findings)
    assert result.metrics["ai_policy_review_status"] == "completed"
    assert result.metrics["ai_policy_review_model"] == "gpt-5-mini"
    assert result.metrics["ai_policy_export"] == [{"Title": "Lease policy context", "Status": "review_prompt"}]
    assert result.metrics["ai_evidence_packs"][0]["Evidence type"] == "Combined AI compact review package"
    assert "Combined AI review completed using gpt-5-mini" in result.metrics["checks_performed"]

def test_optional_ai_full_review_adds_findings_and_export(monkeypatch):
    filler = "Additional extracted review context.\n" * 80
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\nCash and cash equivalents 100 90\nTotal assets 100 90\nEquity 100 90\nTotal equity and liabilities 100 90\n"
                + filler,
                [],
            ),
            PdfPage(
                2,
                "Notes to the financial statements\n1. Significant accounting policies\nRevenue from contracts with customers.\n" + filler,
                [],
            ),
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)
    ai_finding = reviewer.Finding(
        category="AI full review",
        severity="Medium",
        location="Page 2 | Note 1",
        issue="Revenue policy wording needs reviewer confirmation.",
        evidence="AI reviewed supplied policy context and found a possible tailoring issue.",
        recommendation="Review Note 1 and tailor the revenue wording where required.",
        metadata={"match_confidence": "Medium", "page_reference": "Page 2", "note_reference": "Note 1"},
    )
    monkeypatch.setattr(
        ai_review_pipeline,
        "run_combined_ai_review",
        lambda *args, **kwargs: ai_combined_review.CombinedAiReviewResult(
            findings=list(args[4]) + [ai_finding],
            full_export=[{"Title": "Revenue policy", "Status": "review_prompt", "Page reference": "Page 2"}],
            summary="Full AI review found one policy tailoring prompt.",
            executive_memo="Full AI review found one policy tailoring prompt.",
            status="completed",
            model="gpt-5-mini",
            evidence_rows=[{"Evidence type": "Combined AI compact review package", "Source page": "Document extracted pages"}],
            review_mode="standard",
        ),
    )

    result = review_pdf("unused.pdf", options=ReviewOptions(use_ai_full_review=True))

    assert any(f.category == "AI full review" for f in result.findings)
    assert result.metrics["ai_full_review_status"] == "completed"
    assert result.metrics["ai_full_review_model"] == "gpt-5-mini"
    assert result.metrics["ai_full_export"] == [{"Title": "Revenue policy", "Status": "review_prompt", "Page reference": "Page 2"}]
    assert result.metrics["ai_evidence_packs"][0]["Evidence type"] == "Combined AI compact review package"
    assert "Combined AI review completed using gpt-5-mini" in result.metrics["checks_performed"]

def test_ai_review_pipeline_runs_combined_review_once(monkeypatch):
    document = PdfDocument([PdfPage(1, "Notes to the financial statements\n1. Significant accounting policies", [])])
    profile = CompanyProfile(company_name="Test Limited")
    base_finding = reviewer.Finding("Formatting", "Low", "Page 1", "Weak issue", "Evidence", "Review")
    calls = []

    monkeypatch.setattr(ai_review_pipeline, "_AI_PIPELINE_LOCK", __import__("threading").Lock())
    monkeypatch.setattr(
        ai_review_pipeline,
        "run_combined_ai_review",
        lambda *args, **kwargs: calls.append("combined") or ai_combined_review.CombinedAiReviewResult(
            findings=list(args[4]),
            summary="combined ok",
            status="completed",
            model="gpt-5-mini",
            review_mode="standard",
        ),
    )

    result = ai_review_pipeline.run_ai_review_pipeline(
        ai_review_pipeline.AiReviewContext(
            document=document,
            profile=profile,
            note_sections={"1": "Significant accounting policies"},
            policy_map={"revenue": True},
            findings=[base_finding],
            model="gpt-5-mini",
            use_policy_review=True,
            use_full_review=True,
        )
    )

    assert calls == ["combined"]
    assert result.policy_status == "completed"
    assert result.full_status == "completed"
    assert result.finding_status == "completed"
    assert result.findings == [base_finding]

def test_combined_ai_failure_still_allows_finding_cleanup(monkeypatch):
    filler = "Additional extracted review context.\n" * 80
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\nCash and cash equivalents 100 90\nTotal assets 100 90\nEquity 100 90\nTotal equity and liabilities 100 90\n"
                + filler,
                [],
            ),
            PdfPage(
                2,
                "Notes to the financial statements\n1. Significant accounting policies\nRevenue from contracts with customers.\n" + filler,
                [],
            ),
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)
    calls = []

    def fake_combined(*args, **kwargs):
        calls.append("combined")
        return ai_combined_review.CombinedAiReviewResult(
            findings=list(args[4]),
            status="deferred",
            model="gpt-5-mini",
            message="AI review was not completed after automatic retry attempts because the AI service remained temporarily busy.",
            error_rows=[
                {
                    "Stage": "combined_ai_review",
                    "Model": "gpt-5-mini",
                    "Error category": "rate_limit",
                    "Error message": "429 rate limit",
                    "Retry count": "5",
                }
            ],
            review_mode="standard",
        )

    def fake_cleanup(*args, **kwargs):
        calls.append("cleanup")
        return ai_finding_review.AiFindingReviewResult(
            findings=list(args[2]),
            export_rows=[],
            summary="Finding cleanup still ran after combined review failure.",
            status="completed",
            model="gpt-5-mini",
            message="",
            reviewed_count=0,
            suppressed_count=0,
            evidence_rows=[],
        )

    monkeypatch.setattr(ai_review_pipeline, "run_combined_ai_review", fake_combined)
    monkeypatch.setattr(ai_review_pipeline, "run_ai_finding_review", fake_cleanup)

    result = review_pdf("unused.pdf", options=ReviewOptions(use_ai_policy_review=True, use_ai_full_review=True))

    assert calls == ["combined", "cleanup"]
    assert result.metrics["ai_policy_review_status"] == "deferred"
    assert result.metrics["ai_full_review_status"] == "deferred"
    assert result.metrics["ai_finding_review_status"] == "completed"
    assert result.metrics["ai_review_status"] == "Failed after retries / Not completed"
    assert result.metrics["ai_error_log"][0]["Error category"] == "rate_limit"
    assert any("automatic retry attempts" in line for line in result.metrics["checks_skipped"].split("\n"))


def test_combined_ai_failure_logs_real_error_without_counting_as_finding(monkeypatch):
    filler = "Additional extracted review context.\n" * 80
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\nCash and cash equivalents 100 90\nTotal assets 100 90\nEquity 100 90\nTotal equity and liabilities 100 90\n"
                + filler,
                [],
            ),
            PdfPage(
                2,
                "Notes to the financial statements\n1. Significant accounting policies\nRevenue from contracts with customers.\n" + filler,
                [],
            ),
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)
    monkeypatch.setattr(
        ai_review_pipeline,
        "run_combined_ai_review",
        lambda *args, **kwargs: ai_combined_review.CombinedAiReviewResult(
            findings=list(args[4]),
            status="error",
            model="gpt-5-mini",
            message="AI review failed after retries; see AI debug details for the provider error.",
            error_rows=[
                {
                    "Stage": "combined_ai_review",
                    "Exception type": "BadRequestError",
                    "Status code": "400",
                    "Error category": "payload_too_large",
                    "Error message": "context length exceeded",
                    "Model": "gpt-5-mini",
                    "Input token estimate": "200000",
                    "Output token limit": "3500",
                    "Retry count": "1",
                }
            ],
            review_mode="standard",
        ),
    )
    monkeypatch.setattr(
        ai_review_pipeline,
        "run_ai_finding_review",
        lambda *args, **kwargs: ai_finding_review.AiFindingReviewResult(
            findings=list(args[2]),
            export_rows=[],
            summary="",
            status="skipped",
            model="gpt-5-mini",
            message="AI finding cleanup skipped after provider payload error.",
            reviewed_count=0,
            suppressed_count=0,
            evidence_rows=[],
        ),
    )

    result = review_pdf("unused.pdf", options=ReviewOptions(use_ai_full_review=True))

    assert result.metrics["ai_full_review_status"] == "error"
    assert result.metrics["ai_review_status"] == "Failed after retries / Not completed"
    assert result.metrics["ai_error_log"][0]["Error category"] == "payload_too_large"
    assert result.metrics["ai_error_log"][0]["Error message"] == "context length exceeded"
    assert not any("AI review failed" in finding.issue for finding in result.findings)


def test_retry_ai_review_uses_cached_deterministic_result_without_rerunning_checks(monkeypatch):
    document = PdfDocument([
        PdfPage(1, "Statement of financial position\nCash and cash equivalents 100 90\nTotal assets 100 90\nEquity 100 90\nTotal equity and liabilities 100 90", []),
        PdfPage(2, "Notes to the financial statements\n1. Significant accounting policies\nRevenue from contracts with customers", []),
    ])
    deterministic_finding = reviewer.Finding(
        category="Formatting",
        severity="Low",
        location="Page 1",
        issue="Weak deterministic issue.",
        evidence="Evidence.",
        recommendation="Review.",
    )
    cached_result = reviewer.ReviewResult(
        findings=[deterministic_finding],
        metrics={
            "checks_performed": "Statement of financial position totals checked\nCombined AI review completed using gpt-5-mini in Standard mode; 0 deterministic finding(s) reviewed and 0 suppressed.",
            "checks_skipped": "AI review was not completed after automatic retry attempts.\nDetailed note agreement skipped because table extraction confidence is below threshold.",
            "cross_page_export": {},
            "policy_export": [],
        },
    )
    monkeypatch.setattr(reviewer, "check_formatting", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("deterministic checks should not rerun")))
    monkeypatch.setattr(
        ai_review_pipeline,
        "run_combined_ai_review",
        lambda *args, **kwargs: ai_combined_review.CombinedAiReviewResult(
            findings=list(args[4]),
            status="completed",
            model="gpt-5-mini",
            reviewed_count=1,
            suppressed_count=0,
            summary="Retry completed.",
            review_mode="standard",
        ),
    )

    result = reviewer.rerun_ai_review_from_cached_result(
        document,
        "unused.pdf",
        CompanyProfile(company_name="Test Limited"),
        ReviewOptions(use_ai_policy_review=True, use_ai_full_review=True),
        cached_result,
    )

    assert result.findings[0].issue == "Weak deterministic issue."
    assert "Statement of financial position totals checked" in result.metrics["checks_performed"]
    assert "Combined AI review completed using gpt-5-mini" in result.metrics["checks_performed"]
    assert "AI review was not completed after automatic retry attempts" not in result.metrics["checks_skipped"]
    assert "Detailed note agreement skipped" in result.metrics["checks_skipped"]

def test_ai_policy_review_missing_key_is_reported_as_skipped(monkeypatch):
    filler = "Additional extracted policy context.\n" * 80
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\nCash and cash equivalents 100 90\nTotal assets 100 90\nEquity 100 90\nTotal equity and liabilities 100 90\n"
                + filler,
                [],
            ),
            PdfPage(
                2,
                "Notes to the financial statements\n1. Significant accounting policies\nRevenue from contracts with customers ...\n"
                + filler,
                [],
            )
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)
    monkeypatch.delenv("OPENAI_API_KEY", raising=False)

    result = review_pdf("unused.pdf", options=ReviewOptions(use_ai_policy_review=True))

    assert result.metrics["ai_policy_review_status"] == "unavailable"
    assert "OPENAI_API_KEY is not configured" in result.metrics["checks_skipped"]


def test_note_heading_regex_accepts_compact_numeric_headings():
    assert ai_finding_review.NOTE_HEADING_RE.match("3.Cash and cash equivalents")
    assert reviewer.NOTE_HEADING_RE.match("3.Cash and cash equivalents")


def test_ai_finding_review_keeps_explicit_nil_vs_non_zero_note_contradiction():
    finding = reviewer.Finding(
        category="Narrative consistency",
        severity="Medium",
        location="Page 24 | Note 3",
        issue="Note 3 states that a balance was nil, but the same note table shows a non-zero amount.",
        evidence="Cash and cash equivalents at the end of the end of the reporting period was nil. | Note heading: Cash and cash equivalents | Non-zero amount detected in same note: 2,916,467",
        recommendation="Review the narrative disclosure against the note table and correct the wording or the amount presentation.",
    )
    candidates = [{
        "finding_id": "F1",
        "index": 0,
        "category": finding.category,
        "severity": finding.severity,
        "page_reference": "Page 24",
        "note_reference": "Note 3",
        "issue": finding.issue,
        "page_snippet": "Page 24: 3. Cash and cash equivalents Cash and cash equivalents at the end of the end of the reporting period was nil.",
        "note_snippet": "3.Cash and cash equivalents Cash and cash equivalents at the end of the end of the reporting period was nil. 2,916,467 -",
    }]
    parsed = {
        "summary": "test",
        "adjudications": [{
            "finding_id": "F1",
            "decision": "suppress",
            "revised_severity": "Low",
            "status": "likely_false_positive",
            "confidence": "Medium",
            "reason": "layout mismatch",
            "recommended_action": "none",
        }],
    }

    result = ai_finding_review._apply_adjudications([finding], candidates, parsed, "gpt-5-mini")

    assert len(result.findings) == 1
    assert result.findings[0].severity == "Medium"
    assert result.export_rows[0]["AI status"] == "confirmed_exception"


def test_ai_finding_review_does_not_suppress_confirmed_note_mismatch():
    finding = reviewer.Finding(
        category="Notes agreement",
        severity="Medium",
        location="Page 37 | Note 11",
        issue="Amount not located in referenced note for Current tax payable A.",
        evidence="Referenced note: Note 11. Current-year amount 30,105 was not found in the note section.",
        recommendation="Review the note reference and the related tax disclosure.",
        metadata={"note_reference": "Note 11", "page_reference": "Page 37"},
    )
    candidates = [{
        "finding_id": "F1",
        "index": 0,
        "category": finding.category,
        "severity": finding.severity,
        "page_reference": "Page 37",
        "note_reference": "Note 11",
        "issue": finding.issue,
        "page_snippet": "Current tax payable A 11 30,105 -",
        "note_snippet": "Note 11 Going concern The company has adequate support.",
    }]
    parsed = {
        "summary": "test",
        "adjudications": [{
            "finding_id": "F1",
            "decision": "suppress",
            "revised_severity": "Low",
            "status": "likely_false_positive",
            "confidence": "High",
            "reason": "Note 11 only contains going concern wording and does not contain the tax amount, so this is a note linking issue.",
            "recommended_action": "Check the tax note reference.",
        }],
    }

    result = ai_finding_review._apply_adjudications([finding], candidates, parsed, "gpt-4.1")

    assert len(result.findings) == 1
    assert result.suppressed_count == 0
    assert result.findings[0].severity == "Medium"
    assert result.export_rows[0]["Decision"] != "Suppress"
    assert result.export_rows[0]["AI status"] == "confirmed_exception"


def test_ai_policy_evidence_rows_are_page_and_topic_specific():
    document = PdfDocument([
        PdfPage(1, "Statement of profit or loss Revenue 100", []),
        PdfPage(2, "Notes to the financial statements\n1. Significant accounting policies\nRevenue from contracts with customers is recognised over time.", []),
        PdfPage(3, "5. Property, plant and equipment\nPlant and equipment are depreciated over useful lives.", []),
    ])
    rows = ai_policy_review._policy_evidence_rows(
        {
            "1": "1. Significant accounting policies\nRevenue from contracts with customers is recognised over time.",
            "5": "5. Property, plant and equipment\nPlant and equipment are depreciated over useful lives.",
        },
        document,
        {"revenue": True, "ppe": True, "leases": True},
    )

    assert rows[0]["Page reference"] == "Page 2"
    assert rows[0]["Detected topics"] == "revenue"
    assert rows[1]["Page reference"] == "Page 3"
    assert "ppe" in rows[1]["Detected topics"]
    assert "leases" not in rows[0]["Detected topics"]


def test_ai_policy_rows_downgrade_weak_consolidation_and_generic_lease_observations():
    findings, export_rows = ai_policy_review._rows_to_outputs([
        {
            "title": "Consolidation policy",
            "dimension": "policy_relevance",
            "standard_or_topic": "IFRS 10 consolidation",
            "severity": "High",
            "confidence": "High",
            "status": "exception",
            "issue": "Consolidation policy may be missing based on related party sister company wording.",
            "rationale": "Evidence only refers to related parties and sister companies.",
            "recommendation": "Review group structure.",
            "page_reference": "Page 28",
            "note_reference": "Note 20",
            "evidence_snippet": "Related party balances with a sister company were disclosed.",
        },
        {
            "title": "Lease policy",
            "dimension": "standard_context",
            "standard_or_topic": "IFRS 16",
            "severity": "Medium",
            "confidence": "Medium",
            "status": "review_prompt",
            "issue": "Generic IFRS 16 amendment text mentions recognition of a lease asset and lease liability.",
            "rationale": "The wording appears in a new standards/amendments section, not an actual balance note.",
            "recommendation": "Do not elevate unless a lease balance exists.",
            "page_reference": "Page 12",
            "note_reference": "Note 2",
            "evidence_snippet": "New standards amendments refer to recognition of a lease asset and lease liability.",
        },
    ])

    assert [finding.severity for finding in findings] == ["Low", "Low"]
    assert [row["Confidence"] for row in export_rows] == ["Low", "Low"]


def test_optional_ai_finding_review_suppresses_low_false_positive_and_exports_sheet(monkeypatch):
    document = PdfDocument([
        PdfPage(1, "Statement of financial position\nCash and cash equivalents 100 90", []),
        PdfPage(2, "Notes to the financial statements\n1. Significant accounting policies", []),
    ])
    low_finding = reviewer.Finding(
        category="Formatting",
        severity="Low",
        location="Page 1",
        issue="Possible formatting inconsistency detected in a weak text fragment.",
        evidence="Weak fragment only.",
        recommendation="Review manually.",
    )

    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)
    monkeypatch.setattr(reviewer, "check_extraction_quality", lambda _document: [])
    monkeypatch.setattr(reviewer, "check_primary_statement_consistency", lambda _document: ([], [], []))
    monkeypatch.setattr(reviewer, "check_totals_and_rounding", lambda _document: [])
    monkeypatch.setattr(reviewer, "check_formatting", lambda _document, _profile: [low_finding])
    monkeypatch.setattr(reviewer, "check_notes_agreement", lambda _document, cautious_low_confidence=False: [])
    monkeypatch.setattr(reviewer, "_check_note_contradictions", lambda _document: [])
    monkeypatch.setattr(reviewer, "review_notes_1_and_2", lambda *args, **kwargs: ([], []))
    monkeypatch.setattr(reviewer, "check_cross_page_consistency", lambda _document: ([], {}))
    monkeypatch.setattr(
        ai_review_pipeline,
        "run_combined_ai_review",
        lambda *args, **kwargs: ai_combined_review.CombinedAiReviewResult(
            findings=[],
            finding_export=[
                {
                    "Finding ID": "F1",
                    "Category": "Formatting",
                    "Original severity": "Low",
                    "Decision": "Suppress",
                    "Revised severity": "Low",
                    "AI status": "likely_false_positive",
                    "AI confidence": "High",
                    "Page reference": "Page 1",
                    "Note reference": "",
                    "Issue": low_finding.issue,
                    "Reason": "The evidence looks like extraction noise rather than a report defect.",
                    "Recommended action": "Do not elevate unless corroborated by the source PDF.",
                }
            ],
            summary="One weak formatting finding was suppressed as a likely false positive.",
            status="completed",
            model="gpt-5-mini",
            reviewed_count=1,
            suppressed_count=1,
            evidence_rows=[{"Evidence type": "Engine finding adjudication pack", "Finding ID": "F1", "Issue": low_finding.issue}],
            suppressed_rows=[
                {
                    "Finding ID": "F1",
                    "Category": "Formatting",
                    "Decision": "Suppress",
                    "AI status": "likely_false_positive",
                    "Issue": low_finding.issue,
                    "Reason": "The evidence looks like extraction noise rather than a report defect.",
                }
            ],
            review_mode="standard",
        ),
    )

    result = review_pdf("unused.pdf", options=ReviewOptions(use_ai_policy_review=True))

    assert result.findings == []
    assert result.metrics["ai_finding_review_status"] == "completed"
    assert result.metrics["ai_finding_suppressed"] == 1
    assert "Combined AI review completed using gpt-5-mini in Standard mode; 1 deterministic finding(s) reviewed and 1 suppressed." in result.metrics["checks_performed"]

    workbook = openpyxl.load_workbook(BytesIO(build_excel_export(result)), data_only=True)
    assert "AI finding review" in workbook.sheetnames
    assert "AI evidence packs" in workbook.sheetnames
    assert "AI suppressed findings" in workbook.sheetnames
    ai_sheet = workbook["AI finding review"]
    rows = list(ai_sheet.iter_rows(min_row=2, values_only=True))
    assert rows
    assert any(str(row[2] or "").lower() == "low" for row in rows)
    suppressed_sheet = workbook["AI suppressed findings"]
    suppressed_rows = list(suppressed_sheet.iter_rows(min_row=2, values_only=True))
    assert any("Formatting" in [str(cell or "") for cell in row] for row in suppressed_rows)
    evidence_sheet = workbook["AI evidence packs"]
    evidence_rows = list(evidence_sheet.iter_rows(min_row=2, values_only=True))
    assert any("Engine finding adjudication pack" in str(row[0] or "") for row in evidence_rows)

def test_ai_response_parser_accepts_nested_text_value_payload():
    payload = {
        "output": [
            {
                "content": [
                    {
                        "type": "text",
                        "text": {
                            "value": '{"summary":"ok","observations":[{"title":"Policy context","status":"review_prompt"}]}'
                        },
                    }
                ]
            }
        ]
    }

    parsed = _parse_response_json(payload)

    assert parsed["summary"] == "ok"
    assert parsed["observations"][0]["title"] == "Policy context"


def test_ai_response_parser_accepts_fenced_json_with_surrounding_text():
    payload = {
        "output_text": 'Here is the result:\n```json\n{"summary":"ok","observations":[]}\n```\nDone.'
    }

    parsed = _parse_response_json(payload)

    assert parsed == {"summary": "ok", "observations": []}


def test_ai_policy_review_repairs_malformed_json_response(monkeypatch):
    monkeypatch.setenv("OPENAI_API_KEY", "test-key")
    calls = []

    def fake_call(_api_key, payload):
        calls.append(payload)
        if len(calls) == 1:
            return {
                "output_text": (
                    '{"summary":"needs review","observations":[{"title":"Revenue policy","status":"review_prompt"}'
                    '{"title":"Tax policy","status":"ok"}]}'
                )
            }
        return {
            "output_text": (
                '{"summary":"needs review","observations":[{"title":"Revenue policy","status":"review_prompt",'
                '"severity":"Low","confidence":"Medium","issue":"Revenue wording should be reviewed",'
                '"rationale":"Policy text is present but may need tailoring","recommendation":"Review revenue wording",'
                '"page_reference":"Page 10","note_reference":"Note 1","evidence_snippet":"Revenue from contracts with customers"}]}'
            )
        }

    monkeypatch.setattr(ai_policy_review, "_call_openai", fake_call)
    document = PdfDocument([PdfPage(10, "Notes to the financial statements\n1 Significant accounting policies\nRevenue from contracts with customers", [])])

    result = ai_policy_review.run_ai_policy_review(
        document,
        CompanyProfile(company_name="Test Limited", industry="Technology"),
        {"1": "Significant accounting policies. Revenue from contracts with customers."},
        policy_map={"revenue": True},
        model="gpt-test",
    )

    assert result.status == "completed"
    assert len(calls) == 2
    assert result.export_rows
    assert "Revenue" in result.export_rows[0]["Title"]


def test_ai_finding_review_keeps_amount_related_change_without_required_evidence():
    finding = reviewer.Finding(
        category="Notes agreement",
        severity="Medium",
        location="Page 10",
        issue="Amount not located in referenced note.",
        evidence="Statement amount was not found in note snippet.",
        recommendation="Review note reference.",
    )
    candidate = {
        "finding_id": "F1",
        "index": 0,
        "category": finding.category,
        "severity": finding.severity,
        "page_reference": "Page 10",
        "note_reference": "Note 14",
        "issue": finding.issue,
        "evidence": finding.evidence,
        "page_snippet": "",
        "note_snippet": "",
    }
    parsed = {
        "summary": "AI reviewed one finding",
        "adjudications": [
            {
                "finding_id": "F1",
                "decision": "keep",
                "status": "confirmed_exception",
                "confidence": "Medium",
                "reason": "This is likely false positive because no supporting table is visible in the supplied snippets.",
            }
        ],
    }

    result = ai_finding_review._apply_adjudications([finding], [candidate], parsed, "gpt-test")

    assert len(result.findings) == 1
    assert result.suppressed_count == 0
    assert result.export_rows[0]["Decision"] == "Keep"
    assert "required amount evidence" in result.export_rows[0]["Guardrail applied"]

def test_ai_finding_review_keeps_strong_narrative_contradiction_when_ai_is_weak():
    finding = reviewer.Finding(
        category="Narrative consistency",
        severity="Medium",
        location="Page 24",
        issue="Note 3 states that a balance was nil, but the same note table shows a non-zero amount.",
        evidence="Cash and cash equivalents was nil. | Non-zero amount detected in same note: 2,916,467",
        recommendation="Review the narrative disclosure against the note table.",
    )
    candidate = {
        "finding_id": "F1",
        "index": 0,
        "category": finding.category,
        "severity": finding.severity,
        "page_reference": "Page 24",
        "note_reference": "Note 3",
        "issue": finding.issue,
        "page_snippet": "",
        "note_snippet": "",
    }
    parsed = {
        "adjudications": [
            {
                "finding_id": "F1",
                "decision": "keep",
                "status": "insufficient_evidence",
                "confidence": "Low",
                "reason": "The evidence is ambiguous and reviewer attention is needed to confirm if this is a true inconsistency.",
            }
        ]
    }

    result = ai_finding_review._apply_adjudications([finding], [candidate], parsed, "gpt-test")

    assert len(result.findings) == 1
    assert result.findings[0].severity == "Medium"
    assert result.export_rows[0]["Decision"] == "Keep"
    assert result.export_rows[0]["AI status"] == "confirmed_exception"
    assert "explicit nil-style narrative" in result.export_rows[0]["Reason"]


def test_ai_finding_review_includes_page_and_note_context_and_can_suppress_medium():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of financial position",
                        "Current assets 1 1,172,176 811,598",
                        "Total assets 3,548,638 2,926,848",
                        "10",
                    ]
                ),
                [],
            ),
            PdfPage(
                2,
                "\n".join(
                    [
                        "Notes to the financial statements",
                        "18 Cash and cash equivalents",
                        "Access Bank Plc 5,518 13,691",
                        "TOTAL 875,869 605,645",
                        "28",
                    ]
                ),
                [],
            ),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    finding = reviewer.Finding(
        category="Totals and rounding",
        severity="Medium",
        location="Page 28 | Note 18 Cash and cash equivalents",
        issue="Simple note section total does not agree to visible component rows.",
        evidence="Note 18 Cash and cash equivalents, column 1: reported 875,869, visible sum 175,252, difference 700,617",
        recommendation="Recalculate the note section.",
        metadata={"page_reference": "Page 28", "note_reference": "Note 18", "ocr_review": "true"},
    )

    candidates = ai_finding_review._review_candidates(document, [finding])

    assert len(candidates) == 1
    assert "18 Cash and cash equivalents" in candidates[0]["note_snippet"]

    result = ai_finding_review._apply_adjudications(
        [finding],
        candidates,
        {
            "summary": "One ambiguous note-total finding was suppressed as OCR drift.",
            "adjudications": [
                {
                    "finding_id": "F1",
                    "decision": "suppress",
                    "revised_severity": "Low",
                    "status": "likely_false_positive",
                    "confidence": "High",
                    "reason": "The page and note snippets show a split-digit extraction pattern rather than a genuine note-total defect.",
                    "recommended_action": "Do not elevate unless the signed PDF still disagrees after manual recast.",
                    "amount_evidence": {
                        "page_number": "Page 28",
                        "statement_or_note_name": "Note 18 Cash and cash equivalents",
                        "reported_amount": "875,869",
                        "expected_amount": "175,252",
                        "difference": "700,617",
                        "evidence": "The page and note snippets show a split-digit extraction pattern rather than a genuine note-total defect.",
                    },
                }
            ],
        },
        "gpt-5-mini",
    )

    assert result.status == "completed"
    assert result.findings == []
    assert result.suppressed_count == 1
    assert result.export_rows[0]["Note snippet"].startswith("18 Cash and cash equivalents")


def test_ai_finding_review_attaches_source_pdf_when_available(monkeypatch, tmp_path):
    monkeypatch.setenv("OPENAI_API_KEY", "test-key")
    pdf_path = tmp_path / "sample.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n%tiny test pdf\n")
    captured_payloads = []

    def fake_call(_api_key, payload):
        captured_payloads.append(payload)
        return {
            "output_text": (
                '{"summary":"reviewed","adjudications":[{"finding_id":"F1","decision":"keep",'
                '"revised_severity":"Low","status":"review_prompt","confidence":"Medium",'
                '"reason":"Evidence remains suitable for reviewer follow-up.",'
                '"recommended_action":"Review the page."}]}'
            )
        }

    monkeypatch.setattr(ai_finding_review, "_call_openai", fake_call)
    document = PdfDocument([PdfPage(1, "Statement of financial position\nFormatting issue context", [])])
    finding = reviewer.Finding(
        category="Formatting",
        severity="Low",
        location="Page 1",
        issue="Possible formatting issue.",
        evidence="Formatting issue context.",
        recommendation="Review formatting.",
    )

    result = ai_finding_review.run_ai_finding_review(
        document,
        CompanyProfile(company_name="Sample Limited"),
        [finding],
        model="gpt-test",
        pdf_path=pdf_path,
    )

    assert result.status == "completed"
    user_content = captured_payloads[0]["input"][1]["content"]
    assert any(item.get("type") == "input_file" for item in user_content)
    source_rows = [row for row in (result.evidence_rows or []) if row.get("Evidence type") == "Original PDF"]
    assert source_rows and source_rows[0]["Status"] == "Attached"

def test_narrative_dates_do_not_trigger_format_findings():
    document = PdfDocument(
        [
            PdfPage(
                4,
                "Financial Statements for the year ended 31 December 2025\n"
                "The Company was incorporated on 7th May 2019 as a private limited liability company.",
                [],
            ),
            PdfPage(
                36,
                "22. Contingencies\n"
                "The solicitors confirmed that the Company did not have pending legal cases as at 31st December 2025.",
                [],
            ),
        ]
    )

    findings, export = check_cross_page_consistency(document)
    assert not [row for row in export["dates"] if row.get("Comment") == "Inconsistent date format."]
    assert not [finding for finding in findings if finding.category == "Formatting" and "preferred format" in finding.issue]


def test_month_day_year_without_comma_is_preferred_date_format():
    document = PdfDocument(
        [
            PdfPage(
                5,
                "Financial Statements for the year ended December 31 2025\n"
                "Approved by the board on April 4 2026.",
                [],
            )
        ]
    )

    findings, export = check_cross_page_consistency(document)

    assert not [row for row in export["dates"] if row.get("Comment") == "Inconsistent date format."]
    assert not [finding for finding in findings if finding.category == "Formatting" and "preferred format" in finding.issue]


def test_cross_page_consistency_flags_obvious_repeated_word_grammar_issue():
    document = PdfDocument(
        [
            PdfPage(
                12,
                "Notes to the financial statements\n"
                "The company company did not provide any services other than the statutory audit.",
                [],
            )
        ]
    )

    findings, export = check_cross_page_consistency(document)

    assert any(row["Issue"] == "Repeated word detected." for row in export["grammar"])
    assert any("Possible grammatical or drafting issue detected." == finding.issue for finding in findings)


def test_cross_page_consistency_flags_common_spelling_error():
    document = PdfDocument(
        [
            PdfPage(
                18,
                "The deffered tax balance was reviewed by management and agreed to the supporting schedule.",
                [],
            )
        ]
    )

    findings, export = check_cross_page_consistency(document)

    assert any("Possible spelling error" in row["Issue"] for row in export["grammar"])
    assert any(finding.issue == "Possible spelling issue detected." for finding in findings)






def test_cross_page_consistency_ignores_table_header_repeated_words():
    document = PdfDocument(
        [
            PdfPage(
                8,
                "Statement of changes in equity\nAccumulated fund Donation fund fund Total equity",
                [],
            )
        ]
    )

    findings, export = check_cross_page_consistency(document)

    assert not export["grammar"]
    assert not [finding for finding in findings if finding.category == "Formatting"]


def test_cross_page_consistency_ignores_auditor_signature_spacing_noise():
    document = PdfDocument(
        [
            PdfPage(
                12,
                "Independent auditor's report\nFor:Kreston Pedabo Audit services\nChartered Accountants",
                [],
            )
        ]
    )

    findings, export = check_cross_page_consistency(document)

    assert not export["grammar"]
    assert not [finding for finding in findings if finding.category == "Formatting"]




def test_key_amount_consistency_picks_up_follow_on_tax_heading_totals():
    document = PdfDocument(
        [
            PdfPage(
                4,
                "Directors' report\n"
                "Taxation (46,581) (43,638)\n",
                [],
            ),
            PdfPage(
                15,
                "Statement of Profit or Loss\n"
                "Taxation 23 (46,581) (43,638)\n",
                [],
            ),
            PdfPage(
                40,
                "23. Taxation\n"
                "Major components of the tax expense\n"
                "Current\n"
                "46,236 36,721\n"
                "Deferred\n"
                "345 6,917\n"
                "46,581 43,638\n",
                [],
            ),
        ]
    )

    findings, export = check_cross_page_consistency(document)

    assert not findings
    taxation_row = next(row for row in export["key_amounts"] if row["Metric"] == "Taxation")
    assert taxation_row["Pages checked"] == "Pages 4, 15, 40"
    assert "Page 40" in taxation_row["Context"]

    assert not [finding for finding in findings if finding.category == "Formatting"]
    assert not [row for row in export["dates"] if row.get("Comment") == "Inconsistent date format."]


def test_excel_name_export_guard_suppresses_one_page_ocr_artifact():
    from export_utils import clean_name_consistency_rows

    rows = [
        {
            "Name variant 1": "Lai Labode",
            "Page 1": "41, 2, 5, 6",
            "Name variant 2": "Lait Labode",
            "Page 2": "5",
            "Suggested standard spelling": "Lait Labode",
        },
        {
            "Name variant 1": "Mzer Michael Terungwa",
            "Page 1": "5",
            "Name variant 2": "Mzer Micheal Terungwa",
            "Page 2": "41",
            "Suggested standard spelling": "Mzer Michael Terungwa",
        },
    ]

    cleaned = clean_name_consistency_rows(rows)

    assert len(cleaned) == 1
    assert cleaned[0]["Name variant 1"] == "Mzer Michael Terungwa"


def test_key_amount_consistency_summarizes_consistent_rows_without_long_context():
    document = PdfDocument(
        [
            PdfPage(4, "Revenue 8,398,634 7,336,635", []),
            PdfPage(15, "Revenue 13 8,398,634 7,336,635", []),
            PdfPage(44, "Revenue 8,398,634 7,336,635", []),
        ]
    )

    findings, export = check_cross_page_consistency(document)

    assert findings == []
    assert export["key_amounts"] == [
        {
            "Metric": "Revenue",
            "Amount": "8,398,634",
            "Pages checked": "Pages 4, 15, 44",
            "Context": "Consistent across detected occurrences.",
            "Issue": "Consistent",
        }
    ]


def test_amount_match_snippet_returns_complete_extracted_row():
    text = "7 Operating Revenue\nSubscriptions 2 ,783,064\nPrior year 2\n,029,846"
    start = text.index("2 ,783,064")
    snippet = reviewer._amount_snippet_around(text, start, start + len("2 ,783,064"))

    assert snippet == "Subscriptions 2 ,783,064"


def test_page_reference_extracts_page_lists_from_inconsistency_evidence():
    document = PdfDocument(
        [
            PdfPage(5, "Directors' report\nMzer Michael Terungwa", []),
            PdfPage(41, "Directors' report\nMzer Micheal Terungwa", []),
        ]
    )

    findings, _export = check_cross_page_consistency(document)
    name_finding = next(finding for finding in findings if "Name spelt differently" in finding.issue)

    assert name_finding.location == "Pages 5, 41"


def test_excel_export_page_reference_uses_detected_note_pages():
    app_source = Path("app.py").read_text(encoding="utf-8")

    assert "_page_reference_for_finding" in app_source
    assert "_note_page_reference_map" in app_source


def test_rounding_check_flags_bad_visible_total():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "",
                [
                    [
                        ["Description", "2025"],
                        ["Revenue", "100"],
                        ["Other income", "50"],
                        ["Total income", "160"],
                    ]
                ],
            )
        ]
    )

    findings = check_rounding_and_casting(document, tolerance=Decimal("1"))

    assert findings
    assert findings[0].category == "Totals and rounding"


def test_rounding_scale_ignores_narrative_million_amounts_when_presentation_is_thousands():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Presentation currency N'000\nA deferred tax liability of N2 million arose during the year.",
                [],
            )
        ]
    )

    findings = check_rounding_and_casting(document, tolerance=Decimal("1"))

    assert not any("Mixed rounding or scaling" in finding.issue for finding in findings)


def test_column_consistency_ignores_narrative_directors_report_tables():
    document = PdfDocument(
        [
            PdfPage(
                4,
                "Directors' report",
                [
                    [
                        ["Financial statements for the year ended 31 December 2025"],
                        ["Revenue", "2025", "2024"],
                        ["Profit", "100", "90"],
                    ]
                ],
            )
        ]
    )

    findings = check_rounding_and_casting(document, tolerance=Decimal("1"))

    assert not any("only one comparative period" in finding.issue.lower() for finding in findings)


def test_review_memo_zero_high_uses_extraction_quality_next_step():
    from models import Finding, ReviewResult

    result = ReviewResult(
        findings=[
            Finding(
                "Extraction quality",
                "Medium",
                "Table extraction",
                "Some extracted table cells appear to contain multiple merged values.",
                "Detected merged cells.",
                "Review extraction quality.",
            )
        ],
        metrics={
            "findings": 1,
            "high": 0,
            "positive_assurance": "No exceptions noted from line-based checks on primary statements.",
        },
    )

    memo = build_ai_review_memo(result)

    assert "clear high-severity items first" not in memo
    assert "No high-severity exceptions were identified" in memo
    assert "rerun detailed note agreement after table extraction confidence improves" in memo


def test_note_column_is_not_treated_as_amount_column():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "",
                [
                    [
                        ["Description", "Note", "2025", "2024"],
                        ["Revenue", "5", "100", "90"],
                        ["Other income", "6", "50", "40"],
                        ["Total income", "99", "150", "130"],
                    ]
                ],
            )
        ]
    )

    findings = check_rounding_and_casting(document, tolerance=Decimal("1"))

    assert findings == []


def test_statement_line_parser_handles_split_amounts_and_note_columns():
    assert _amounts_from_statement_line("Other Revenue 9 307,482 1 89,751")[-2:] == [
        Decimal("307482"),
        Decimal("189751"),
    ]
    assert _amounts_from_statement_line("Office accommodation costs 10 3 9,772 1 7,996")[-2:] == [
        Decimal("39772"),
        Decimal("17996"),
    ]
    assert _amounts_from_statement_line("Total Liabilities 1 41,411 1 54,819")[-2:] == [
        Decimal("141411"),
        Decimal("154819"),
    ]
    assert _amounts_from_statement_line("Net cash inflow from financing activities - ( 16,128)")[-2:] == [
        Decimal("0"),
        Decimal("-16128"),
    ]


def test_ocr_statement_row_parser_separates_note_number_from_amounts(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of financial position",
                        "Share capital 8 10,000 10,000",
                        "Cash and cash equivalents 7 39,387 193,627",
                    ]
                ),
                [],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    rows = reviewer._statement_row_parses(document.pages[0].text)
    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    debug_rows = result.metrics["ocr_statement_rows"]

    assert rows["share capital"].note_ref == "8"
    assert rows["share capital"].amounts == (Decimal("10000"), Decimal("10000"))
    assert rows["cash and cash equivalents"].note_ref == "7"
    assert rows["cash and cash equivalents"].amounts == (Decimal("39387"), Decimal("193627"))
    assert rows["cash and cash equivalents"].correction_applied == "No"
    assert "share capital | 8 | 10,000 | 10,000" in debug_rows
    assert "cash and cash equivalents | 7 | 39,387 | 193,627" in debug_rows


def test_ocr_statement_row_parser_extracts_sfp_note_numbers_without_merging_amounts():
    rows = reviewer._statement_row_parses(
        "\n".join(
            [
                "Statement of financial position",
                "Investment property 3 4,072,229 4,112,524",
                "Trade and other receivables 5 1,910,631 131,254",
                "Financial liabilities 9 5,356,392 4,555,742",
            ]
        )
    )

    assert rows["investment property"].note_ref == "3"
    assert rows["investment property"].amounts == (Decimal("4072229"), Decimal("4112524"))
    assert rows["trade and other receivables"].note_ref == "5"
    assert rows["trade and other receivables"].amounts == (Decimal("1910631"), Decimal("131254"))
    assert rows["financial liabilities"].note_ref == "9"
    assert rows["financial liabilities"].amounts == (Decimal("5356392"), Decimal("4555742"))


def test_ocr_statement_row_parser_keeps_split_bracketed_current_year_amount():
    rows = reviewer._statement_row_parses(
        "\n".join(
            [
                "Statement of profit or loss",
                "Loss before taxation (221,494) (173,516)",
                "Taxation (226,684) 53,127",
                "Loss for the year (448, 178) (120,389)",
            ]
        )
    )

    assert rows["profit after tax"].amounts == (Decimal("-448178"), Decimal("-120389"))


def test_ocr_statement_amount_parser_normalizes_decimal_thousands_separator():
    rows = reviewer._statement_row_parses(
        "\n".join(
            [
                "Statement of profit or loss",
                "Loss before taxation (173,516) (681,559)",
                "Taxation 53.127 253.124",
            ]
        )
    )

    assert rows["taxation"].amounts == (Decimal("53127"), Decimal("253124"))


def test_simple_note_amount_parser_normalizes_split_digits_and_spaces():
    assert reviewer._simple_note_amounts_from_line("Annual Public Lecture 2 55 664") == [Decimal("255"), Decimal("664")]
    assert reviewer._simple_note_amounts_from_line("Account payable 7 1,392 59,081") == [Decimal("71392"), Decimal("59081")]
    assert reviewer._simple_note_amounts_from_line("Medical 2 2,537 11,280") == [Decimal("22537"), Decimal("11280")]


def test_ocr_income_alignment_infers_missing_current_year_after_tax_from_prior_match():
    rows = reviewer._statement_row_parses(
        "\n".join(
            [
                "Statement of profit or loss",
                "2022 2021",
                "Loss before taxation (221,494) (173,516)",
                "Taxation (226,684) 53,127",
                "Loss for the year (120,389)",
            ]
        )
    )

    assert rows["profit after tax"].amounts == (Decimal("-448178"), Decimal("-120389"))
    assert rows["profit after tax"].confidence == "Low-Medium"


def test_ocr_income_alignment_does_not_infer_single_amount_without_prior_cast_match():
    rows = reviewer._statement_row_parses(
        "\n".join(
            [
                "Statement of profit or loss",
                "2022 2021",
                "Loss before taxation (221,494) (173,516)",
                "Taxation (226,684) 53,127",
                "Loss for the year (999,999)",
            ]
        )
    )

    assert rows["profit after tax"].amounts == (Decimal("-999999"),)


def test_ocr_statement_row_parser_does_not_concatenate_two_digit_note_with_revenue():
    rows = reviewer._statement_row_parses("Statement of profit or loss\nRevenue 13 707,189 297,041")

    assert rows["revenue"].note_ref == "13"
    assert rows["revenue"].amounts == (Decimal("707189"), Decimal("297041"))
    assert rows["revenue"].correction_applied == "No"


def test_statement_row_parser_ignores_implicit_note_refs_on_statement_totals_and_cash_summary_rows():
    rows = reviewer._statement_row_parses(
        "\n".join(
            [
                "Statement of financial position",
                "Current assets 1 1,172,176 811,598",
                "Statement of cash flows",
                "Total cash movement for the year 7 0,224 (15,461)",
                "Cash and cash equivalents at the beginning of the year 8 75,869 605,645",
            ]
        )
    )

    assert rows["current assets"].note_ref == ""
    assert rows["current assets"].amounts == (Decimal("1"), Decimal("1172176"), Decimal("811598"))[-2:]
    assert rows["total cash movement for the year"].note_ref == ""
    assert rows["cash at beginning"].note_ref == ""


def test_primary_statement_checks_run_from_line_text_when_tables_are_low_confidence():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of income and expenditure",
                        "Subscriptions 242,511 120,424",
                        "Registrations 75,392 81,050",
                        "Operating Revenue 7 2 ,783,064 2,029,846",
                        "Gross Operating Revenue 3,100,967 2,231,320",
                        "Operating Expenditure 8 (1,269,506) (986,920)",
                        "Gross Revenue 1,831,461 1,244,400",
                        "Other Revenue 9 307,482 1 89,751",
                        "Finance Income 5 2,200 40,216",
                        "TOTAL INCOME 2,191,143 1 ,474,367",
                        "Office accommodation costs 10 3 9,772 1 7,996",
                        "Personnel costs 633,447 441,731",
                        "Administrative costs 871,713 589,872",
                        "Finance Expenses 1 1,013 14,599",
                        "TOTAL EXPENDITURE 1,555,945 1 ,064,198",
                        "SURPLUS OF INCOME OVER EXPENDITURE 635,198 410,169",
                        "TOTAL COMPREHENSIVE INCOME 635,198 410,169",
                    ]
                ),
                [],
            )
        ]
    )

    findings, performed, skipped = check_primary_statement_consistency(document)

    assert findings == []
    assert any("total income checked" in item for item in performed)
    assert any("total expenditure checked" in item for item in performed)


def test_single_page_sfp_extract_runs_limited_scope_checks_without_full_afs_highs(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Funtierra Limited",
                        "Statement of Financial Position",
                        "As at 31 December 2025",
                        "N’000",
                        "Non-current assets 4,072,229 4,112,524",
                        "Current assets 2,160,355 324,881",
                        "Total assets 6,232,584 4,437,405",
                        "Equity 876,192 (118,337)",
                        "Liabilities 5,356,392 4,555,742",
                        "Total equity and liabilities 6,232,584 4,437,405",
                        "Statement extract detail " * 60,
                    ]
                ),
                [],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    memo = build_ai_review_memo(result)

    assert result.metrics["document_scope"] == "Limited-scope statement extract"
    assert result.metrics["detected_profile"]["Document scope"] == "Limited-scope statement extract"
    assert result.metrics["detected_profile"]["Currency"] == "NGN / N'000"
    assert result.metrics["high"] == 0
    assert not any("IAS 1" in finding.location for finding in result.findings)
    assert any(finding.category == "Document scope" for finding in result.findings)
    assert "Statement of financial position: total assets checked from line-extracted rows." in result.metrics["checks_performed"]
    assert "Limited-scope review performed on Statement of Financial Position only." in memo
    assert "Full financial statement completeness" in result.metrics["checks_skipped"]


def test_detected_company_name_removes_short_ocr_prefix():
    document = PdfDocument([PdfPage(1, "Fl Funtierra Limited\nFinancial statements", [])])

    assert infer_detected_profile(document)["Company name"] == "Funtierra Limited"


def test_ocr_profit_or_loss_rows_are_parsed_with_fuzzy_labels():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of profit or loss",
                        "Revenve 1,000 900",
                        "Loss before taxation (200) (100)",
                        "Taxation (50) (30)",
                        "Loss after tax (250) (130)",
                    ]
                ),
                [],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )

    findings, performed, skipped = check_primary_statement_consistency(document)

    assert not findings
    assert any("profit/loss after tax checked" in item for item in performed)


def test_ocr_sfp_requested_rows_are_parsed_with_fuzzy_labels():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of financial position",
                        "Non current assets 600 500",
                        "Current assets 400 300",
                        "Total assets 1,000 800",
                        "Equity 700 550",
                        "Liabilities 300 250",
                        "Total liabilities and equity 1,000 800",
                    ]
                ),
                [],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )

    findings, performed, skipped = check_primary_statement_consistency(document)

    assert not findings
    assert any("total assets checked from line-extracted rows" in item for item in performed)
    assert any("equity and liabilities equation checked" in item for item in performed)



def test_sfp_equity_liabilities_uses_total_liabilities_when_financial_liabilities_row_exists():
    document = PdfDocument(
        [
            PdfPage(
                19,
                "\n".join(
                    [
                        "Statement of financial position",
                        "Assets",
                        "Total assets 33,264,220 17,882,380 28,059,917 15,380,685",
                        "Liabilities",
                        "Financial liabilities 17,478,416 8,364,386 9,532,167 6,910,153",
                        "Total liabilities 35,549,090 17,628,420 27,294,566 14,830,227",
                        "Equity",
                        "Total equity (2,284,870) 253,960 765,351 550,458",
                        "Total liabilities and equity 33,264,220 17,882,380 28,059,917 15,380,685",
                    ]
                ),
                [],
            )
        ]
    )

    findings, performed, skipped = check_primary_statement_consistency(document)

    assert any("equity and liabilities equation checked" in item for item in performed)
    assert not any("Equity plus liabilities" in finding.issue for finding in findings)
    assert not findings

def test_ocr_cash_beginning_and_end_rows_are_parsed():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of cash flows",
                        "Net increase in cash and cash equivalents 100 80",
                        "Cash and cash equivalents at beginning of year 250 170",
                        "Cash and cash equivalents at end of year 350 250",
                    ]
                ),
                [],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )

    findings, performed, skipped = check_primary_statement_consistency(document)

    assert not findings
    assert any("opening plus movement checked to closing" in item for item in performed)


def test_cash_flow_statement_keeps_clean_row_parse_over_noisier_statement_note_line_parse():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of cash flows",
                        "Net cash inflow from operating activities (a+b) 3 40,664 3 68,868",
                        "Net cash absorbed in investing activities ( 70,440) ( 103,685)",
                        "Net cash inflow from financing activities - ( 16,128)",
                        "Net increase in cash and cash equivalents 2 70,224 2 49,055",
                        "Cash and cash equivalents at the beginning of the year 6 05,645 3 56,590",
                        "Cash and cash equivalents as at the end of the year 18 8 75,869 6 05,645",
                    ]
                ),
                [],
            )
        ]
    )

    findings, performed, skipped = check_primary_statement_consistency(document)

    assert not any("Operating, investing, and financing cash flows agree to net increase in cash" in f.issue for f in findings)
    assert not any("Opening cash plus total movement" in f.issue for f in findings)
    assert any("Statement of cash flows: net cash increase checked." == item for item in performed)
    assert any("Statement of cash flows: opening plus movement checked to closing." == item for item in performed)


def test_cash_flow_statement_checks_net_movement_when_investing_section_is_absent_but_zero():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of cash flows",
                        "Net cash generated from/(used in) operating activities 2,579 49,881",
                        "Net cash used in financing activities - (39,917)",
                        "Total cash movement for the year 2,579 9,964",
                        "Loss on foreign exchange (2,579) (9,964)",
                        "Cash and cash equivalents at the end of the year - -",
                    ]
                ),
                [],
            )
        ]
    )

    findings, performed, skipped = check_primary_statement_consistency(document)

    assert any("Statement of cash flows: net cash increase checked." == item for item in performed)
    assert not any("net cash increase checked" in item.lower() for item in skipped)


def test_arithmetic_skips_merged_numeric_cells():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "",
                [[["Description", "2025", "2024"], ["Revenue", "100 90", ""], ["Total revenue", "100", "90"]]],
            )
        ]
    )

    findings = check_rounding_and_casting(document, tolerance=Decimal("1"))

    assert getattr(document, "skipped_table_details", [])
    assert not any("does not agree" in finding.issue.lower() for finding in findings)


def test_arithmetic_skips_five_year_summary():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "",
                [
                    [
                        ["Five year financial summary", "2025", "2024", "2023"],
                        ["Revenue", "100", "90", "80"],
                        ["Total assets", "300", "280", "250"],
                        ["Total equity", "200", "190", "170"],
                    ]
                ],
            )
        ]
    )

    findings = check_rounding_and_casting(document, tolerance=Decimal("1"))

    assert any("multi-year summary" in detail.lower() for detail in getattr(document, "skipped_table_details", []))
    assert not any("does not agree" in finding.issue.lower() for finding in findings)


def test_arithmetic_skips_value_added_statement():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "",
                [
                    [
                        ["Value added statement", "2025", "2024"],
                        ["Revenue", "100", "90"],
                        ["Bought in materials", "40", "35"],
                        ["Total value added", "60", "55"],
                    ]
                ],
            )
        ]
    )

    findings = check_rounding_and_casting(document, tolerance=Decimal("1"))

    assert any("value-added statement" in detail.lower() for detail in getattr(document, "skipped_table_details", []))
    assert not any("does not agree" in finding.issue.lower() for finding in findings)


def test_low_confidence_table_skips_are_grouped():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "",
                [
                    [["Value added statement", "2025", "2024"], ["Revenue", "100", "90"], ["Total", "100", "90"]],
                    [["Five year financial summary", "2025", "2024"], ["Assets", "300", "280"], ["Equity", "200", "190"]],
                ],
            )
        ]
    )

    findings = check_rounding_and_casting(document, tolerance=Decimal("1"))
    grouped = getattr(document, "skipped_table_details", [])

    assert len(grouped) == 2
    assert any("Page 1, table 1" in detail for detail in grouped)
    assert any("Page 1, table 2" in detail for detail in grouped)


def test_skipped_table_details_include_reviewer_and_source_pages(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                10,
                "Statement page with financial table\n9",
                [
                    [
                        ["Value added statement", "2025", "2024"],
                        ["Revenue", "100", "90"],
                        ["Total value added", "100", "90"],
                    ]
                ],
            )
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions())

    details = result.metrics["skipped_table_details"]
    assert details[0]["Page"] == "9"
    assert details[0]["Source PDF page"] == "10"
    assert result.metrics["skipped_table_summary"][0]["Pages affected"] == "Page 9"


def test_ai_rate_limit_message_does_not_tell_user_to_refresh():
    ai_policy_review._AI_RATE_LIMIT_UNTIL = 0
    ai_policy_review._set_rate_limit_block(90)

    message = ai_policy_review._friendly_ai_error_message(RuntimeError("429 rate limit exceeded"))

    assert "refresh" not in message.lower()
    assert "deterministic review" in message.lower()
    assert "automatic retry" in message.lower()
    assert "20 second(s)" not in message
    assert ai_policy_review._rate_limit_wait_seconds() <= 120


def test_generic_arithmetic_skips_ocr_reconstructed_tables():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\nRevenue 100 90\nTotal revenue 999 999",
                [[["Description", "2025", "2024"], ["Revenue", "100", "90"], ["Total revenue", "999", "999"]]],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
        ocr_tables=1,
    )

    findings = check_rounding_and_casting(document, tolerance=Decimal("1"))

    assert any("generic arithmetic checks were skipped" in finding.issue.lower() for finding in findings)
    assert not any("does not agree" in finding.issue.lower() for finding in findings)


def test_arithmetic_respects_repeated_table_headers_in_note_blocks():
    document = PdfDocument(
        [
            PdfPage(
                20,
                "Notes to the financial statements\nN'000",
                [
                    [
                        ["December", "31", "2025"],
                        ["N'", "000", "000"],
                        ["Trade and other receivable", "288,706", "288,706"],
                        ["Cash", "875,869", "875,869"],
                        ["Note (s) N'", "000", "000"],
                        ["Trade and other payables", "141,411", "141,411"],
                        ["Note (s) N'", "000", "000"],
                        ["Trade and other payables", "154,819", "154,819"],
                        ["Note (s) N'", "000", "000"],
                        ["Trade and other payables", "141,411", "1", "54,819"],
                        ["Total Liabilities", "141,411", "154,819"],
                    ]
                ],
            )
        ]
    )

    findings = check_rounding_and_casting(document, tolerance=Decimal("1"))

    assert not any("does not agree" in finding.issue.lower() for finding in findings)


def test_arithmetic_stops_when_new_note_heading_starts():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "",
                [
                    [
                        ["Description", "2025"],
                        ["Other payables", "127,575"],
                        ["Withholding tax - State", "7,968"],
                        ["Accrued fees", "5,869"],
                        ["Total", "141,411"],
                        ["20 Receivables & Advances", ""],
                        ["Receivables", "2,000,000"],
                        ["Total", "2,000,000"],
                    ]
                ],
            )
        ]
    )

    findings = check_rounding_and_casting(document, tolerance=Decimal("1"))

    assert not any("visible sum 2,141,411" in finding.evidence for finding in findings)


def test_notes_check_flags_missing_note_heading():
    document = PdfDocument([PdfPage(1, "Revenue Note 7 100\nProfit for the year 40", [])])

    findings = check_notes_agreement(document)

    assert any("note 7" in finding.issue.lower() for finding in findings)


def test_note_reference_detection_rejects_years_and_large_numbers():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of profit or loss\nRevenue 2025 4001 3001\nOther Revenue 9 307,482 189,751",
                [],
            ),
            PdfPage(2, "Notes\n9 Other Revenue\nFair value gain 307,482", []),
        ]
    )

    findings = check_notes_agreement(document)

    assert not any("note 2025" in finding.issue.lower() for finding in findings)
    assert not any("note 4001" in finding.issue.lower() for finding in findings)
    assert not any("note 3001" in finding.issue.lower() for finding in findings)
    assert not any("note 9" in finding.issue.lower() and "not found" in finding.issue.lower() for finding in findings)


def test_note_heading_detection_handles_multiline_header_with_year_columns():
    text = "NOTE 9 2025 2024\nOther Revenue N'000 N'000\nFair Value Gain 100 90"

    headings = _note_headings(text)

    assert headings["9"] == "Other Revenue"


def test_note_heading_detection_handles_combined_and_split_table_headings():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of income and expenditure",
                        "Personnel costs 11 633,447 441,731",
                    ]
                ),
                [],
            ),
            PdfPage(
                2,
                "\n".join(
                    [
                        "NOTES TO THE FINANCIAL STATEMENTS",
                        "Note - 2025 7 8",
                        "Cost of Generating",
                        "Revenue Heads Operating Revenue Operating Revenue Net Surplus",
                        "N'000 N'000 N'000",
                        "2,783,064 1,269,506 1,513,130",
                        "10 Office accommodation N'000 N'000",
                        "Repairs 39,772 17,996",
                        "11(a) Staff costs N'000 N'000",
                        "Salaries 633,447 441,731",
                        "16 Inventories 2025 2024",
                        "N'000 N'000",
                        "Study packs 7,601 15,969",
                    ]
                ),
                [],
            )
        ]
    )

    headings = _note_headings_by_page(document)

    assert headings["7"][0] == "Operating Revenue"
    assert headings["8"][0] == "Operating Expenditure"
    assert headings["10"][0] == "Office accommodation"
    assert headings["11"][0] == "Staff costs"
    assert headings["16"][0] == "Inventories"


def test_disclosure_only_notes_are_not_flagged_as_unreferenced():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\nCash Note 18 100",
                [],
            ),
            PdfPage(
                2,
                "18 Cash and cash equivalents\nCash 100\n25 Capital commitments\nThere were no capital commitments.\n26 Contingent liabilities\nNone.\n27 Subsequent events disclosure\nThere were no subsequent events.\n28 Related party transactions\nNo transactions.",
                [],
            ),
        ]
    )

    findings = check_notes_agreement(document)

    assert not any("note 25 exists but was not referenced" in finding.issue.lower() for finding in findings)
    assert not any("note 26 exists but was not referenced" in finding.issue.lower() for finding in findings)
    assert not any("note 27 exists but was not referenced" in finding.issue.lower() for finding in findings)
    assert not any("note 28 exists but was not referenced" in finding.issue.lower() for finding in findings)


def test_notes_check_flags_possible_wrong_note_reference_when_item_is_in_another_note():
    filler = "Revenue 100 90\n" * 80
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of income and expenditure\nOther Revenue 7 307,482 189,751\n" + filler,
                [],
            ),
            PdfPage(
                2,
                "Notes to the financial statements\n7 Operating revenue\nSubscriptions 242,511 120,424\nRegistrations 75,392 81,050\n9 Other Revenue\nFair value gain 307,482 189,751\n",
                [],
            ),
        ]
    )

    findings = check_notes_agreement(document)

    wrong_ref = [finding for finding in findings if "possible wrong note reference" in finding.issue.lower()]
    assert wrong_ref
    assert wrong_ref[0].severity == "High"
    assert "Other Revenue references Note 7" in wrong_ref[0].issue
    assert "Note 9 appears to be a stronger match" in wrong_ref[0].issue
    assert wrong_ref[0].metadata["statement"] == "Statement of income and expenditure"
    assert wrong_ref[0].metadata["line_item"] == "Other Revenue"
    assert wrong_ref[0].metadata["referenced_note"] == "7"
    assert wrong_ref[0].metadata["suggested_note"] == "9"


def test_note_reference_compatibility_is_generic_not_ppe_only():
    filler = "Statement narrative text for extraction confidence. " * 80
    document = PdfDocument(
        [
            PdfPage(1, "Statement of financial position\nIntangible assets 3 84,014 70,000\n" + filler, []),
            PdfPage(
                2,
                "\n".join(
                    [
                        "Notes to the financial statements",
                        "3 Investment property",
                        "Fair value property 1,000 900",
                        "4 Intangible assets",
                        "Software 84,014 70,000",
                    ]
                ),
                [],
            ),
        ]
    )

    findings = check_notes_agreement(document)

    wrong_ref = [finding for finding in findings if "note heading mismatch" in finding.issue.lower()]
    assert wrong_ref
    assert "Intangible Assets references Note 3" in wrong_ref[0].issue
    assert "Note 4" in wrong_ref[0].issue
    assert wrong_ref[0].metadata["suggested_note"] == "4"


def test_note_reference_compatibility_flags_liability_heading_mismatch():
    filler = "Statement narrative text for extraction confidence. " * 80
    document = PdfDocument(
        [
            PdfPage(1, "Statement of financial position\nTrade and other payables 5 526,917 311,755\n" + filler, []),
            PdfPage(
                2,
                "\n".join(
                    [
                        "Notes to the financial statements",
                        "5 Trade and other receivables",
                        "Receivables 10,000 9,000",
                        "6 Trade and other payables",
                        "Accruals 526,917 311,755",
                    ]
                ),
                [],
            ),
        ]
    )

    findings = check_notes_agreement(document)

    wrong_ref = [finding for finding in findings if "note heading mismatch" in finding.issue.lower()]
    assert wrong_ref
    assert "Trade And Other Payables references Note 5" in wrong_ref[0].issue
    assert wrong_ref[0].metadata["suggested_note"] == "6"


def test_note_reference_compatibility_does_not_flag_matching_generic_heading():
    filler = "Statement narrative text for extraction confidence. " * 80
    document = PdfDocument(
        [
            PdfPage(1, "Statement of financial position\nCash and cash equivalents 7 39,387 193,627\n" + filler, []),
            PdfPage(
                2,
                "Notes to the financial statements\n7 Cash and cash equivalents\nBank balances 39,387 193,627",
                [],
            ),
        ]
    )

    findings = check_notes_agreement(document)

    assert not any("note heading mismatch" in finding.issue.lower() for finding in findings)
    assert not any("possible wrong note reference" in finding.issue.lower() for finding in findings)


def test_wrong_note_reference_check_respects_low_confidence_gate():
    document = PdfDocument(
        [
            PdfPage(1, "Statement of income and expenditure\nOther Revenue 7 307,482 189,751", []),
            PdfPage(
                2,
                "Notes to the financial statements\n7 Operating revenue\nSubscriptions 242,511\n9 Other Revenue\nFair value gain 307,482",
                [],
            ),
        ]
    )

    normal_findings = check_notes_agreement(document)
    cautious_findings = check_notes_agreement(document, cautious_low_confidence=True)

    assert not any("possible wrong note reference" in finding.issue.lower() for finding in normal_findings)
    assert not any("possible wrong note reference" in finding.issue.lower() for finding in cautious_findings)


def test_cash_flow_wrong_note_reference_requires_unmistakably_stronger_alternative():
    document = PdfDocument(
        [
            PdfPage(1, "Statement of cash flows\nImpairment losses and reversals 5A 16,228", []),
            PdfPage(
                2,
                "\n".join(
                    [
                        "Notes to the financial statements",
                        "5 Financial assets",
                        "See note 6A for movement in expected credit loss allowance.",
                        "5A Financial assets at amortised cost",
                        "Gross balance 100,000 90,000",
                        "6 Trade and other receivables",
                        "6A Movement in credit loss allowances",
                        "Reversal recognised 16,228 0",
                    ]
                ),
                [],
            ),
        ]
    )

    findings = check_notes_agreement(document)

    assert not any("possible wrong note reference" in finding.issue.lower() for finding in findings)


def test_cash_flow_subnote_amount_gap_is_not_elevated_as_hard_exception():
    document = PdfDocument(
        [
            PdfPage(1, "Statement of cash flows\nImpairment losses and reversals 5A 16,228", []),
            PdfPage(
                2,
                "\n".join(
                    [
                        "Notes to the financial statements",
                        "5 Financial assets",
                        "5A Financial assets at amortised cost",
                        "Gross balance 100,000 90,000",
                        "6 Trade and other receivables",
                        "6A Movement in credit loss allowances",
                        "Reversal recognised 16,228 0",
                    ]
                ),
                [],
            ),
        ]
    )

    findings = check_notes_agreement(document)

    assert not any(
        finding.category == "Notes agreement"
        and "not found in the related note text" in finding.issue.lower()
        for finding in findings
    )


def test_simple_note_text_casting_skips_breakdown_sections():
    page = PdfPage(
        33,
        "\n".join(
            [
                "17. Other operating gains/(losses)",
                "Net foreign exchange gains/(losses) 87,178 (385,317)",
                "Breakdown of realized and unrealized exchange difference",
                "Realized exchange gains/(losses) 52,003 (16,162)",
                "Unrealized exchange gains/(losses) 35,175 (369,155)",
                "87,178 (385,317)",
            ]
        ),
        [],
    )

    findings = reviewer._check_simple_note_text_casting(page, Decimal("1"))

    assert not any("simple note section total does not agree" in finding.issue.lower() for finding in findings)


def test_cautious_note_reference_validation_uses_detected_headings_when_sections_are_weak():
    document = PdfDocument(
        [
            PdfPage(1, "Statement of income and expenditure\nOther Revenue 7 307,482 189,751", []),
            PdfPage(
                2,
                "Notes to the financial statements\n7 Operating Revenue\nRevenue schedule text without matching amount\n9 Other Revenue\nNarrative only",
                [],
            ),
        ]
    )

    findings = check_notes_agreement(document, cautious_low_confidence=True)
    wrong_ref = [finding for finding in findings if "possible wrong note reference" in finding.issue.lower()]

    assert wrong_ref == []


def test_cautious_note_reference_validation_flags_explicit_missing_note_as_review_prompt():
    document = PdfDocument(
        [
            PdfPage(1, "Statement of financial position\nCash Note 18 875,869 605,645", []),
            PdfPage(2, "Notes to the financial statements\n19 Trade and other payables\nOther payables 141,411", []),
        ]
    )

    findings = check_notes_agreement(document, cautious_low_confidence=True)
    missing = [finding for finding in findings if "referenced note not found" in finding.issue.lower()]

    assert missing
    assert missing[0].severity == "Low"
    assert missing[0].metadata["line_item"] == "Cash"
    assert missing[0].metadata["referenced_note"] == "18"


def test_cautious_face_to_note_amount_agreement_flags_amount_found_elsewhere():
    document = PdfDocument(
        [
            PdfPage(1, "Statement of financial position\nCash Note 18 875,869 605,645", []),
            PdfPage(
                2,
                "\n".join(
                    [
                        "Notes to the financial statements",
                        "18 Cash and cash equivalents",
                        "Bank balances 100,000 90,000",
                        "19 Cash balances",
                        "Bank balances 875,869 605,645",
                    ]
                ),
                [],
            ),
        ]
    )

    findings = check_notes_agreement(document, cautious_low_confidence=True)
    amount_prompts = [finding for finding in findings if "amount appears in another note" in finding.issue.lower()]

    assert amount_prompts
    assert amount_prompts[0].severity == "Medium"
    assert amount_prompts[0].metadata["statement"] == "Statement of financial position"
    assert amount_prompts[0].metadata["line_item"] == "Cash"
    assert amount_prompts[0].metadata["referenced_note"] == "18"
    assert amount_prompts[0].metadata["alternative_note_found"] == "19"
    assert amount_prompts[0].metadata["current_year_amount_found"] == "No"
    assert amount_prompts[0].metadata["prior_year_amount_found"] == "No"
    assert amount_prompts[0].metadata["amount_match_confidence"] == "Medium"


def test_cautious_face_to_note_amount_agreement_keeps_low_amount_not_located_out_of_findings(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of income and expenditure\nOther Revenue Note 9 307,482 189,751", []),
            PdfPage(
                2,
                "Notes to the financial statements\n9 Other Revenue\nFair value gain 100,000 90,000\n10 Office accommodation\nRent 50,000 40,000",
                [],
            ),
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    findings = check_notes_agreement(document, cautious_low_confidence=True)
    amount_prompts = [finding for finding in findings if "amount not located in referenced note" in finding.issue.lower()]
    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    rows = result.metrics["note_agreement_results"]

    assert not amount_prompts
    assert any(
        row["Line item"] == "Other Revenue"
        and row["Result"] == "Review prompt"
        and row["Reason"] == "Amount not located in referenced note."
        for row in rows
    )


def test_cautious_face_to_note_amount_agreement_does_not_flag_when_amounts_are_in_referenced_note():
    document = PdfDocument(
        [
            PdfPage(1, "Statement of income and expenditure\nOther Revenue Note 9 307,482 189,751", []),
            PdfPage(2, "Notes to the financial statements\n9 Other Revenue\nFair value gain 307,482 189,751", []),
        ]
    )

    findings = check_notes_agreement(document, cautious_low_confidence=True)

    assert not any("amount not located" in finding.issue.lower() for finding in findings)
    assert not any("amount appears in another note" in finding.issue.lower() for finding in findings)


def test_cautious_face_to_note_amount_agreement_matches_normalized_visible_amounts(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of income and expenditure\nOperating Revenue Note 7 2,783,064 2,029,846", []),
            PdfPage(2, "Notes to the financial statements\n7 Operating Revenue\nSubscriptions 2 ,783,064\nPrior year 2\n,029,846", []),
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    row = next(row for row in result.metrics["note_agreement_results"] if row["Line item"] == "Operating Revenue")

    assert row["Result"] == "Passed"
    assert row["Current year amount found in referenced note?"] == "Yes"
    assert row["Prior year amount found in referenced note?"] == "Yes"
    assert "normalized amount" in row["Matching method"]
    assert "2 ,783,064" in row["Matched text snippet from referenced note"]


def test_cautious_face_to_note_amount_agreement_does_not_suggest_amount_only_alternative(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of income and expenditure\nOperating Revenue Note 7 2,783,064 2,029,846", []),
            PdfPage(
                2,
                "Notes to the financial statements\n7 Operating Revenue\nSubscriptions 100,000 90,000\n8 Operating Expenditure\nExpenses 2,783,064 2,029,846",
                [],
            ),
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    row = next(row for row in result.metrics["note_agreement_results"] if row["Line item"] == "Operating Revenue")

    assert row["Alternative note found"] == ""
    assert row["Result"] == "Review prompt"
    assert not any("Operating Revenue amount appears in Note 8" in finding.issue for finding in result.findings)


def test_cash_note_reference_does_not_suggest_non_cash_alternative(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of financial position\nCash At End Of The Year Note 7 100 90", []),
            PdfPage(
                2,
                "Notes to the financial statements\n3 Investment property\nCash At End Of The Year 100 90\n7 Directors' interests in shares\nNarrative only\n"
                + ("Additional OCR note text for coverage.\n" * 80),
                [],
            ),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert not any(
        finding.metadata
        and finding.metadata.get("line_item") == "Cash At End Of The Year"
        and finding.metadata.get("suggested_note") == "3"
        for finding in result.findings
    )


def test_revenue_note_reference_does_not_suggest_investment_property_without_revenue_context(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of profit or loss\nRevenue Note 13 707,189 297,041", []),
            PdfPage(
                2,
                "Notes to the financial statements\n3 Investment property\nFair value movement 707,189 297,041\n13 Revenue\nNarrative only\n"
                + ("Additional OCR note text for coverage.\n" * 80),
                [],
            ),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert not any(
        finding.metadata
        and finding.metadata.get("line_item") == "Revenue"
        and finding.metadata.get("suggested_note") == "3"
        for finding in result.findings
    )


def test_ocr_revenue_heading_prompt_stays_in_note_results_not_exception_register(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of profit or loss\nRevenue Note 13 707,189 297,041", []),
            PdfPage(
                2,
                "Notes to the financial statements\n3 Rental income\nRental income 707,189 297,041\n13 Investment property\nNarrative only\n"
                + ("Additional OCR note text for coverage.\n" * 80),
                [],
            ),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    row = next(row for row in result.metrics["note_agreement_results"] if row["Line item"] == "Revenue")

    assert row["Alternative note found"] == "3"
    assert row["Match confidence"] == "Low"
    assert row["Result"] == "Skipped"
    assert "debug" in row["Reason"].lower()
    assert not any(finding.metadata and finding.metadata.get("line_item") == "Revenue" for finding in result.findings)


def test_ocr_revenue_heading_prompt_runs_for_direct_revenue_heading(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of profit or loss\nOther Revenue Note 7 307,482 189,751", []),
            PdfPage(
                2,
                "Notes to the financial statements\n7 Operating revenue\nRevenue 100 90\n9 Other Revenue\nOther Revenue 307,482 189,751\n"
                + ("Additional OCR note text for coverage.\n" * 80),
                [],
            ),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    row = next(row for row in result.metrics["note_agreement_results"] if row["Line item"] == "Other Revenue")

    assert row["Alternative note found"] == "9"
    assert row["Result"] == "Review prompt"
    assert result.metrics["note_reference_findings"] == 0


def test_ocr_revenue_note_reference_does_not_suggest_weak_tenant_note(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of profit or loss\nRevenue Note 13 707,189 297,041", []),
            PdfPage(
                2,
                "Notes to the financial statements\n10B Tenant deposits\nTenant balances and rent deposits 707,189 297,041\n13 Investment property\nNarrative only\n"
                + ("Additional OCR note text for coverage.\n" * 80),
                [],
            ),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    row = next(row for row in result.metrics["note_agreement_results"] if row["Line item"] == "Revenue")

    assert row["Alternative note found"] == ""
    assert row["Result"] == "Skipped"
    assert not any(finding.metadata and finding.metadata.get("suggested_note") == "10B" for finding in result.findings)


def test_revenue_amount_agreement_does_not_suggest_tenant_note_without_approved_heading(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of profit or loss\nRevenue Note 13 707,189 297,041", []),
            PdfPage(
                2,
                "Notes to the financial statements\n10B Tenant balances\nTenant rental balances 707,189 297,041\n13 Revenue\nNarrative only",
                [],
            ),
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    row = next(row for row in result.metrics["note_agreement_results"] if row["Line item"] == "Revenue")

    assert row["Alternative note found"] == ""
    assert not any(finding.metadata and finding.metadata.get("suggested_note") == "10B" for finding in result.findings)


def test_cautious_face_to_note_amount_agreement_skips_non_face_linked_lines(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of financial position\nCurrent liabilities Note 19 141,411 154,819", []),
            PdfPage(2, "Notes to the financial statements\n19 Trade and other payables\nOther payables 141,411 154,819", []),
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    row = next(row for row in result.metrics["note_agreement_results"] if row["Line item"] == "Current Liabilities")

    assert row["Result"] == "Skipped"
    assert row["Reason"] == "Skipped - not a face-linked note line"


def test_statement_amount_parser_does_not_merge_note_number_into_cash_amount():
    line = "Cash and cash equivalents at the end of the year 3 88,741 154,449"

    parsed = reviewer._parse_ocr_statement_row(line)

    assert parsed is not None
    assert parsed.label == "cash at end"
    assert parsed.amounts == (Decimal("88741"), Decimal("154449"))
    assert "388,741" not in reviewer._amount_tokens_from_statement_line(line)


def test_cash_flow_opening_movement_closing_allows_exchange_effect_after_note_token():
    text = """
Statement of Cash Flows
2025 2024
Cash flows from operating activities
Net cash generated from/(used in) operating activities 327,438 (273,155)
Cash flows from investing activities
Net cash used in investing activities (209,437) (826,228)
Cash flows from financing activities
Net cash (used in)/generated from financing activities (181,174) 1,244,762
Total cash movement for the year (63,173) 145,379
Cash and cash equivalents at the beginning of the year 154,449 9,070
Loss on foreign exchange on cash and cash equivalents (2,535) -
Cash and cash equivalents at the end of the year 3 88,741 154,449
"""
    page = PdfPage(14, text, [])

    findings, performed, skipped = reviewer._check_cash_flow_text(page, Decimal("1"))

    assert not findings
    assert "Statement of cash flows: opening plus movement checked to closing." in performed
    assert not skipped


def test_cash_flow_activity_subtotal_without_note_is_not_face_linked():
    item = reviewer.StatementNoteLine(
        statement_name="Statement of cash flows",
        page_number=14,
        line="Net cash (used in)/generated from financing activities (181,174) 1,244,762",
        line_item="Cash /generated from financing activities",
        ref="",
        amounts=(Decimal("-181174"), Decimal("1244762")),
        explicit_ref=False,
    )

    assert reviewer._note_agreement_skip_reason(item) == "not a face-linked note line"


def test_note_agreement_excludes_cash_flow_subtotals_without_explicit_note(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of cash flows",
                        "Net cash inflow from operating activities a 141,411 154,819",
                        "Net cash absorbed in investing activities b (20,000) (15,000)",
                        "Net increase in cash and cash equivalents a+b 121,411 139,819",
                    ]
                ),
                [],
            ),
            PdfPage(2, "Notes to the financial statements\n18 Cash and cash equivalents\nCash 121,411 139,819", []),
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert result.metrics["note_agreement_results"] == []
    assert not any("cash inflow" in finding.issue.lower() for finding in result.findings)


def test_note_agreement_keeps_cash_flow_rows_with_explicit_note(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of cash flows\nCash and cash equivalents at end of year Note 18 875,869 605,645", []),
            PdfPage(2, "Notes to the financial statements\n18 Cash and cash equivalents\nCash 875,869 605,645", []),
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert any(row["Line item"] == "Cash And Cash Equivalents At End Of Year" and row["Result"] == "Passed" for row in result.metrics["note_agreement_results"])


def test_note_agreement_excludes_value_added_and_five_year_summary(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of value added\nSurplus for the year 9 307,482 189,751", []),
            PdfPage(2, "Five year financial summary\nRevenue 7 2,783,064 2,029,846", []),
            PdfPage(3, "Notes to the financial statements\n7 Revenue\nRevenue 2,783,064 2,029,846\n9 Other Revenue\nOther revenue 307,482 189,751", []),
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert result.metrics["note_agreement_results"] == []


def test_cautious_face_to_note_amount_agreement_does_not_treat_amount_digits_as_note_refs():
    document = PdfDocument(
        [
            PdfPage(1, "Statement of financial position\nTotal Current Assets 1,172,176 811,598", []),
            PdfPage(2, "Notes to the financial statements\n8 Operating expenditure\nExpenses 811,598", []),
        ]
    )

    findings = check_notes_agreement(document, cautious_low_confidence=True)

    assert not any("current assets references note 8" in finding.issue.lower() for finding in findings)


def test_note_agreement_results_skip_rows_where_only_note_number_was_detected_as_amount(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of financial position\nOther Financial Assets Note 6 6", []),
            PdfPage(2, "Notes to the financial statements\n6 Other financial assets\nNarrative only", []),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    row = next(row for row in result.metrics["note_agreement_results"] if row["Line item"] == "Other Financial Assets")

    assert row["Referenced note"] == "6"
    assert row["Current year amount"] == ""
    assert row["Result"] == "Skipped"
    assert "no reliable statement amount" in row["Reason"].lower()


def test_review_pdf_reports_cautious_note_reference_override_as_performed(monkeypatch):
    noisy_table = [["Description", "2025", "2024"]]
    noisy_table.extend([["Line item", "100 90", ""] for _ in range(40)])
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of income and expenditure\nRevenue 100 90\n" * 80,
                [noisy_table],
            )
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert "Cautious face-to-note amount agreement performed in review-prompt mode." in result.metrics["checks_performed"]
    assert "Cautious note-reference validation performed in review-prompt mode; no possible wrong note references detected." in result.metrics["checks_performed"]
    assert "Cautious note-reference validation skipped" not in result.metrics["checks_skipped"]
    assert result.metrics["cautious_note_validation_enabled"] is True
    assert result.metrics["note_validation_mode"] == "review_prompt"
    assert result.metrics["note_reference_rows_detected"] == 0
    assert result.metrics["note_headings_detected"] == 0
    assert result.metrics["note_reference_findings"] == 0


def test_note_agreement_results_include_passed_and_review_prompt_rows(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of financial position",
                        "Cash Note 18 875,869 605,645",
                        "Trade payables Note 19 141,411 154,819",
                    ]
                ),
                [],
            ),
            PdfPage(
                2,
                "\n".join(
                    [
                        "Notes to the financial statements",
                        "18 Cash and cash equivalents",
                        "Cash at bank 875,869 605,645",
                        "19 Trade and other payables",
                        "Other payables 141,411 100,000",
                    ]
                ),
                [],
            ),
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    rows = result.metrics["note_agreement_results"]

    assert any(row["Line item"] == "Cash" and row["Result"] == "Passed" for row in rows)
    assert any(row["Line item"] == "Trade Payables" and row["Result"] == "Review prompt" for row in rows)
    assert any(row["Prior year amount found in referenced note?"] == "No" for row in rows)


def test_notes_agreement_is_conservative_for_ocr_documents():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\nRevenue Note 7 100\nNotes to the financial statements\n8 Revenue",
                [],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )

    findings = check_notes_agreement(document)

    assert len(findings) == 1
    assert findings[0].category == "Extraction quality"
    assert "skipped for an ocr-assisted document" in findings[0].issue.lower()


def test_non_exception_extraction_skip_messages_do_not_enter_findings():
    document = PdfDocument([PdfPage(1, "Statement of financial position", [])], ocr_used=True, ocr_pages=1)
    skipped_notice = reviewer.Finding(
        "Extraction quality",
        "Low",
        "Notes agreement",
        "Detailed note-reference reconciliation was skipped for an OCR-assisted document.",
        "OCR can misread note columns and note tables.",
        "Use OCR output for navigation, but rely on manual review for detailed agreement.",
    )

    result = reviewer._build_result(document, [skipped_notice])

    assert result.findings == []
    assert result.metrics["findings"] == 0
    assert "Detailed note-reference reconciliation was skipped" in result.metrics["checks_skipped"]
    assert any(
        row["Result"].startswith("Skipped") and "Detailed note-reference reconciliation" in row["Check"]
        for row in result.metrics["check_results"]
    )


def test_note_heading_detection_rejects_years_and_ocr_noise():
    text = "2021 Statement of changes\n2022 Revenue\n89B OCR garbage\n4 Revenue\n5 Trade and other receivables"

    headings = _note_headings(text)

    assert "2021" not in headings
    assert "2022" not in headings
    assert "89B" not in headings
    assert headings["4"] == "Revenue"


def test_note_heading_detection_rejects_report_furniture_and_entity_names():
    text = "\n".join(
        [
            "1 Significant accounting policies",
            "1S Example Holdings Limited",
            "4 Example Holdings Limited",
            "5 Directors",
            "7 Financial Statements for the year ended December 31, 2021",
            "9 _ Financial liabilities",
            "10 Trade and other payables",
        ]
    )

    headings = _note_headings(text)

    assert headings["1"] == "Significant accounting policies"
    assert "1S" not in headings
    assert "4" not in headings
    assert "5" not in headings
    assert "7" not in headings
    assert headings["9"] == "_ Financial liabilities"
    assert headings["10"] == "Trade and other payables"


def test_note_heading_detection_rejects_narrative_suffix_heading():
    text = "\n".join(
        [
            "Notes to the financial statements",
            "10A Advances from tenants",
            "10B This represents the advance payment made by tenants for future rental periods",
            "11 Trade and other payables",
        ]
    )

    headings = _note_headings(text)

    assert headings["10A"] == "Advances from tenants"
    assert "10B" not in headings
    assert headings["11"] == "Trade and other payables"


def test_note_heading_detection_accepts_clear_continued_heading_when_original_split():
    document = PdfDocument(
        [
            PdfPage(32, "Notes to the financial statements\nIntangible assets\nCost 100 90", []),
            PdfPage(33, "4. Intangible assets (continued)\nAccumulated amortisation 20 10", []),
        ]
    )

    headings = _note_headings_by_page(document)

    assert headings["4"] == ("Intangible assets", 33)


def test_missing_statement_note_reference_includes_statement_page(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                14,
                "Statement of financial position\nDeferred tax Note 9 100 90\nTotal assets 100 90\nTotal equity and liabilities 100 90\n"
                + ("Primary statement context.\n" * 70),
                [],
            ),
            PdfPage(35, "Notes to the financial statements\n10 Deferred tax\nDeferred tax 100 90\n" + ("Note context.\n" * 70), []),
        ]
    )
    monkeypatch.setattr(PdfDocument, "table_extraction_confidence", property(lambda self: 100))

    findings = check_notes_agreement(document)

    missing = next(finding for finding in findings if "Statement references note 9" in finding.issue)
    assert missing.location == "Page 14"
    assert "Page 14" in missing.evidence


def test_note_internal_total_skips_complex_movement_schedule_noise(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                20,
                "\n".join(
                    [
                        "Notes to the financial statements",
                        "6 Trade and other receivables",
                        "Movement in loss allowance",
                        "Opening balance 100",
                        "Charge for the year 20",
                        "Utilised during the year (10)",
                        "Closing balance 130",
                        "Total 999",
                    ]
                )
                + "\n"
                + ("Note context.\n" * 70),
                [],
            )
        ]
    )
    monkeypatch.setattr(PdfDocument, "table_extraction_confidence", property(lambda self: 100))

    findings = check_notes_agreement(document)

    assert not any("subtotal or total does not agree" in finding.issue.lower() for finding in findings)


def test_skipped_table_summary_groups_notes_arithmetic_skips(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\nTotal assets 100 90\nTotal equity and liabilities 100 90\n"
                + ("Primary context.\n" * 70),
                [],
            ),
            PdfPage(10, "Notes to the financial statements\n3 Revenue\nRevenue 100 90", [[["Revenue", "2025", "2024"], ["Fees", "60", "50"], ["Subscriptions", "40", "40"], ["Total", "100", "90"]]]),
            PdfPage(11, "4 Expenses\nExpenses 60 50", [[["Expenses", "2025", "2024"], ["Admin", "60", "50"], ["Total", "60", "50"]]]),
        ]
    )
    monkeypatch.setattr(PdfDocument, "table_extraction_confidence", property(lambda self: 100))
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions())
    summary = result.metrics["skipped_table_summary"]

    notes_group = next(row for row in summary if row["Skipped check group"] == "Notes tables - manual review recommended")
    assert notes_group["Pages affected"] == "Page 11"
    assert notes_group["Tables affected"] == "1"
    assert notes_group["Can automated check be fixed?"] == "Partially"
    assert "may merge" in notes_group["Why reviewer should review"].lower()


def test_simple_note_table_casts_when_structure_is_clear(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\nTotal assets 100 90\nTotal equity and liabilities 100 90\n"
                + ("Primary context.\n" * 70),
                [],
            ),
            PdfPage(
                10,
                "Notes to the financial statements\n3 Revenue\n",
                [[["Revenue", "2025", "2024"], ["Fees", "60", "55"], ["Subscriptions", "40", "35"], ["Total", "100", "90"]]],
            ),
        ]
    )
    monkeypatch.setattr(PdfDocument, "table_extraction_confidence", property(lambda self: 100))
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions())

    assert any(
        row["Result"] == "Passed" and "Simple note table on Page 10" in row["Check"]
        for row in result.metrics["check_results"]
    )
    assert not result.metrics["skipped_table_details"]


def test_simple_note_table_flags_wrong_total(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\nTotal assets 100 90\nTotal equity and liabilities 100 90\n"
                + ("Primary context.\n" * 70),
                [],
            ),
            PdfPage(
                10,
                "Notes to the financial statements\n3 Revenue\n",
                [[["Revenue", "2025", "2024"], ["Fees", "60", "55"], ["Subscriptions", "40", "35"], ["Total", "105", "90"]]],
            ),
        ]
    )
    monkeypatch.setattr(PdfDocument, "table_extraction_confidence", property(lambda self: 100))
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions())

    assert any(
        finding.category == "Totals and rounding"
        and finding.severity == "Medium"
        and "Simple note table total does not agree" in finding.issue
        for finding in result.findings
    )


def test_simple_note_text_section_casts_when_table_grid_is_poor(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\nTotal assets 100 90\nTotal equity and liabilities 100 90\n"
                + ("Primary context.\n" * 70),
                [],
            ),
            PdfPage(
                10,
                "\n".join(
                    [
                        "Notes to the financial statements",
                        "17. Employee costs",
                        "Basic 385,648 -",
                        "Other payroll levies 26,366 -",
                        "Other allowance 96,146 -",
                        "Pension cost 23,003 -",
                        "531,163 -",
                    ]
                ),
                [[["Financial Statements for the year ended", "31", "2025"], ["N '", "000", "000"], ["Header fragment", "", ""]]],
            ),
        ]
    )
    monkeypatch.setattr(PdfDocument, "table_extraction_confidence", property(lambda self: 100))
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions())

    assert any(
        row["Result"] == "Passed" and "Simple note section on Page 10" in row["Check"]
        for row in result.metrics["check_results"]
    )
    assert not result.metrics["skipped_table_details"]


def test_complex_note_table_is_not_casted(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\nTotal assets 100 90\nTotal equity and liabilities 100 90\n"
                + ("Primary context.\n" * 70),
                [],
            ),
            PdfPage(
                10,
                "Notes to the financial statements\n4 Trade and other payables\nMaturity analysis",
                [[["Trade and other payables", "2025", "2024"], ["Less than 30 days", "100", "90"], ["31 to 90 days", "20", "10"], ["Total", "120", "100"]]],
            ),
        ]
    )
    monkeypatch.setattr(PdfDocument, "table_extraction_confidence", property(lambda self: 100))
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions())

    assert not any("Simple note table on Page 10" in row["Check"] for row in result.metrics["check_results"])
    assert not any("Simple note table total does not agree" in finding.issue for finding in result.findings)


def test_front_matter_tables_are_not_reported_as_skipped_audit_checks(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                2,
                "General Information\nDirectors Lai Labode\nCompany registration number RC 12345",
                [[["Financial Statements for the year ended", "31", "2025"], ["FRC/", "2022", "861283"], ["Director", "Lai", "Labode"]]],
            ),
            PdfPage(
                13,
                "Statement of financial position\nTotal assets 100 90\nTotal equity and liabilities 100 90\n"
                + ("Primary context.\n" * 70),
                [],
            ),
            PdfPage(
                30,
                "Notes to the financial statements\n3 Revenue\nRevenue 100 90",
                [[["Revenue", "2025", "2024"], ["Fees", "100", "90"], ["Total", "100", "90"]]],
            ),
        ]
    )
    monkeypatch.setattr(PdfDocument, "table_extraction_confidence", property(lambda self: 100))
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions())
    details = result.metrics["skipped_table_details"]

    assert not any(row.get("Page") == "2" for row in details)
    assert any(row.get("Page") == "30" for row in details)


def test_share_capital_directors_report_table_casts_as_passed():
    document = PdfDocument(
        [
            PdfPage(
                5,
                "Directors' Report\n4. Share capital (continued)\n"
                "Issued\n"
                "FishBone & Lestr 995 995 994,560 994,560\n"
                "Lai Labode 1,027 1,027 1,027,442 1,027,442\n"
                "Cashbridge Global Leasing Company 772 772 772,136 772,136\n"
                "Owolabi Awosan Olusegun 465 465 464,694 464,694\n"
                "Daar Communications Limited 250 250 249,511 249,511\n"
                "Hatleys Consults Limited 202 202 201,545 201,545\n"
                "Banklink Africa Private Equity Limited 190 190 190,201 190,201\n"
                "Haylett Perry Limited 138 138 138,311 138,311\n"
                "Vidtesot Limited 93 93 93,113 93,113\n"
                "Obasi Chimaobi Ogbonnaya 84 84 84,286 84,286\n"
                "Adanijo Olanrewaju 82 82 82,392 82,392\n"
                "Okoraofor Ezichi Sapphire 64 64 64,295 64,295\n"
                "Olashore Taiwo Olagoke 51 51 50,923 50,923\n"
                "Mzer Michael Terungwa 26 26 26,339 26,339\n"
                "Adeoye Simileoluwa 26 26 26,339 26,339\n"
                "Isaac Olowokere 4 4 3,659 3,659\n"
                "Sokenu Evaristus Tolulope 1 1 896 896\n"
                "Sylvant Limited 3 3 2,886 2,886\n"
                "Claycounty 100 100 99,647 99,647\n"
                "Walter Castle Limited 225 225 225,057 225,057\n"
                "Kelmarid farms limited 202 202 201,768 201,768\n"
                "5,000 5,000 5,000,000 5,000,000\n",
                [],
            )
        ]
    )

    findings = reviewer.check_totals_and_rounding(document)

    assert any(f.severity == "Passed" and "Share capital/shareholding table" in f.issue for f in findings)
    assert not any(f.severity != "Passed" for f in findings)


def test_share_capital_directors_report_table_flags_wrong_total():
    document = PdfDocument(
        [
            PdfPage(
                5,
                "Directors' Report\nShare capital\nIssued\n"
                "Alpha Limited 100 100 100,000 100,000\n"
                "Beta Limited 200 200 200,000 200,000\n"
                "302 300 300,000 300,000\n",
                [],
            )
        ]
    )

    findings = reviewer.check_totals_and_rounding(document)

    assert any(
        f.category == "Totals and rounding"
        and f.severity == "Medium"
        and "Share capital or shareholding table total does not agree" in f.issue
        for f in findings
    )




def test_low_confidence_sfp_fallback_component_casting_is_skipped():
    page = PdfPage(
        7,
        "Statement of financial position\n"
        "Investment property 1,000 900\n"
        "Property plant and equipment 2,000 1,800\n"
        "Intangible assets 100 90\n"
        "Total non - current assets 3,500 2,790\n",
        [],
    )
    document = PdfDocument([page])

    findings, performed, skipped = reviewer._check_sfp_text(page, Decimal("1"), document=document)

    assert not findings
    assert not performed
    assert any("fallback component casting skipped" in item for item in skipped)


def test_share_capital_table_handles_dropped_repeated_small_amount():
    document = PdfDocument(
        [
            PdfPage(
                5,
                "Directors' Report\nShare capital\nIssued\n"
                "Alpha Limited 100 100 100,000 100,000\n"
                "Beta Limited 200 200 200,000 200,000\n"
                "Small Holder 3 2,886 2,886\n"
                "303 303 302,886 302,886\n",
                [],
            )
        ]
    )

    findings = reviewer.check_totals_and_rounding(document)

    assert any(f.severity == "Passed" and "Extraction corrections applied" in f.evidence for f in findings)
    assert not any(f.severity != "Passed" for f in findings)






def test_company_as_lessee_note_heading_is_valid():
    assert reviewer._valid_note_heading("4", "Leases (Company as lessee)")



def test_printed_page_map_interpolates_missing_statement_footer():
    document = PdfDocument(
        [
            PdfPage(1, "Cover", []),
            PdfPage(2, "Contents\nStatement of Changes in Equity 15\nStatement of Cash Flows 16\nFive-Year Financial Summary 44", []),
            PdfPage(15, "Statement of profit or loss\n14", []),
            PdfPage(16, "Statement of Changes in Equity\nBalance at 31 December 2025 100\n15", []),
            PdfPage(17, "Statement of Cash Flows\nCash at end of year 100", []),
            PdfPage(18, "Notes to the financial statements\n17", []),
        ]
    )

    assert reviewer._reviewer_page_number(document, 17) == 16
    assert reviewer._reviewer_page_number(document, 2) == 2


def test_changes_statement_page_is_inferred_from_contents_when_rotated():
    document = PdfDocument(
        [
            PdfPage(
                3,
                "Contents\nStatement of Financial Position 13\nStatement of Profit or Loss 14\n"
                "Statement of Changes in Equity 15\nStatement of Cash Flows 16",
                [],
            ),
            PdfPage(
                14,
                "Statement of Financial Position\nAssets\nTotal assets 100 90\nTotal equity and liabilities 100 90",
                [],
            ),
            PdfPage(
                15,
                "Statement of Profit or Loss\nRevenue 100 90\nProfit before tax 10 9\nTaxation (1) (1)\nProfit after tax 9 8",
                [],
            ),
            PdfPage(16, "(€8z‘6€L) = (p16‘E0rT) TTHPEE= i z pT8IZ7‘E6OLrTT", []),
            PdfPage(
                17,
                "Statement of Cash Flows\nCash flows from operating activities\nCash and cash equivalents at end of year 20 10",
                [],
            ),
        ]
    )

    page = reviewer._find_statement_page(document, "Statement of changes in equity")

    assert page is not None
    assert page.number == 16

def test_rotated_gibberish_page_is_marked_for_ocr_retry():
    gibberish = "(â‚¬8zâ€˜6â‚¬L) = (p16â€˜E0rT) TTHPEE= i z pT8IZ7â€˜E6OLrTT"
    clean = "Statement of Changes in Equity\nBalance at 1 January 2025 5,000 10,000"

    assert extraction._should_retry_with_rotated_ocr(gibberish)
    assert not extraction._should_retry_with_rotated_ocr(clean)


def test_changes_in_equity_check_includes_direct_equity_movements():
    page = PdfPage(
        16,
        "Statement of Changes in Equity\n"
        "Balance at 1 January 2025 5,000 1,670,895 1,675,895 179,284 (1,103,914) 751,265\n"
        "Profit for the year - - - - 341,832 341,832\n"
        "Contribution by owners of the Company - 57,298 57,298 - - 57,298\n"
        "Balance at 31 December 2025 5,000 1,728,193 1,733,193 179,284 (762,082) 1,150,395\n",
        [],
    )

    findings, performed, skipped = reviewer._check_accumulated_fund_text(
        page,
        Decimal("1"),
        False,
        PdfDocument([page]),
    )

    assert performed
    assert not skipped
    assert not findings


def test_simple_note_text_casting_can_infer_one_noisy_missing_amount_from_total():
    page = PdfPage(
        38,
        "Notes to the Financial Statements\n"
        "17. Operating expenses\n"
        "Advertising 421,366 188,606\n"
        "Entertainment 2,478 D123\n"
        "Travel expenses 199,281 148,968\n"
        "623,125 342,697\n",
        [],
    )

    findings = reviewer._check_simple_note_text_casting(page, Decimal("1"))

    assert findings
    assert findings[0].severity == "Passed"
    assert "5,123" in findings[0].evidence


def test_note_heading_inferred_when_extraction_drops_note_number():
    document = PdfDocument(
        [
            PdfPage(
                14,
                "Statement of Financial Position as at 31 December 2025\n"
                "2025 2024\nNote(s) N'000 N'000\n"
                "Deferred tax 9 13,629 13,284\n"
                "Trade and other payables 10 2,969,630 2,881,552\n"
                "Total Equity and Liabilities 4,544,782 3,898,493",
                [],
            ),
            PdfPage(
                35,
                "Notes to the Financial Statements\n2025 2024\nN'000 N'000\n"
                "Deferred tax\n"
                "The deferred tax assets and deferred tax liability relate to income tax.\n"
                "Deferred tax liability 13,629 13,284\n"
                "10. Trade and other payables\nTrade payables 427,359 634,025",
                [],
            ),
        ]
    )

    headings = reviewer._note_headings_by_page(document)
    rows = reviewer._note_agreement_result_rows(document)
    deferred_tax_row = next(row for row in rows if row["Line item description"] == "Deferred Tax")

    assert headings["9"] == ("Deferred tax", 35)
    assert deferred_tax_row["Review result"] == "Passed"
    assert deferred_tax_row["Current year amount found in referenced note?"] == "Yes"
    assert deferred_tax_row["Prior year amount found in referenced note?"] == "Yes"


def test_note_heading_detection_starts_after_notes_heading_when_present():
    document = PdfDocument(
        [
            PdfPage(1, "Directors' report\n7 Directors interests in shares\n8 Employment and employees", []),
            PdfPage(2, "Notes to the financial statements\n7 Revenue\nRevenue 100 90\n8 Cost of sales\nCost 50 40", []),
        ]
    )

    headings = _note_headings_by_page(document)

    assert headings["7"] == ("Revenue", 2)
    assert headings["8"] == ("Cost of sales", 2)
    assert "Directors interests" not in headings["7"][0]


def test_ocr_note_heading_detection_ignores_directors_report_before_notes_section():
    document = PdfDocument(
        [
            PdfPage(1, "Directors' report\n7 Directors' interests in shares\nEmployment and employees\n", []),
            PdfPage(2, "Notes to the financial statements\n1 Accounting policies\n2 Investment property\n7 Trade and other payables\n", []),
        ],
        ocr_used=True,
        ocr_pages=2,
    )

    headings = _note_headings_by_page(document)

    assert headings["7"] == ("Trade and other payables", 2)
    assert all("directors" not in title.lower() for title, _page in headings.values())


def test_ocr_note_headings_never_include_pages_before_notes_start_page():
    document = PdfDocument(
        [
            PdfPage(5, "Directors' report\n7 Directors' interests in shares\n11 Employment and employees", []),
            PdfPage(14, "Notes to the financial statements\n1 Accounting policies\n3 Investment property\n13 Revenue\n", []),
        ],
        ocr_used=True,
        ocr_pages=2,
    )

    headings = _note_headings_by_page(document)

    assert headings["13"] == ("Revenue", 14)
    assert "7" not in headings
    assert "11" not in headings


def test_note_headings_are_not_overwritten_from_statement_references():
    document = PdfDocument(
        [
            PdfPage(1, "Statement of profit or loss\nRevenue Note 13 707,189 297,041", []),
            PdfPage(14, "Notes to the financial statements\n3 Investment property\n13 Revenue from contracts with customers", []),
        ]
    )

    headings = _note_headings_by_page(document)

    assert headings["3"] == ("Investment property", 14)
    assert headings["13"] == ("Revenue from contracts with customers", 14)


def test_notes_start_page_requires_actual_notes_heading_not_front_matter_phrase(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(2, "Independent auditor's report\nThe notes to the financial statements are part of our audit.", []),
            PdfPage(14, "Notes to the financial statements\n1 Accounting policies\n13 Revenue", []),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert result.metrics["notes_section_start_page"] == 14


def test_ocr_note_validation_skips_when_notes_section_start_is_not_detected():
    document = PdfDocument(
        [
            PdfPage(1, "Statement of financial position\nCash Note 7 100 90", []),
            PdfPage(2, "Directors' report\n7 Directors' interests in shares", []),
        ],
        ocr_used=True,
        ocr_pages=2,
    )

    findings = check_notes_agreement(document, cautious_low_confidence=True)

    assert any("notes section start was not detected" in finding.issue.lower() for finding in findings)
    assert not any("possible wrong note reference" in finding.issue.lower() for finding in findings)


def test_review_pdf_does_not_report_ocr_note_validation_performed_without_notes_boundary(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of financial position\nCash Note 7 100 90", []),
            PdfPage(2, "Directors' report\n7 Directors' interests in shares", []),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert "Heading-based note-reference validation performed" not in result.metrics["checks_performed"]
    assert "OCR note-reference validation skipped because the notes section start was not detected." in result.metrics["checks_skipped"]
    assert result.metrics["note_validation_mode"] == "skipped"
    assert result.metrics["note_headings_detected"] == 0


def test_ocr_note_validation_status_stays_out_of_exception_register(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of financial position\nCash Note 7 100 90\n" + ("OCR statement context.\n" * 40), []),
            PdfPage(2, "Notes to the financial statements\n7 Cash and cash equivalents\nCash 100 90\n" + ("OCR note context.\n" * 40), []),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert "Heading-based note-reference validation performed in OCR review-prompt mode." in result.metrics["checks_performed"]
    assert not any("Heading-based note-reference validation was run" in finding.issue for finding in result.findings)


def test_ocr_notes_start_page_accepts_fuzzy_notes_heading_variants():
    document = PdfDocument(
        [
            PdfPage(2, "Independent auditor's report\nThe notes to the financial statements are part of our audit.", []),
            PdfPage(14, "NOTES FORMING PART OF THE FINANCIAL STATEMENTS\n1 Accounting policies\n13 Revenue", []),
        ],
        ocr_used=True,
        ocr_pages=2,
    )

    assert reviewer._notes_start_page(document) == 14
    assert "NOTES FORMING PART" in reviewer._format_notes_heading_snippet(document)


def test_ocr_notes_start_page_uses_strong_heading_candidate_with_note_heading_on_same_line():
    document = PdfDocument(
        [
            PdfPage(2, "Independent auditor's report\nThe notes to the financial statements are part of our audit.", []),
            PdfPage(14, "Notes to the Financial Statements 1. Significant accounting policies\n2 Revenue\nRevenue 100 90", []),
        ],
        ocr_used=True,
        ocr_pages=2,
    )

    assert reviewer._notes_start_page(document) == 14
    snippet = reviewer._format_notes_heading_snippet(document)
    assert "Confidence" in snippet
    assert "Significant accounting policies" in snippet
    headings = _note_headings_by_page(document)
    assert headings["1"] == ("Significant accounting policies", 14)
    debug = reviewer._format_note_heading_debug(document)
    assert "Confidence:" in debug
    assert "Source:" in debug


def test_ocr_note_heading_rejects_amount_table_rows_after_notes_start():
    document = PdfDocument(
        [
            PdfPage(
                14,
                "\n".join(
                    [
                        "Notes to the Financial Statements",
                        "1. Significant accounting policies",
                        "1 Taxation (226,684) 53,127",
                        "2 Revenue",
                    ]
                ),
                [],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )

    headings = _note_headings_by_page(document)

    assert headings["1"] == ("Significant accounting policies", 14)
    assert headings["2"] == ("Revenue", 14)


def test_ocr_note_headings_stop_before_value_added_and_five_year_summary_pages():
    document = PdfDocument(
        [
            PdfPage(14, "Notes to the Financial Statements\n1. Significant accounting policies\n7 Cash and cash equivalents", []),
            PdfPage(31, "Statement of value added\n1 Depreciation 10,000 9,000\n7 Notes to the Financial Statements", []),
            PdfPage(32, "Five-year financial summary\n8 Assets 13233 12000", []),
        ],
        ocr_used=True,
        ocr_pages=3,
    )

    headings = _note_headings_by_page(document)

    assert headings["1"] == ("Significant accounting policies", 14)
    assert headings["7"] == ("Cash and cash equivalents", 14)
    assert "8" not in headings


def test_note_heading_rejects_repeated_notes_header_as_title():
    document = PdfDocument([PdfPage(14, "Notes to the Financial Statements\n7 Notes to the Financial Statements\n7 Cash and cash equivalents", [])])

    headings = _note_headings_by_page(document)

    assert headings["7"] == ("Cash and cash equivalents", 14)


def test_ocr_note_headings_keep_clear_note_lines_and_reject_noise():
    document = PdfDocument(
        [
            PdfPage(
                14,
                "\n".join(
                    [
                        "Notes to the Financial Statements",
                        "1 Significant accounting policies",
                        "The financial statements are prepared under IFRS",
                        "3 Investment property",
                        "Investment property is measured at fair value",
                        "5 Trade and other receivables",
                        "5 Trade and other receivables 1,910,631 131,254",
                        "6 Other financial assets",
                        "7 Notes to the Financial Statements",
                        "7 Cash and cash equivalents",
                        "8 Share capital",
                        "9 Financial liabilities",
                        "10 Trade and other payables",
                        "12 Taxation",
                        "13 Revenue",
                        "14 Direct costs",
                        "15 Other operating income",
                        "16 Other operating gains",
                        "17 Administrative expenses",
                        "18 Finance cost",
                        "20 Related parties",
                        "21 Going concern",
                        "23 Financial instruments and risk management",
                        "23 Credit risk",
                        "23 Liquidity risk",
                        "24 This represents the advance payment made by customers",
                        "25 Financial instruments are accounted for at amortised cost",
                    ]
                ),
                [],
            ),
            PdfPage(31, "Statement of value added\n26 Depreciation 10,000 9,000", []),
            PdfPage(32, "Five-year financial summary\n27 Revenue 13233 12000", []),
        ],
        ocr_used=True,
        ocr_pages=3,
    )

    headings = _note_headings_by_page(document)

    expected = {
        "1": "Significant accounting policies",
        "3": "Investment property",
        "5": "Trade and other receivables",
        "6": "Other financial assets",
        "7": "Cash and cash equivalents",
        "8": "Share capital",
        "9": "Financial liabilities",
        "10": "Trade and other payables",
        "12": "Taxation",
        "13": "Revenue",
        "14": "Direct costs",
        "15": "Other operating income",
        "16": "Other operating gains",
        "17": "Administrative expenses",
        "18": "Finance cost",
        "20": "Related parties",
        "21": "Going concern",
        "23": "Financial instruments and risk management",
    }
    assert {ref: title for ref, (title, _page) in headings.items()} == expected


def test_ocr_note_headings_reject_financial_instrument_subheadings_without_replacing_main_heading():
    document = PdfDocument(
        [
            PdfPage(
                14,
                "\n".join(
                    [
                        "Notes to the Financial Statements",
                        "23 Financial instruments and risk management",
                        "23 Credit risk",
                        "23 Liquidity risk",
                        "23 Market risk",
                    ]
                ),
                [],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )

    headings = _note_headings_by_page(document)

    assert headings == {"23": ("Financial instruments and risk management", 14)}


def test_ocr_notes_heading_accepts_repeating_report_header_with_numbered_policy(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(3, "Statement of financial position\nTotal assets 100 90\nTotal equity and liabilities 100 90", []),
            PdfPage(
                14,
                "Example Limited\nFinancial statements for the year ended 31 December 2025\n"
                "Directors' and independent auditor's reports\n"
                "Notes to the Financial Statements\n"
                "1. Significant accounting policies\n2 Revenue\nRevenue 100 90",
                [],
            ),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert result.metrics["notes_section_start_page"] == 14
    assert result.metrics["note_structure_confidence"] != "0%"
    assert any(row["Accepted"] == "Yes" and row["Page"] == "14" for row in result.metrics["notes_heading_candidates"])
    assert _note_headings_by_page(document)["1"] == ("Significant accounting policies", 14)


def test_ocr_notes_start_infers_note_1_from_material_accounting_policies():
    document = PdfDocument(
        [
            PdfPage(3, "Statement of financial position\nTotal assets 100 90\nTotal equity and liabilities 100 90", []),
            PdfPage(
                14,
                "Financial statements for the year ended 31 December 2025\n"
                "Notes to the Financial Statements\n"
                "Material accounting policies\n"
                "The material accounting policies applied in preparing these financial statements are set out below.\n"
                "1.1 Basis of preparation\n"
                "2 New Standards and Interpretations",
                [],
            ),
        ],
        ocr_used=True,
        ocr_pages=2,
    )

    headings = _note_headings_by_page(document)

    assert headings["1"] == ("Material accounting policies", 14)


def test_ocr_notes_heading_candidates_reject_contents_page():
    document = PdfDocument(
        [
            PdfPage(
                2,
                "Statement of financial position 12\n"
                "Statement of comprehensive income 13\n"
                "Statement of cash flows 14\n"
                "Notes to the Financial Statements 15\n"
                "Value Added Statement 39\n"
                "Five Year Financial Summary 40",
                [],
            ),
            PdfPage(
                16,
                "Notes to the Financial Statements\n"
                "1. Material accounting policies\n"
                "2 New Standards and Interpretations",
                [],
            ),
        ],
        ocr_used=True,
        ocr_pages=2,
    )

    candidates = reviewer._notes_heading_candidate_rows(document)

    contents_row = next(row for row in candidates if row["Page"] == "2")
    assert contents_row["Accepted"] == "No"


def test_ocr_notes_start_page_searches_raw_page_text_with_broken_lines(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(3, "Statement of financial position\nTotal assets 100 90\nTotal equity and liabilities 100 90", []),
            PdfPage(
                14,
                "NOtes to\n the Financial\n Statements\n1 Significant accounting policies\n2 Revenue\nRevenue 100 90",
                [],
            ),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert result.metrics["notes_section_start_page"] == 14
    candidates = result.metrics["notes_heading_candidates"]
    assert any(row["Accepted"] == "Yes" and row["Page"] == "14" for row in candidates)
    assert any(row["Normalized snippet"] for row in candidates)


def test_ocr_notes_start_page_accepts_notes_to_accounts_variant_and_records_candidate_rows(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of financial position\nTotal assets 100 90\nTotal equity and liabilities 100 90", []),
            PdfPage(12, "Notes to the accounts\n1. Significant accounting policies\n2 Revenue\nRevenue 100 90", []),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert result.metrics["notes_section_start_page"] == 12
    candidates = result.metrics["notes_heading_candidates"]
    assert any(row["Accepted"] == "Yes" and row["Page"] == "12" for row in candidates)


def test_ocr_notes_candidate_scan_not_blocked_by_later_false_statement_page(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(3, "Statement of financial position\nTotal assets 100 90\nTotal equity and liabilities 100 90", []),
            PdfPage(14, "Notes to the Financial Statements\n1. Significant accounting policies\n2 Revenue", []),
            PdfPage(30, "Statement of financial position five year summary\nTotal assets 100 90", []),
        ],
        ocr_used=True,
        ocr_pages=3,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert result.metrics["notes_section_start_page"] == 14


def test_ocr_notes_start_page_falls_back_to_significant_accounting_policies_after_statements(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(3, "Statement of financial position\nTotal assets 100 90\nTotal equity and liabilities 100 90", []),
            PdfPage(14, "1. Significant accounting policies\n2 Revenue\nRevenue 100 90", []),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert result.metrics["notes_section_start_page"] == 14
    assert any("Significant accounting policies" in row["Raw OCR snippet"] and row["Accepted"] == "Yes" for row in result.metrics["notes_heading_candidates"])
    assert any(row["Normalized snippet"] for row in result.metrics["notes_heading_candidates"])


def test_printed_footer_page_numbers_are_used_for_notes_metrics(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(28, "Notes to the Financial Statements\n14. Taxation\nCurrent tax 10 9\n27\nDRAFT", []),
            PdfPage(29, "15. Revenue\nRevenue 100 90\n28\nDRAFT", []),
        ]
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions())

    assert result.metrics["printed_page_map"]["28"] == 27
    assert result.metrics["notes_section_start_page"] == 27
    assert "Note 14 | Page 27 | Page 27 | Taxation" in result.metrics["note_headings"]


def test_printed_footer_page_number_detects_bottom_page_before_draft():
    assert reviewer._printed_footer_page_number("Some text\n27\nDRAFT") == 27
    assert reviewer._printed_footer_page_number("Some text\nPage 27") == 27


def test_clean_note_title_repairs_common_ocr_heading_typos():
    assert reviewer._clean_note_title("Trade and othet payables") == "Trade and other payables"
    assert reviewer._clean_note_title("Depreciation and amottisation") == "Depreciation and amortisation"


def test_notes_heading_candidates_include_rejected_diagnostic_rows(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of financial position\nTotal assets 100 90\nTotal equity and liabilities 100 90\n" + ("OCR filler\n" * 80), []),
            PdfPage(10, "Notes and financial statement extracts\nNarrative only\n" + ("OCR filler\n" * 80), []),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert any(row["Accepted"] == "No" for row in result.metrics["notes_heading_candidates"])


def test_excel_export_wires_notes_heading_candidates_sheet():
    app_source = Path("app.py").read_text(encoding="utf-8")

    assert 'sheet_name="Notes heading candidates"' in app_source
    assert 'writer.book["Notes heading candidates"]' in app_source


def test_ai_finding_review_normalizes_physical_pages_to_printed_pages():
    document = PdfDocument(
        [
            PdfPage(15, "Statement of Changes in Equity\n14\nDRAFT", []),
            PdfPage(16, "Notes to the Financial Statements\n1.Material accounting policies\n15\nDRAFT", []),
        ]
    )
    finding = reviewer.Finding(
        category="Presentation",
        severity="Medium",
        location="Page 15 | Statement of cash flows",
        issue="Statement of cash flows page carries incorrect 'Statement of Changes in Equity' heading.",
        evidence="The cash flow statement appears to be presented under a 'Statement of Changes in Equity' title.",
        recommendation="Update the heading.",
    )

    result = ai_finding_review.run_ai_finding_review(document, CompanyProfile(), [finding], model="gpt-5-mini")

    # no api key path should short-circuit before this point in normal runs, so inspect candidate helpers directly
    page_ref = ai_finding_review._finding_page_reference(document, finding, {})
    snippet = ai_finding_review._page_snippet(document, [15], ["statement", "cash", "flows"], "")

    assert page_ref == "Page 14"
    assert snippet.startswith("Page 14:")


def test_ocr_heading_based_note_reference_validation_runs_as_review_prompt(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, ("Statement of profit or loss\nOther Revenue Note 7 307,482 189,751\n" * 20), []),
            PdfPage(2, "Notes to the financial statements\n7 Operating revenue\nRevenue 100 90\n9 Other Revenue\nOther revenue 307,482 189,751", []),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert result.metrics["note_validation_mode"] == "review_prompt"
    row = next(row for row in result.metrics["note_agreement_results"] if row["Line item"] == "Other Revenue")
    assert row["Alternative note found"] == "9"
    assert row["Result"] == "Review prompt"
    assert not any("possible wrong note reference" in finding.issue.lower() for finding in result.findings)


def test_ifrs_16_checklist_triggers_from_actual_lease_balance_disclosure():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Accounting policies\nIFRS 16 is applied to lease liability balances at commencement date.\nCash and cash equivalents 500",
                [],
            )
        ]
    )

    findings = check_standard_checklist(document, CompanyProfile(checklist_areas=("IFRS 16",)))

    assert any("IFRS 16" in finding.location for finding in findings)


def test_policy_check_does_not_trigger_lease_policy_from_generic_contract_wording():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Accounting policies\nIFRS 16 is applied to lease contracts at commencement date.\nCash and cash equivalents 500",
                [],
            )
        ]
    )

    findings = check_policy_relevance(document, CompanyProfile())

    assert not any("leases policy" in finding.issue.lower() for finding in findings)


def test_policy_check_flags_industry_mismatch_and_superseded_standard():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Accounting policies\nBiological assets are measured under IAS 41.\nLeases are accounted for under IAS 17.",
                [],
            )
        ]
    )

    findings = check_policy_relevance(document, CompanyProfile(industry="Technology"))

    assert any("inconsistent with the stated industry" in finding.issue.lower() for finding in findings)
    assert any("superseded" in finding.issue.lower() for finding in findings)


def test_currency_check_ignores_generic_policy_currency_words():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Accounting policies\nForeign currency transactions may be denominated in Dollar or Euro as examples.\nStatement of financial position\nPresented in N'000",
                [],
            )
        ]
    )

    findings = check_formatting(document, CompanyProfile())

    assert not any("currency marker" in finding.issue.lower() for finding in findings)


def test_ocr_formatting_ignores_standalone_fragments_in_unstructured_summary():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Five-year financial summary\nRevenue growth history\n13233\nRevenue 13233 12000\n2021 2022 2023\n",
                [],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )

    findings = check_formatting(document, CompanyProfile(reporting_currency="NGN"))

    assert not any("thousands separators" in finding.issue.lower() for finding in findings)


def test_ocr_formatting_ignores_financial_review_fragments_without_reliable_table():
    document = PdfDocument(
        [
            PdfPage(
                32,
                "5 year financial review\nRevenue 13233 12000\nAssets 55555 44444\n",
                [],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )

    findings = check_formatting(document, CompanyProfile(reporting_currency="NGN"))

    assert not any("thousands separators" in finding.issue.lower() for finding in findings)


def test_ocr_formatting_ignores_low_confidence_summary_table_fragments():
    document = PdfDocument(
        [
            PdfPage(
                32,
                "Five-year financial summary",
                [
                    [
                        ["Revenue", "13233 12000"],
                        ["Assets", "55555 44444"],
                    ]
                ],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )

    findings = check_formatting(document, CompanyProfile(reporting_currency="NGN"))

    assert not any("thousands separators" in finding.issue.lower() for finding in findings)


def test_formatting_flags_unclear_share_capital_unit_headings():
    document = PdfDocument(
        [
            PdfPage(
                4,
                "Directors' Report\n"
                "6. Share capital\n"
                "2025 2024 2025 2024\n"
                "Issued N '000 N '000 Number of shares\n"
                "Greystone One Holdings Limited 1,250,000 1,250,000 1,250,000 1,250,000\n"
                "Greystone Two Holdings Limited 1,250,000 1,250,000 1,250,000 1,250,000\n"
                "2,500,000 2,500,000 2,500,000 2,500,000\n",
                [],
            )
        ]
    )

    findings = check_formatting(document, CompanyProfile())

    issue = next((finding for finding in findings if "unit headings" in finding.issue.lower()), None)
    assert issue is not None
    assert issue.location == "Page 4"
    assert issue.severity == "Low"
    assert issue.metadata["check_type"] == "share_capital_presentation"
    assert "N '000" in issue.evidence
    assert "Number of shares" in issue.evidence


def test_formatting_does_not_flag_clear_share_capital_unit_headings():
    document = PdfDocument(
        [
            PdfPage(
                4,
                "Share capital\n"
                "Issued | 2025 N '000 | 2024 N '000 | 2025 Number of shares | 2024 Number of shares\n"
                "Greystone One Holdings Limited 1,250,000 1,250,000 1,250,000 1,250,000\n"
                "Greystone Two Holdings Limited 1,250,000 1,250,000 1,250,000 1,250,000\n"
                "Total 2,500,000 2,500,000 2,500,000 2,500,000\n",
                [],
            )
        ]
    )

    findings = check_formatting(document, CompanyProfile())

    assert not any("unit headings" in finding.issue.lower() for finding in findings)


def test_currency_normalization_accepts_naira_aliases_and_rejects_invalid_code():
    assert normalize_reporting_currency("NGN") == "NGN"
    assert normalize_reporting_currency("Naira") == "NGN"
    assert normalize_reporting_currency("₦") == "NGN"
    assert normalize_reporting_currency("N’000") == "NGN"
    assert normalize_reporting_currency("N '000") == "NGN"
    assert normalize_reporting_currency("N ‘000") == "NGN"
    assert normalize_reporting_currency("₦’000") == "NGN"
    assert normalize_reporting_currency("NGN’000") == "NGN"
    assert normalize_reporting_currency("NGN / N'000") == "NGN"
    assert normalize_reporting_currency("N000") == "NGN"
    assert normalize_reporting_currency("NGB") == ""


def test_detected_profile_infers_upload_only_context():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "NATIONAL INSTITUTE OF PROFESSIONAL ADMINISTRATORS OF NIGERIA\nFinancial Statements\nYear ended December 31, 2025\nPrepared in accordance with IFRS and presented in N'000.\nPrincipal activities are professional membership services, professional development, training and certification for members and associates. This paragraph continues with extracted report boilerplate that should not be pasted in full.\nSubscriptions 200\nMembers fund 300\nCash and cash equivalents 100\nTrade and other receivables 50",
                [],
            )
        ]
    )

    profile = infer_detected_profile(document)

    assert profile["Company name"] == "National Institute of Professional Administrators of Nigeria"
    assert profile["Year end"] == "December 31 2025"
    assert profile["Currency"] == "NGN / N'000"
    assert profile["Framework"] == "IFRS"
    assert "professional body" in profile["Entity type"].lower()
    assert profile["Principal activities"] == "Professional membership body, including member services, professional development, training, and certification."


def test_detected_profile_classifies_limited_property_company_before_professional_body_terms():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Example Property Investments Limited",
                        "Directors' report",
                        "The company has share capital and directors.",
                        "The principal activity is property investment.",
                        "Investment property 1,000 900",
                        "Professional advisers are listed below.",
                    ]
                ),
                [],
            )
        ]
    )

    profile = infer_detected_profile(document)

    assert profile["Entity type"] == "Private company / property investment company"


def test_generic_professional_body_pattern_detects_membership_entity_without_company_name_hardcode():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "SYNTHETIC COUNCIL OF REGISTERED PROFESSIONALS",
                        "Statement of income and expenditure",
                        "Subscriptions 1,200 1,100",
                        "Members' fund 5,000 4,700",
                        "Accumulated fund 5,000 4,700",
                        "Principal activities are membership services, training, certification and professional development for fellows and associates.",
                    ]
                ),
                [],
            )
        ]
    )

    profile = infer_detected_profile(document)

    assert profile["Company name"] == "Synthetic Council of Registered Professionals"
    assert profile["Entity type"] == "Non-profit / professional body"
    assert "Professional membership body" in profile["Principal activities"]


def test_generic_scanned_private_company_pattern_detects_company_and_runs_sfp_check(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Example Property Investments Limited",
                        "Directors' report",
                        "The company has ordinary shares and share capital.",
                    ]
                ),
                [],
            ),
            PdfPage(
                2,
                "\n".join(
                    [
                        "Statement of financial position",
                        "Non-current assets 600 500",
                        "Current assets 400 300",
                        "Total assets 1,000 800",
                        "Equity 700 550",
                        "Liabilities 300 250",
                        "Total equity and liabilities 1,000 800",
                    ]
                )
                + "\n"
                + ("Additional OCR statement text for coverage.\n" * 80),
                [],
            ),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    profile = infer_detected_profile(document)
    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert profile["Entity type"] == "Private company / property investment company"
    assert "Statement of financial position: total assets checked" in result.metrics["checks_performed"]
    assert any(row["Result"] == "Passed" and "total assets" in row["Check"].lower() for row in result.metrics["check_results"])


def test_ocr_statement_rows_include_row_confidence(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\nCash and cash equivalents 39,387 193,627\nTotal assets 39,387 193,627\nTotal equity and liabilities 39,387 193,627\n"
                + ("Additional OCR statement text for coverage.\n" * 80),
                [],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert "cash and cash equivalents" in result.metrics["ocr_statement_rows"]
    assert "Low-Medium" in result.metrics["ocr_statement_rows"]


def test_text_confidence_separates_table_confidence():
    noisy_table = [["Description", "2025", "2024"], ["Revenue", "100 90", ""], ["Total", "100", "90"]]
    document = PdfDocument(
        [PdfPage(1, "Statement of financial position\n" + ("Revenue 100 90\n" * 120), [noisy_table])]
    )

    assert document.extraction_confidence >= 90
    assert document.table_extraction_confidence < document.extraction_confidence


def test_formatting_ignores_registration_numbers_and_nonfinancial_negatives():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "FRC registration number 816300\nCertificate number 375217\nAddress 12345 Lagos\nThis text has extraction artefact -12345",
                [],
            )
        ]
    )

    findings = check_formatting(document, CompanyProfile())

    assert not any("thousands separators" in finding.issue.lower() for finding in findings)
    assert not any("negative" in finding.issue.lower() for finding in findings)


def test_superseded_standard_context_classifies_current_policy_as_high_and_transition_as_low():
    current_policy = PdfDocument(
        [PdfPage(1, "Accounting policies\nFinancial instruments are accounted for under IAS 39.", [])]
    )
    transition_note = PdfDocument(
        [PdfPage(2, "New and amended standards\nIAS 39 was replaced by IFRS 9 and is discussed as transition history.", [])]
    )

    current_findings = check_policy_relevance(current_policy, CompanyProfile())
    transition_findings = check_policy_relevance(transition_note, CompanyProfile())

    assert any(finding.severity == "High" and "Page 1" in finding.location for finding in current_findings)
    assert any(finding.severity == "Low" and "Page 2" in finding.location for finding in transition_findings)
    assert all("Context:" in finding.evidence for finding in current_findings if finding.category == "Accounting policies")


def test_consolidation_policy_is_not_triggered_by_generic_group_wording():
    document = PdfDocument(
        [PdfPage(1, "Accounting policies\nThe group of standards issued by the IASB is reviewed annually.", [])]
    )

    findings = check_policy_relevance(document, CompanyProfile())

    assert not any("consolidation-related" in finding.issue.lower() for finding in findings)


def test_revenue_policy_recognises_revenue_from_contracts_with_customers():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of profit or loss\nRevenue 1,000 900\nNotes to the financial statements\nAccounting policies\nRevenue from contracts with customers is recognised when control transfers.",
                [],
            )
        ]
    )

    findings = check_policy_relevance(document, CompanyProfile())

    assert not any("revenue-related balances" in finding.issue.lower() for finding in findings)


def test_revenue_policy_recognises_ocr_variant_revenue_from_contract_with_customer():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of profit or loss\nRevenue 1,000 900\nNotes to the financial statements\nAccounting policies\nRevenue from contract with customer is recognised when control transfers.",
                [],
            )
        ]
    )

    policy_findings = check_policy_relevance(document, CompanyProfile())
    checklist_findings = check_standard_checklist(document, CompanyProfile())

    assert not any("revenue-related balances" in finding.issue.lower() for finding in policy_findings)
    assert not any("IFRS 15" in finding.location for finding in checklist_findings)


def test_revenue_policy_recognises_broken_ocr_revenue_contracts_customers_phrase():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of profit or loss\nRevenue 1,000 900\nNotes to financial statements\nAccounting policies\nRevenue contracts customers recognised when control transfers.",
                [],
            )
        ]
    )

    policy_findings = check_policy_relevance(document, CompanyProfile())

    assert not any("revenue-related balances" in finding.issue.lower() for finding in policy_findings)


def test_revenue_policy_recognises_split_ocr_revenue_from_contracts_with_customers_phrase():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of profit or loss\nRevenue 1,000 900\nNotes to financial statements\n"
                "Accounting policies\nRevenue from\ncontracts with\ncustorners is recognised when control transfers.",
                [],
            )
        ]
    )

    policy_findings = check_policy_relevance(document, CompanyProfile())
    checklist_findings = check_standard_checklist(document, CompanyProfile())

    assert not any("revenue-related balances" in finding.issue.lower() for finding in policy_findings)
    assert not any("IFRS 15" in finding.location for finding in checklist_findings)


def test_ifrs_16_not_triggered_by_generic_standards_amendments_text():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Notes to the financial statements\nNew standards and amendments\nIFRS 16 Leases amendments are effective for annual periods beginning after the reporting date.",
                [],
            )
        ]
    )

    policy_findings = check_policy_relevance(document, CompanyProfile())
    checklist_findings = check_standard_checklist(document, CompanyProfile())

    assert not any("leases policy" in finding.issue.lower() for finding in policy_findings)
    assert not any("IFRS 16" in finding.location for finding in checklist_findings)


def test_ifrs_16_not_triggered_by_right_of_use_wording_inside_amendments_section():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Notes to the financial statements\nNew standards and amendments\nAmendments mention right-of-use asset and lease liability examples effective for annual periods beginning after the reporting date.",
                [],
            )
        ]
    )

    policy_findings = check_policy_relevance(document, CompanyProfile())
    checklist_findings = check_standard_checklist(document, CompanyProfile())

    assert not any("leases policy" in finding.issue.lower() for finding in policy_findings)
    assert not any("IFRS 16" in finding.location for finding in checklist_findings)


def test_ifrs_16_not_triggered_by_generic_deferred_tax_single_transaction_text():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Notes to the financial statements\nNew standards and amendments\n"
                "Deferred tax related to assets and liabilities arising from a single transaction includes examples for right-of-use asset and lease liability.",
                [],
            )
        ]
    )

    policy_findings = check_policy_relevance(document, CompanyProfile())
    checklist_findings = check_standard_checklist(document, CompanyProfile())

    assert not any("leases policy" in finding.issue.lower() for finding in policy_findings)
    assert not any("IFRS 16" in finding.location for finding in checklist_findings)


def test_ifrs_16_not_triggered_by_generic_new_and_amended_standards_text():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Notes to the financial statements\nNew and amended standards issued but not effective\n"
                "The amendments to IFRS 16 discuss lease liability and right-of-use asset measurement examples.",
                [],
            )
        ]
    )

    policy_findings = check_policy_relevance(document, CompanyProfile())
    checklist_findings = check_standard_checklist(document, CompanyProfile())

    assert not any("leases policy" in finding.issue.lower() for finding in policy_findings)
    assert not any("IFRS 16" in finding.location for finding in checklist_findings)


def test_ifrs_16_not_triggered_by_theoretical_lease_asset_liability_wording_even_if_forced():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Notes to the financial statements\nNew standards and deferred tax amendments\n"
                "The recognition of a lease asset and lease liability may arise from theoretical IFRS 16 examples.",
                [],
            )
        ]
    )

    policy_findings = check_policy_relevance(document, CompanyProfile())
    checklist_findings = check_standard_checklist(document, CompanyProfile(checklist_areas=("IFRS 16",)))

    assert not any("leases policy" in finding.issue.lower() for finding in policy_findings)
    assert not any("IFRS 16" in finding.location for finding in checklist_findings)


def test_ifrs_16_triggers_from_actual_lease_liability_amount():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Notes to the financial statements\nLease liabilities\nCurrent lease liability 125\nNon-current lease liability 430",
                [],
            )
        ]
    )

    findings = check_standard_checklist(document, CompanyProfile())

    assert any("IFRS 16" in finding.location for finding in findings)


def test_ifrs_16_triggers_from_actual_lease_arrangement_disclosure():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Notes to the financial statements\nThe company leases office premises under a cancellable lease arrangement.",
                [],
            )
        ]
    )

    checklist_findings = check_standard_checklist(document, CompanyProfile())

    assert any("IFRS 16" in finding.location for finding in checklist_findings)


def test_ifrs_16_not_triggered_by_generic_lease_contract_policy_wording():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Accounting policies\nIFRS 16 is applied to lease contracts at commencement date.",
                [],
            )
        ]
    )

    policy_findings = check_policy_relevance(document, CompanyProfile())
    checklist_findings = check_standard_checklist(document, CompanyProfile())

    assert not any("leases policy" in finding.issue.lower() for finding in policy_findings)
    assert not any("IFRS 16" in finding.location for finding in checklist_findings)


def test_detected_profile_does_not_suggest_ifrs_16_from_generic_new_standards_text():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Notes to the financial statements\nNew standards and amendments\nIFRS 16 right-of-use asset and lease liability examples are discussed as amendments.",
                [],
            )
        ]
    )

    profile = infer_detected_profile(document)

    assert "IFRS 16" not in profile["Suggested checklist areas"]


def test_cash_flow_net_increase_row_is_not_misread_as_note_reference_split_digit():
    row = _parse_ocr_statement_row("Net increase in cash and cash equivalents 2 70,224 2 49,055")

    assert row is not None
    assert row.note_ref == ""
    assert row.amounts == (Decimal("270224"), Decimal("249055"))


def test_ocr_sfp_statement_specific_checks_run_only_on_confident_rows():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\n" + ("Non-current assets current assets total assets equity liabilities\n" * 80),
                [
                    [
                        ["Description", "2025"],
                        ["Non-current assets", "100"],
                        ["Current assets", "50"],
                        ["Total assets", "153"],
                        ["Equity", "60"],
                        ["Liabilities", "90"],
                        ["Total equity and liabilities", "150"],
                    ]
                ],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )

    findings = check_rounding_and_casting(document, tolerance=Decimal("1"))

    assert any("Non-current assets + current assets" in finding.issue for finding in findings)


def test_ocr_sfp_line_based_checks_run_when_tables_are_not_structured():
    document = PdfDocument(
        [
            PdfPage(
                1,
                ("Statement of financial position\n" * 40)
                + "\n".join(
                    [
                        "Non-current assets 100 90",
                        "Current assets 50 40",
                        "Total assets 153 130",
                        "Equity 60 50",
                        "Liabilities 90 80",
                        "Total equity and liabilities 150 130",
                    ]
                ),
                [],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )

    findings = check_rounding_and_casting(document, tolerance=Decimal("1"))

    assert any("Non-current assets + current assets" in finding.issue for finding in findings)


def test_confidence_metrics_separate_ocr_text_from_statement_structure(monkeypatch):
    document = PdfDocument(
        [PdfPage(1, "OCR text without parseable statement rows\n" * 80, [])],
        ocr_used=True,
        ocr_pages=1,
        ocr_tables=3,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert result.metrics["ocr_text_coverage"] == "100%"
    assert result.metrics["statement_structure_confidence"] == "0%"
    assert result.metrics["table_arithmetic_confidence"] == "0%"
    assert result.metrics["ocr_tables"] == 3


def test_statement_structure_confidence_reflects_checkable_rows_not_headings_only(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of profit or loss\nRevenue text without amounts\n" + ("OCR filler\n" * 80), []),
            PdfPage(2, "Statement of financial position\nTotal assets 500 400\n" + ("OCR filler\n" * 80), []),
            PdfPage(3, "Statement of changes in equity\nMovement narrative only\n" + ("OCR filler\n" * 80), []),
            PdfPage(4, "Statement of cash flows\nCash flow narrative only\n" + ("OCR filler\n" * 80), []),
        ],
        ocr_used=True,
        ocr_pages=4,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert int(result.metrics["statement_structure_confidence"].rstrip("%")) < 40
    assert "Statement of financial position: equity and liabilities equation checked" not in result.metrics["checks_performed"]


def test_fuzzy_ocr_statement_page_detection_classifies_distorted_headings(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statememt of financiaI position\nTotal assets 100 90\nTotal equity and liabilities 100 90\n" * 20, []),
            PdfPage(2, "Statment of profit or Ioss\nRevenue 100 90\nProfit 10 8\n" * 20, []),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert "Statement of financial position | Page 1" in result.metrics["primary_statement_pages"]
    assert "Statement of income and expenditure | Page 2" in result.metrics["primary_statement_pages"]
    assert result.metrics["statement_structure_confidence"] != "0%"


def test_review_report_includes_ocr_statement_rows_debug(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of financial position",
                        "Non current assets 600 500",
                        "Current assets 400 300",
                        "Total assets 1,000 800",
                        "Equity 700 550",
                        "Liabilities 300 250",
                        "Total equity and liability 1,000 800",
                    ]
                )
                + "\n"
                + ("Additional OCR statement text for coverage.\n" * 80),
                [],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert "non-current assets" in result.metrics["ocr_statement_rows"]
    assert "total equity and liabilities" in result.metrics["ocr_statement_rows"]
    assert "Statement of financial position: equity and liabilities equation checked" in result.metrics["checks_performed"]


def test_ocr_statement_rows_ignore_title_and_date_lines(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Example Property Investments Limited financial statements for the year ended 31 December 2021 2020",
                        "Statement of profit or loss and other comprehensive income",
                        "Revenue 707,189 297,041",
                        "Loss before taxation (173,516) (681,559)",
                        "Taxation 253,124",
                        "Loss after taxation (120,389) (428,435)",
                    ]
                )
                + "\n"
                + ("Additional OCR statement text for coverage.\n" * 80),
                [],
            ),
            PdfPage(
                2,
                "\n".join(
                    [
                        "Statement of financial position as at 31 December 2021 2020",
                        "Non-current assets 1,200,000 1,100,000",
                        "Current assets 300,000 250,000",
                        "Total assets 1,500,000 1,350,000",
                        "Equity 900,000 850,000",
                        "Liabilities 600,000 500,000",
                        "Trade and other receivables 1,910,631 131,254",
                        "Cash and cash equivalents 739,387 193,627",
                        "Financial liabilities 5,356,392 4,555,742",
                        "Total equity and liabilities 1,500,000 1,350,000",
                    ]
                )
                + "\n"
                + ("Additional OCR statement text for coverage.\n" * 80),
                [],
            ),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    rows = result.metrics["ocr_statement_rows"]

    assert "revenue |  | 707,189 | 297,041" in rows
    assert "profit before tax |  | -173,516 | -681,559" in rows
    assert "taxation |  | 253,124" in rows
    assert "profit after tax |  | -120,389 | -428,435" in rows
    assert "non-current assets |  | 1,200,000 | 1,100,000" in rows
    assert "trade and other receivables |  | 1,910,631 | 131,254" in rows
    assert "cash and cash equivalents |  | 739,387 | 193,627" in rows
    assert "financial liabilities |  | 5,356,392 | 4,555,742" in rows
    assert "total equity and liabilities |  | 1,500,000 | 1,350,000" in rows
    assert "financial statements for the year ended" not in rows.lower()
    assert "Statement of financial position: equity and liabilities equation checked" in result.metrics["checks_performed"]


def test_ocr_statement_rows_ignore_table_of_contents_pages(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                2,
                "\n".join(
                    [
                        "Contents",
                        "Statement of profit or loss ........ 10",
                        "Statement of financial position .... 11",
                        "Statement of changes in equity ..... 12",
                        "Statement of cash flows ............ 13",
                        "Equity 11",
                    ]
                ),
                [],
            ),
            PdfPage(
                11,
                "\n".join(
                    [
                        "Statement of financial position",
                        "Investment property 4,900,000 4,850,000",
                        "Trade and other receivables 1,910,631 131,254",
                        "Cash and cash equivalents 739,387 193,627",
                        "Total assets 7,550,018 5,174,881",
                        "Equity 2,193,626 619,139",
                        "Financial liabilities 5,356,392 4,555,742",
                        "Total equity and liabilities 7,550,018 5,174,881",
                    ]
                )
                + "\n"
                + ("Additional OCR statement text for coverage.\n" * 80),
                [],
            ),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    rows = result.metrics["ocr_statement_rows"]

    assert "equity | 11" not in rows.lower()
    assert "Statement of financial position | Page 11" in rows
    assert "financial liabilities |  | 5,356,392 | 4,555,742" in rows


def test_ocr_statement_rows_drive_partial_statement_checks(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of profit or loss",
                        "Revenue 707,189 297,041",
                        "Loss before taxation (173,516) (681,559)",
                        "Taxation 53,127 253,124",
                        "Loss after taxation (120,389) (428,435)",
                    ]
                )
                + "\n"
                + ("Additional OCR statement text for coverage.\n" * 80),
                [],
            ),
            PdfPage(
                2,
                "\n".join(
                    [
                        "Statement of financial position",
                        "Investment property 4,900,000 4,850,000",
                        "Total non-current assets 4,900,000 4,850,000",
                        "Trade and other receivables 1,910,631 131,254",
                        "Cash and cash equivalents 739,387 193,627",
                        "Total current assets 2,650,018 324,881",
                        "Total assets 7,550,018 5,174,881",
                        "Equity 2,193,626 619,139",
                        "Financial liabilities 5,356,392 4,555,742",
                        "Total equity and liabilities 7,550,018 5,174,881",
                    ]
                )
                + "\n"
                + ("Additional OCR statement text for coverage.\n" * 80),
                [],
            ),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert "Income statement: revenue, tax, and profit/loss after tax checked" in result.metrics["checks_performed"]
    assert "Statement of financial position: total assets checked from extracted asset rows." in result.metrics["checks_performed"]
    assert "Statement of financial position: equity and liabilities equation checked" in result.metrics["checks_performed"]
    assert "Statement-specific OCR checks were skipped" not in "\n".join(finding.issue for finding in result.findings)
    assert result.metrics["checks_passed_count"] >= 2
    check_results = result.metrics["check_results"]
    assert any(
        row["Result"] == "Passed" and "total assets" in row["Check"].lower()
        for row in check_results
    )


def test_ocr_income_statement_runs_current_year_only_when_tax_row_has_one_amount(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of profit or loss",
                        "Revenue 707,189 297,041",
                        "Loss before taxation (173,516) (681,559)",
                        "Taxation 253,124",
                        "Loss after taxation (120,389) (428,435)",
                    ]
                )
                + "\n"
                + ("Additional OCR statement text for coverage.\n" * 80),
                [],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert "Income statement: revenue, tax, and profit/loss after tax checked" in result.metrics["checks_performed"]
    tax_findings = [finding for finding in result.findings if "Profit/loss after tax" in finding.issue]
    assert tax_findings
    assert all(finding.severity in {"Medium", "Low"} for finding in tax_findings)
    assert all("Possible mismatch from OCR line extraction" in finding.issue for finding in tax_findings)
    assert "Loss/profit before taxation raw line" in tax_findings[0].evidence
    assert "Taxation raw line" in tax_findings[0].evidence
    assert "Extracted values" in tax_findings[0].evidence


def test_ocr_income_statement_prompt_includes_corroborating_report_values(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Directors' report\nThe loss before taxation was (173,516). Tax credit was 53,127. Loss after taxation was (120,389).",
                [],
            ),
            PdfPage(
                2,
                "\n".join(
                    [
                        "Statement of profit or loss",
                        "Revenue 707,189 297,041",
                        "Loss before taxation (473,516) (681,559)",
                        "Taxation 53,127",
                        "Loss after taxation (120,389) (428,435)",
                    ]
                )
                + "\n"
                + ("Additional OCR statement text for coverage.\n" * 80),
                [],
            ),
            PdfPage(3, "Statement of cash flows\nLoss before taxation (173,516)\nTaxation 53,127", []),
        ],
        ocr_used=True,
        ocr_pages=3,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    tax_findings = [finding for finding in result.findings if "Profit/loss after tax" in finding.issue]

    assert tax_findings
    assert tax_findings[0].severity == "Low"
    assert tax_findings[0].issue.startswith("OCR conflict - manual confirmation required")
    assert "Corroborating OCR values" in tax_findings[0].evidence
    assert "Corroboration assessment" in tax_findings[0].evidence
    assert "before tax=+1" not in tax_findings[0].evidence
    assert "before tax=(173,516)" in tax_findings[0].evidence
    assert "Page 1" in tax_findings[0].evidence
    assert "Page 3" in tax_findings[0].evidence


def test_ocr_income_incomplete_current_year_after_tax_uses_clear_manual_confirmation_wording(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Directors' report\nThe loss before taxation was (221,494). Taxation was (226,684). Loss after taxation was (448,178).",
                [],
            ),
            PdfPage(
                2,
                "\n".join(
                    [
                        "Statement of profit or loss",
                        "Revenue 707,189 297,041",
                        "Loss before taxation (221,494) (173,516)",
                        "Taxation (226,684) 53,127",
                        "Loss for the year (999,999)",
                    ]
                )
                + "\n"
                + ("Additional OCR statement text for coverage.\n" * 80),
                [],
            ),
        ],
        ocr_used=True,
        ocr_pages=2,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    tax_findings = [finding for finding in result.findings if "Profit/loss after tax" in finding.issue]

    assert not tax_findings
    assert "current-year after-tax value not confidently extracted" in result.metrics["checks_skipped"]
    assert "Corroborating lines indicate (448,178)" in result.metrics["checks_skipped"]
    assert "Manual confirmation required" in result.metrics["checks_skipped"]


def test_ocr_income_incomplete_after_tax_skips_arithmetic_instead_of_mixing_years(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of profit or loss",
                        "Revenue 707,189 297,041",
                        "Loss before taxation (221,494) (173,516)",
                        "Taxation (226,684) 53,127",
                        "Loss for the year (999,999)",
                    ]
                )
                + "\n"
                + ("Additional OCR statement text for coverage.\n" * 80),
                [],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))

    assert "Skipped / OCR conflict - current-year after-tax value not confidently extracted." in result.metrics["checks_skipped"]
    assert not any("Expected 5,190" in finding.evidence for finding in result.findings)
    assert not any("reported -120,389" in finding.evidence for finding in result.findings)


def test_ocr_statement_mismatches_are_review_prompts_not_high(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of financial position",
                        "Total assets 100 90",
                        "Total equity and liabilities 80 70",
                    ]
                )
                + "\n"
                + ("Additional OCR statement text for coverage.\n" * 80),
                [],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )
    monkeypatch.setattr(reviewer, "extract_pdf", lambda _path: document)

    result = review_pdf("unused.pdf", options=ReviewOptions(run_cautious_note_agreement=True))
    total_findings = [finding for finding in result.findings if finding.category == "Totals and rounding"]

    assert total_findings
    assert all(finding.severity in {"Medium", "Low"} for finding in total_findings)
    assert all("Possible mismatch from OCR line extraction" in finding.issue for finding in total_findings)


def test_ocr_sfp_statement_specific_checks_skip_low_confidence_rows():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\nTotal assets ########",
                [[["Description", "2025"], ["Non-current assets", "100 50"], ["Total assets", "150"]]],
            )
        ],
        ocr_used=True,
        ocr_pages=1,
    )

    findings = check_rounding_and_casting(document, tolerance=Decimal("1"))

    assert any("statement-specific ocr checks were skipped" in finding.issue.lower() for finding in findings)
    assert not any(finding.category == "Totals and rounding" for finding in findings)


def test_formatting_flags_missing_comparative_period():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position 2025\nRevenue 100\nTotal assets 500",
                [],
            )
        ]
    )

    findings = check_formatting(document, CompanyProfile())

    assert any("only one reporting year" in finding.issue.lower() for finding in findings)


def test_notes_check_flags_eps_formula_difference():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of profit or loss\nProfit for the year Note 9 100\n"
                + ("ordinary shares earnings per share disclosure\n" * 80)
                + "Notes to the financial statements\n9 Earnings per share\nProfit attributable 100\nWeighted average shares 20\nBasic EPS 8",
                [],
            )
        ]
    )

    findings = check_notes_agreement(document)

    assert any("eps calculation" in finding.issue.lower() for finding in findings)


def test_standard_checklist_flags_revenue_disclosure_gap():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of profit or loss\nContract asset 500\nNotes to the financial statements\nRevenue is recognised when control transfers.",
                [],
            )
        ]
    )

    findings = check_standard_checklist(document, CompanyProfile())

    assert any(finding.category == "Standards checklist" and "IFRS 15" in finding.location for finding in findings)


def test_standard_checklist_specific_triggers_avoid_irrelevant_eps_and_segments():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Customer segment credit risk ageing\nSegment by customer type\nOrdinary members attended training",
                [],
            )
        ]
    )

    findings = check_standard_checklist(document, CompanyProfile())

    assert not any("IAS 33" in finding.location for finding in findings)
    assert not any("IFRS 8" in finding.location for finding in findings)


def test_eps_note_check_does_not_trigger_from_contingencies_or_steps_word_fragment():
    findings: list[reviewer.Finding] = []

    reviewer._check_eps_note(
        findings,
        "25",
        "25. Contingencies\nManagement outlined the next steps in resolving the matter.\nNo earnings per claim were presented.",
    )

    assert findings == []


def test_standard_checklist_suppresses_tax_exempt_and_no_subsequent_events():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "The entity is tax-exempt and not subject to income tax.\nSubsequent events\nThere were no subsequent events after the reporting period.",
                [],
            )
        ]
    )

    findings = check_standard_checklist(document, CompanyProfile())

    assert not any("IAS 12" in finding.location for finding in findings)
    assert not any("IAS 10" in finding.location for finding in findings)


def test_standard_checklist_can_be_forced_by_area():
    document = PdfDocument([PdfPage(1, "Notes to the financial statements\nLease expense 20", [])])

    findings = check_standard_checklist(document, CompanyProfile(checklist_areas=("IFRS 16",)))

    assert any("IFRS 16" in finding.location for finding in findings)


def test_standard_checklist_disabled_for_local_gaap():
    document = PdfDocument([PdfPage(1, "Revenue 500", [])])

    findings = check_standard_checklist(document, CompanyProfile(presentation_standard="Local GAAP"))

    assert findings == []


def test_extraction_quality_flags_scanned_pdf_like_document():
    document = PdfDocument([PdfPage(1, "", []), PdfPage(2, "", [])])

    findings = check_extraction_quality(document)

    assert findings
    assert findings[0].category == "Extraction quality"


def test_extraction_quality_flags_placeholder_values():
    document = PdfDocument([PdfPage(1, "Revenue #####\nTotal income #####", [])])

    findings = check_extraction_quality(document)

    assert any("Unreadable" in finding.issue for finding in findings)
    assert any("too low" in finding.issue for finding in findings)


def test_extraction_quality_flags_merged_numeric_cells():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Revenue 100 90\n" * 120,
                [[["Description", "2025", "2024"], ["Revenue", "100 90", ""], ["Total", "100", "90"]]],
            )
        ]
    )

    findings = check_extraction_quality(document)

    assert any("merged values" in finding.issue.lower() for finding in findings)


def test_clean_text_document_has_high_extraction_confidence():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of financial position\n" + ("Revenue 100 90\n" * 120),
                [[["Description", "2025", "2024"], ["Revenue", "100", "90"], ["Total revenue", "100", "90"]]],
            )
        ]
    )

    assert document.extraction_profile == "text-based"
    assert document.extraction_confidence >= 90


def test_text_based_document_with_layout_noise_is_not_blocked():
    noisy_table = [["Description", "2025", "2024"]]
    noisy_table.extend([["Line item", "100 90", ""] for _ in range(20)])
    document = PdfDocument(
        [PdfPage(1, "Statement of financial position\n" + ("Revenue 100 90\n" * 120), [noisy_table])]
    )

    findings = check_extraction_quality(document)

    assert document.extraction_confidence >= 50
    assert not any("too low" in finding.issue for finding in findings)


def test_ocr_fallback_does_not_crash_when_tesseract_missing(monkeypatch):
    monkeypatch.setattr(extraction, "_resolve_tesseract", lambda: "")
    base_document = PdfDocument([PdfPage(1, "", [])])

    document = extract_pdf_with_ocr("missing.pdf", base_document, ReviewOptions(use_ocr=True))

    assert document.ocr_error
    assert document.pages == base_document.pages


def test_review_options_enable_ocr_without_changing_defaults():
    options = ReviewOptions(use_ocr=True, ocr_max_pages=10, ocr_dpi=150)

    assert options.use_ocr is True
    assert options.ocr_max_pages == 10
    assert options.ocr_dpi == 150


def test_ocr_line_to_table_row_splits_label_and_amounts():
    row = _line_to_table_row("Revenue from contracts 12,500 10,250")

    assert row == ["Revenue from contracts", "12,500", "10,250"]


def test_ocr_line_to_table_row_drops_note_reference_column():
    row = _line_to_table_row("Cash and cash equivalents 24 11,343,889 5,102,400")

    assert row == ["Cash and cash equivalents", "11,343,889", "5,102,400"]


def test_ocr_line_to_table_row_ignores_docusign_header():
    row = _line_to_table_row("DocuSign Envelope ID 4895-2047-947")

    assert row is None


def test_ocr_table_reconstruction_builds_candidate_table():
    lines = [
        {"text": "Revenue 100 90"},
        {"text": "Cost of sales (60) (50)"},
        {"text": "Total profit 40 40"},
    ]

    tables = _reconstruct_ocr_tables(lines)

    assert tables == [[["Revenue", "100", "90"], ["Cost of sales", "(60)", "(50)"], ["Total profit", "40", "40"]]]


def test_metrics_include_ocr_table_count():
    document = PdfDocument(
        [PdfPage(1, "Revenue 100\n" * 120, [[["Revenue", "100"], ["Total revenue", "100"]]])],
        ocr_used=True,
        ocr_pages=1,
        ocr_tables=1,
    )

    findings = check_extraction_quality(document)

    assert findings[0].severity == "Low"
    assert "reconstructed 1 table" in findings[0].evidence
    assert document.ocr_tables == 1


def test_notes_agreement_flags_missing_note_reference_on_face_statement(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of profit or loss\nRevenue 10,000 8,000", []),
            PdfPage(2, "Notes to the financial statements\n1. Revenue\nSales of goods 10,000", []),
        ]
    )
    monkeypatch.setattr(PdfDocument, "table_extraction_confidence", property(lambda self: 100))
    findings = check_notes_agreement(document)
    finding = next((f for f in findings if "lacks a note reference" in f.issue), None)
    assert finding is not None
    assert finding.severity == "Medium"
    assert "Revenue" in finding.issue
    assert "Note 1" in finding.issue
    assert "Suggested Note 1" in finding.evidence


def test_notes_agreement_flags_missing_note_reference_when_no_note_found(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of profit or loss\nAssets 5,000 4,000", []),
            PdfPage(2, "Notes to the financial statements\n1. Revenue\nSales of goods 10,000", []),
        ]
    )
    monkeypatch.setattr(PdfDocument, "table_extraction_confidence", property(lambda self: 100))
    findings = check_notes_agreement(document)
    finding = next((f for f in findings if "has no note reference" in f.issue), None)
    assert finding is not None
    assert finding.severity == "Low"
    assert "Assets" in finding.issue
    assert "no matching note was found" in finding.issue


def test_notes_agreement_does_not_flag_cash_flow_working_capital_lines_without_note_refs(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of Cash Flows\n"
                "Profit/(loss) before taxation 125,897 (538)\n"
                "Loans and advance (5,568,284) -\n"
                "Other receivables (1,961,318) -\n"
                "Other payable 3,180,280 538\n"
                "Total cash movement for the year 2,916,467 -\n"
                "Cash and cash equivalents at the end of the year 3 2,916,467 -",
                [],
            ),
            PdfPage(
                2,
                "Notes to the financial statements\n"
                "3. Cash and cash equivalents\nTotal 2,916,467\n"
                "4. Other receivables\nTotal 1,961,318\n"
                "5. Loans and advances\nTotal 5,568,284\n"
                "9. Other payables\nTotal 3,180,280",
                [],
            ),
        ]
    )
    monkeypatch.setattr(PdfDocument, "table_extraction_confidence", property(lambda self: 100))

    findings = check_notes_agreement(document)

    issues = [f.issue for f in findings]
    assert not any("Profit/(loss) before taxation lacks a note reference" in issue for issue in issues)
    assert not any("Loans And Advance lacks a note reference" in issue for issue in issues)
    assert not any("Other Receivables lacks a note reference" in issue for issue in issues)
    assert not any("Other Payable lacks a note reference" in issue for issue in issues)
    assert not any("Cash Movement For The Year lacks a note reference" in issue for issue in issues)


def test_skipped_note_tables_do_not_create_findings(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Notes to the financial statements\n17. Operating expenses\nAdvertising 421,366 188,606\nTravel 199,281 148,968\nTotal 1,888,693 1,725,336",
                [[["Line item", "2025", "2024"], ["Advertising", "421,366", "188,606"], ["Travel", "199,281", "148,968"], ["Total", "1,888,693", "1,725,336"]]],
            )
        ]
    )
    monkeypatch.setattr(PdfDocument, "table_extraction_confidence", property(lambda self: 100))
    monkeypatch.setattr(reviewer, "_notes_start_page", lambda _document: 1)
    monkeypatch.setattr(reviewer, "_classified_primary_statement_pages", lambda _document: {})
    monkeypatch.setattr(reviewer, "_check_simple_note_table_casting", lambda *_args, **_kwargs: [])
    monkeypatch.setattr(reviewer, "_check_simple_note_text_casting", lambda *_args, **_kwargs: [])
    document.skipped_table_details = []

    findings = reviewer.check_totals_and_rounding(document)

    assert getattr(document, "skipped_table_details", [])
    assert not any("table(s) skipped for generic arithmetic review" in finding.issue for finding in findings)


def test_notes_agreement_matches_negative_statement_amount_to_positive_tax_note_total(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(1, "Statement of profit or loss\nTaxation 19 (43,722) -", []),
            PdfPage(
                2,
                "Notes to the financial statements\n19. Taxation\n"
                "Major components of the tax expense\n"
                "Current tax 41,069 -\nDeferred tax 2,653 -\n43,722 -",
                [],
            ),
        ]
    )
    monkeypatch.setattr(PdfDocument, "table_extraction_confidence", property(lambda self: 100))

    findings = check_notes_agreement(document)

    assert not any(
        finding.category == "Notes agreement"
        and finding.location == "Note 19"
        and "not found in the related note text" in finding.issue
        for finding in findings
    )


def test_parse_statement_note_line_detects_implicit_note_before_dash_placeholders():
    parsed = reviewer._parse_statement_note_line(
        "Property, plant and equipment 3 - -",
        12,
        "Statement of financial position",
    )

    assert parsed is not None
    assert parsed.ref == "3"
    assert parsed.line_item == "property plant and equipment"
    assert parsed.amounts == ()


def test_notes_agreement_does_not_flag_cash_flow_balance_or_interest_lines_without_note_refs(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of cash flows\n"
                "Interest income 37,617 586\n"
                "Cash from financing activities (24,332) (43,639)\n"
                "Cash and cash equivalents at the beginning of the year 87,815 39,968\n"
                "Effect of exchange rate movement on cash balances (142,822) -\n",
                [],
            ),
            PdfPage(
                2,
                "Notes to the financial statements\n"
                "8. Cash and cash equivalents\n87,815 39,968\n"
                "18. Investment income\n37,617 586",
                [],
            ),
        ]
    )
    monkeypatch.setattr(PdfDocument, "table_extraction_confidence", property(lambda self: 100))

    findings = check_notes_agreement(document)
    issues = [f.issue for f in findings]

    assert not any("Interest Income lacks a note reference" in issue for issue in issues)
    assert not any("Cash From Financing Activities lacks a note reference" in issue for issue in issues)
    assert not any("Cash And Cash Equivalents At The Beginning Of The Year lacks a note reference" in issue for issue in issues)
    assert not any("Effect Of Exchange Rate Movement On Cash Balances lacks a note reference" in issue for issue in issues)


def test_notes_agreement_skips_cash_flow_adjustment_amount_checks_even_with_explicit_note_refs(monkeypatch):
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of cash flows\n"
                "Loss on disposal of fixed asset 3 - 26,383\n"
                "Depreciation 13 - 30,216\n"
                "Finance costs 15 - 868\n",
                [],
            ),
            PdfPage(
                2,
                "Notes to the financial statements\n"
                "3. Property, plant and equipment\nCarrying amount 100 90\n"
                "13. Depreciation expense\n30,216 30,216\n"
                "15. Finance costs\n868 868",
                [],
            ),
        ]
    )
    monkeypatch.setattr(PdfDocument, "table_extraction_confidence", property(lambda self: 100))

    findings = check_notes_agreement(document)

    assert not any(
        finding.category == "Notes agreement"
        and "not found in the related note text" in finding.issue
        for finding in findings
    )


def test_simple_note_text_casting_ignores_page_footer_like_total_lines():
    page = PdfPage(
        28,
        "\n".join(
            [
                "14. Operating expenses",
                "Auditors remuneration 4,677 5,422",
                "Bank charges - 28",
                "Professional fees 40,985 49,806",
                "Office expenses 352 21,487",
                "Loss on disposal - 6,426",
                "Restructuring expense 180 1,652",
                "Fines and penalties 10,113 -",
                "Security - 503",
                "Technical service fees 34,418 372,009",
                "Telecommunication expenses 3,978 3,805",
                "Transportation and travelling - 11,646",
                "94,703 472,784",
                "27",
                "DRAFT",
            ],
        ),
        [],
    )

    findings = reviewer._check_simple_note_text_casting(page, Decimal("1"))

    assert not any(f.severity != "Passed" and "Note 14 Operating expenses" in f.location for f in findings)


def test_line_item_not_face_linked_covers_profit_before_financing_and_cash_flow_buckets():
    assert reviewer._line_item_not_face_linked(
        "Profit before financing and income taxes",
        "Statement of profit or loss and other comprehensive income",
        False,
    )
    assert reviewer._line_item_not_face_linked(
        "Cash from financing activities",
        "Statement of cash flows",
        False,
    )
    assert reviewer._line_item_not_face_linked(
        "Cash at the beginning of the year",
        "Statement of cash flows",
        False,
    )


def test_simple_note_text_casting_skips_reconciliation_style_note_sections():
    page = PdfPage(
        37,
        "\n".join(
            [
                "14. Deposit for Shares",
                "As at 1 January 351,370 439,789",
                "Movement in the year - (88,419)",
                "As at 31 December 351,370 351,370",
            ],
        ),
        [],
    )

    findings = reviewer._check_simple_note_text_casting(page, Decimal("1"))

    assert not findings



def test_not_elevated_review_prompts_are_not_counted_as_skipped_checks():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "Statement of profit or loss\nRevenue Note 9 100 90\nNotes to the financial statements\n9 Revenue\n",
                [],
            )
        ]
    )
    finding = reviewer.Finding(
        "Notes agreement",
        "Low",
        "Page 1 | Note 9",
        "Amount not located in referenced note.",
        "Revenue Note 9 amount could not be corroborated from the detected note text.",
        "Review the source statement and note manually.",
        {"match_confidence": "Low", "referenced_note": "9", "page_reference": "Page 1"},
    )

    result = reviewer._build_result(
        document,
        [finding],
        checks_performed=["Cautious face-to-note amount agreement performed in review-prompt mode."],
    )

    assert result.metrics["review_prompts_not_elevated_count"] == 1
    assert "low-confidence review prompt" not in result.metrics["checks_skipped"].lower()
    assert result.metrics["checks_skipped"] == "No major checks skipped."


def test_skipped_cash_flow_check_becomes_manual_review_required():
    document = PdfDocument(
        [
            PdfPage(
                9,
                "Statement of cash flows\nCash generated from operations 100 90",
                [],
            )
        ]
    )

    result = reviewer._build_result(
        document,
        [],
        checks_skipped=["Statement of cash flows: Page 9 skipped because table structure was not confidently parsed."],
    )

    manual_findings = [finding for finding in result.findings if finding.category == "Manual review"]
    assert len(manual_findings) == 1
    assert manual_findings[0].location == "Page 9"
    assert "manual review required" in manual_findings[0].issue.lower()
    assert any(
        row["Result"] == "Manual review required" and "cash flows" in row["Check"].lower()
        for row in result.metrics["check_results"]
    )


def test_ai_openai_call_uses_bounded_timeout(monkeypatch):
    captured = {}

    class FakeResponse:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self):
            return b'{"output_text":"{}"}'

    def fake_urlopen(_req, timeout):
        captured["timeout"] = timeout
        return FakeResponse()

    monkeypatch.setattr(ai_policy_review, "_rate_limit_wait_seconds", lambda: 0)
    monkeypatch.setattr(ai_policy_review.request, "urlopen", fake_urlopen)

    ai_policy_review._call_openai("test-key", {"model": "test", "input": []})

    assert captured["timeout"] == ai_policy_review.AI_REQUEST_TIMEOUT_SECONDS
    assert captured["timeout"] < 60


def test_ai_openai_call_retries_rate_limit_with_retry_after(monkeypatch):
    from urllib import error as url_error

    attempts = []
    sleeps = []

    class FakeResponse:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self):
            return b'{"output_text":"{}"}'

    def fake_sleep(seconds):
        sleeps.append(seconds)
        ai_policy_review._AI_RATE_LIMIT_UNTIL = 0

    def fake_urlopen(_req, timeout):
        attempts.append(timeout)
        if len(attempts) == 1:
            raise url_error.HTTPError(
                url="https://api.openai.com/v1/responses",
                code=429,
                msg="Too Many Requests",
                hdrs={"Retry-After": "7"},
                fp=BytesIO(b"rate limited"),
            )
        return FakeResponse()

    ai_policy_review._AI_RATE_LIMIT_UNTIL = 0
    monkeypatch.setattr(ai_policy_review, "AI_MAX_ATTEMPTS", 2)
    monkeypatch.setattr(ai_policy_review.request, "urlopen", fake_urlopen)
    monkeypatch.setattr(ai_policy_review, "_sleep_before_ai_retry", fake_sleep)

    result = ai_policy_review._call_openai("test-key", {"model": "test", "input": []})

    assert result == {"output_text": "{}"}
    assert len(attempts) == 2
    assert 7 in sleeps


def test_skipped_table_summary_marks_supplementary_schedules_as_intentional_exclusions():
    document = PdfDocument(
        [
            PdfPage(
                7,
                "Supplementary schedules",
                [
                    [["Value added statement", "2025", "2024"], ["Revenue", "100", "90"], ["Total value added", "100", "90"]],
                    [["Five year financial summary", "2025", "2024"], ["Revenue", "100", "90"], ["Total assets", "200", "180"]],
                ],
            )
        ]
    )

    check_rounding_and_casting(document, tolerance=Decimal("1"))
    summary = reviewer._skipped_table_summary_rows(document)

    value_added = next(row for row in summary if row["Skipped check group"] == "Value-added statement")
    multi_year = next(row for row in summary if row["Skipped check group"] == "Multi-year summary")
    assert value_added["Can automated check be fixed?"] == "Not applicable"
    assert multi_year["Can automated check be fixed?"] == "Not applicable"
    assert "supplementary" in value_added["Why reviewer should review"].lower()
    assert "several years" in multi_year["Why reviewer should review"].lower()


def test_parse_skipped_check_adds_cross_source_page_context():
    row = parse_skipped_check(
        "Cross-source income-to-equity linkage skipped because only one of income or equity reference lines was confidently parsed. "
        "Available evidence: income result Page 12; equity movement not parsed."
    )

    assert row["Check area"] == "Cross-source income-to-equity linkage"
    assert row["Page reference"] == "Page 12"
    assert "income result Page 12" in row["Reason skipped"]
    assert row["Can automated check be fixed?"] == "Partially"


def test_greystone_style_drafting_issues_are_flagged_with_page_context():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "The annual report and financial statements on pagess 10 to 36 were approved by the Board.",
                        "Financial liabilities at fair value through profit or loss or loss are recognised.",
                        "Transaction costs are recognised in profit or loss or loss.",
                        "For these leases, the payment is recognised as a direct cost (note ) on a straight-line basis.",
                        "Ordinary shares are classified as . Mandatorily redeemable preference shares are liabilities.",
                        "Deferred tax asset of N284,9451,338 was not recognised.",
                        "Effect of development levy - % 1,955 - % -",
                        "including governance, strategy, risk management, metrics and tragets, and the effect of climated related matters",
                    ]
                ),
                [],
            )
        ]
    )

    findings, export = check_cross_page_consistency(document)
    issues = [row["Issue"] for row in export["grammar"]]

    assert any("pagess" in issue for issue in issues)
    assert sum(1 for issue in issues if "profit or loss or loss" in issue) == 1
    assert any("Blank note reference" in issue for issue in issues)
    assert any("Incomplete sentence" in issue for issue in issues)
    assert any("malformed thousands" in issue for issue in issues)
    assert any("missing percentage" in issue for issue in issues)
    assert any("tragets" in issue for issue in issues)
    assert any("climated" in issue for issue in issues)
    assert all(f.location == "Page 1" for f in findings if f.category == "Formatting")


def test_cash_flow_source_amount_tieout_flags_one_unit_differences_without_sign_noise():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of Comprehensive Income",
                        "2025 2024",
                        "Other operating income 14 116,364 88,383",
                        "Finance costs 17 (324,994) (438,700)",
                        "(Loss)/profit for the year (245,475) 4,891,754",
                    ]
                ),
                [],
            ),
            PdfPage(
                2,
                "\n".join(
                    [
                        "Statement of Cash Flows",
                        "2025 2024",
                        "Amortisation of right-of-use asset 3 - 124,980",
                        "Interest received on loan 4 (116,364) (88,383)",
                        "Finance costs 17 324,995 438,700",
                        "Total cash at end of the year 6 215,166 314,032",
                    ]
                ),
                [],
            ),
            PdfPage(
                3,
                "\n".join(
                    [
                        "Notes to the Financial Statements",
                        "16. Other operating expenses",
                        "Depreciation - 124,979",
                        "17. Finance costs",
                        "Interest expenses 324,994 438,700",
                    ]
                ),
                [],
            ),
        ]
    )

    findings, performed = reviewer._check_cash_flow_supporting_amounts(document, Decimal("1"))

    assert performed
    assert findings
    evidence = findings[0].evidence
    assert "difference 1" in evidence
    assert "difference 649,989" not in evidence
    assert "Interest received on loan" not in evidence


def test_supporting_disclosure_note_reference_tieout_ignores_standards_dates():
    document = PdfDocument(
        [
            PdfPage(1, "Statement of financial position\nOther receivables 5 323,614 120,512", []),
            PdfPage(
                2,
                "\n".join(
                    [
                        "Notes to the Financial Statements",
                        "5. Other receivables",
                        "Other receivables 323,614 120,512",
                        "8. Borrowings",
                        "Loans 3,224,255 3,147,828",
                    ]
                ),
                [],
            ),
            PdfPage(
                3,
                "\n".join(
                    [
                        "20. Risk management",
                        "The standard is effective for annual reporting periods beginning on or after 1 January 2027.",
                        "Credit risk",
                        "Other receivables 5 1,474,662 - 1,474,662",
                        "Borrowings 8 3,224,254 - 3,224,254",
                    ]
                ),
                [],
            ),
        ]
    )

    findings, performed = reviewer._check_supporting_disclosure_note_reference_amounts(document, Decimal("1"))

    assert performed
    assert findings
    evidence = findings[0].evidence
    assert "Other receivables references Note 5" in evidence
    assert "Borrowings references Note 8" in evidence
    assert "beginning on or after" not in evidence


def test_supplementary_summary_consistency_flags_current_year_mismatches():
    document = PdfDocument(
        [
            PdfPage(
                1,
                "\n".join(
                    [
                        "Statement of Financial Position as at 31 December 2025",
                        "Retained income (2,624,236) (2,378,761)",
                        "Total Liabilities 4,311,564 3,942,462",
                        "Total Equity and Liabilities 1,689,828 1,566,201",
                    ]
                ),
                [],
            ),
            PdfPage(
                2,
                "\n".join(
                    [
                        "Statement of Comprehensive Income",
                        "Revenue 13 - 389,382",
                        "Other operating income 14 116,364 88,383",
                        "Other operating expenses 16 (50,027) (160,667)",
                        "Operating profit 80,262 5,358,295",
                        "Finance costs 17 (324,994) (438,700)",
                        "(Loss)/profit before taxation (244,732) 4,919,595",
                        "Taxation 12 (743) (27,841)",
                        "(Loss)/profit for the year (245,475) 4,891,754",
                    ]
                ),
                [],
            ),
            PdfPage(
                3,
                "\n".join(
                    [
                        "Five Year Financial Summary",
                        "2025 2024 2023",
                        "Other operating expenses (50,026) (160,667) (262,281)",
                        "Operating profit/(loss) 80,263 5,358,295 (4,496,406)",
                        "Finance costs (324,995) (438,700) (306,083)",
                        "Retained income (2,624,235) (2,378,761) (7,270,515)",
                        "Total liabilities 4,311,563 3,942,462 9,179,029",
                    ]
                ),
                [],
            ),
        ]
    )

    findings, performed = reviewer._check_supplementary_summary_consistency(document, Decimal("1"))

    assert performed
    assert findings
    evidence = findings[0].evidence
    assert "Other operating expenses" in evidence
    assert "Operating profit/loss" in evidence
    assert "Total liabilities" in evidence

def test_ai_openai_call_defers_when_another_ai_request_is_active(monkeypatch):
    ai_policy_review._AI_RATE_LIMIT_UNTIL = 0
    monkeypatch.setattr(ai_policy_review, "AI_REQUEST_LOCK_TIMEOUT_SECONDS", 0.1)
    acquired = ai_policy_review._AI_REQUEST_LOCK.acquire(blocking=False)
    assert acquired
    try:
        try:
            ai_policy_review._call_openai("test-key", {"model": "test", "input": []})
        except RuntimeError as exc:
            message = str(exc)
        else:
            raise AssertionError("Expected a busy AI request to be deferred.")
    finally:
        ai_policy_review._AI_REQUEST_LOCK.release()

    assert "AI service busy" in message
    assert ai_policy_review._is_rate_limit_error(RuntimeError(message))


def test_review_document_wires_canonical_qc_into_metrics_and_check_results():
    document = PdfDocument([
        PdfPage(
            1,
            """Example Limited
Statement of Financial Position
2025 2024
N'000 N'000
Non-current assets 60 50
Current assets 40 40
Total assets 100 90
Total equity 30 30
Total liabilities 70 60
Total equity and liabilities 100 90
""",
            [],
        )
    ])
    result = reviewer.review_document(document, "example.pdf", CompanyProfile(), ReviewOptions())
    canonical_rows = result.metrics.get("canonical_recalculation_checks", [])
    assert canonical_rows
    assert any(row.get("Check") == "SFP total assets cast" and row.get("Status") == "Pass" for row in canonical_rows)
    assert any(
        row.get("Check") == "Canonical QC - SFP total assets cast" and row.get("Result") == "Passed"
        for row in result.metrics.get("check_results", [])
    )
    assert result.metrics.get("canonical_extraction_audit")

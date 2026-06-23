from __future__ import annotations

import json
from pathlib import Path

import openpyxl

import cli
import job_runner
from models import CompanyProfile, Finding, ReviewOptions, ReviewResult


def _sample_result() -> ReviewResult:
    return ReviewResult(
        findings=[
            Finding(
                category="Formatting",
                severity="Low",
                location="Page 5",
                issue="Example issue",
                evidence="Example evidence",
                recommendation="Example recommendation",
                metadata={"page_reference": "Page 5", "note_reference": "Note 2"},
            )
        ],
        metrics={
            "pages": 12,
            "tables": 4,
            "findings": 1,
            "high": 0,
            "medium": 0,
            "low": 1,
            "checks_performed_count": 3,
            "checks_passed_count": 2,
            "checks_skipped_count": 1,
            "checks_performed": "Check A\nCheck B",
            "checks_skipped": "Check C",
            "positive_assurance": "No exceptions noted from line-based checks on primary statements.",
            "ai_policy_review_status": "completed",
            "ai_policy_review_message": "",
            "ai_policy_review_summary": "AI policy completed.",
            "ai_policy_review_model": "gpt-5-mini",
            "ai_finding_review_status": "skipped",
            "ai_finding_review_message": "No eligible findings.",
            "ai_finding_review_summary": "",
            "ai_finding_reviewed": 0,
            "ai_finding_suppressed": 0,
            "detected_profile": {
                "Company name": "Example Limited",
                "Entity type": "Private company",
                "Document scope": "Full financial statements",
            },
        },
    )


def test_write_review_outputs_creates_standard_artifacts(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(job_runner, "review_pdf", lambda *args, **kwargs: _sample_result())

    result, summary = job_runner.write_review_outputs(
        tmp_path / "input.pdf",
        CompanyProfile(),
        ReviewOptions(use_ai_policy_review=True),
        tmp_path / "exports",
    )

    assert result.metrics["ai_policy_review_status"] == "completed"
    files = summary["output_files"]
    assert set(files) == {"excel", "markdown", "json_summary"}

    excel_path = Path(files["excel"])
    markdown_path = Path(files["markdown"])
    json_path = Path(files["json_summary"])
    assert excel_path.exists()
    assert markdown_path.exists()
    assert json_path.exists()

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload["company_name"] == "Example Limited"
    assert payload["ai_policy_review_status"] == "completed"
    assert payload["ai_finding_review_status"] == "skipped"

    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    assert "Summary" in workbook.sheetnames
    assert "AI policy judgement" in workbook.sheetnames


def test_ai_verification_errors_flag_unavailable_statuses() -> None:
    result = _sample_result()
    result.metrics["ai_policy_review_status"] = "unavailable"
    result.metrics["ai_finding_review_status"] = "deferred"

    errors = job_runner.ai_verification_errors(
        result,
        require_ai_available=True,
        require_ai_policy_completed=True,
        require_ai_finding_available=True,
    )

    assert any("policy review status is unavailable" in error.lower() for error in errors)
    assert any("finding review status is deferred" in error.lower() for error in errors)


def test_cli_returns_verification_exit_code_when_ai_is_unavailable(monkeypatch, tmp_path: Path) -> None:
    sample = _sample_result()
    sample.metrics["ai_policy_review_status"] = "unavailable"
    sample.metrics["ai_finding_review_status"] = "unavailable"

    monkeypatch.setattr(cli, "review_pdf", lambda *args, **kwargs: sample)
    monkeypatch.setattr(cli, "findings_to_markdown", lambda result: "report")
    monkeypatch.setattr(
        cli,
        "write_review_outputs",
        lambda *args, **kwargs: (
            sample,
            {"output_files": {}, "ai_policy_review_status": "unavailable", "ai_finding_review_status": "unavailable"},
        ),
    )
    monkeypatch.setattr(
        "sys.argv",
        [
            "cli.py",
            str(tmp_path / "input.pdf"),
            "--export-dir",
            str(tmp_path / "exports"),
            "--require-ai-available",
        ],
    )

    assert cli.main() == 2

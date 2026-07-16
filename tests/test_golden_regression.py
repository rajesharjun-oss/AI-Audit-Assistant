from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from models import CompanyProfile, PdfDocument, PdfPage, ReviewOptions
from report_exports import build_excel_export
from reviewer import review_document

ROOT = Path(__file__).resolve().parents[1]
GOLDEN_DIR = ROOT / "tests" / "golden_cases"


def _case_files() -> list[Path]:
    return sorted(GOLDEN_DIR.glob("*.json"))


def _load_case(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _build_document(case: dict) -> PdfDocument:
    pages = [
        PdfPage(
            int(page["number"]),
            str(page.get("text", "")),
            page.get("tables", []) or [],
        )
        for page in case.get("pages", [])
    ]
    document_data = case.get("document", {}) or {}
    return PdfDocument(
        pages,
        ocr_used=bool(document_data.get("ocr_used", False)),
        ocr_pages=int(document_data.get("ocr_pages", 0) or 0),
        ocr_tables=int(document_data.get("ocr_tables", 0) or 0),
        ocr_error=str(document_data.get("ocr_error", "") or ""),
    )


def _profile(data: dict | None) -> CompanyProfile:
    data = data or {}
    return CompanyProfile(
        company_name=data.get("company_name", ""),
        industry=data.get("industry", ""),
        reporting_currency=data.get("reporting_currency", ""),
        expected_policies=tuple(data.get("expected_policies", [])),
        significant_transactions=tuple(data.get("significant_transactions", [])),
        presentation_standard=data.get("presentation_standard", "IFRS"),
        checklist_areas=tuple(data.get("checklist_areas", [])),
    )


def _options(data: dict | None) -> ReviewOptions:
    data = data or {}
    return ReviewOptions(
        use_ocr=bool(data.get("use_ocr", False)),
        ocr_max_pages=data.get("ocr_max_pages", 60),
        ocr_dpi=int(data.get("ocr_dpi", 300) or 300),
        run_cautious_note_agreement=bool(data.get("run_cautious_note_agreement", False)),
        use_ai_policy_review=bool(data.get("use_ai_policy_review", False)),
        use_ai_full_review=bool(data.get("use_ai_full_review", False)),
    )


def _summary_map(workbook) -> dict[str, object]:
    rows = workbook["Summary"].iter_rows(min_row=2, values_only=True)
    return {str(metric): value for metric, value in rows if metric}


def _sheet_text(workbook, sheet_name: str) -> str:
    if sheet_name not in workbook.sheetnames:
        return ""
    values: list[str] = []
    for row in workbook[sheet_name].iter_rows(values_only=True):
        values.extend(str(value) for value in row if value is not None)
    return "\n".join(values)


@pytest.mark.parametrize("case_path", _case_files(), ids=lambda path: path.stem)
def test_golden_regression_case(case_path: Path, tmp_path: Path):
    case = _load_case(case_path)
    document = _build_document(case)
    result = review_document(document, case_path.name, _profile(case.get("profile")), _options(case.get("options")))
    output = tmp_path / f"{case_path.stem}_exception_register.xlsx"
    output.write_bytes(build_excel_export(result))

    assert output.exists()
    assert output.stat().st_size > 0

    workbook = openpyxl.load_workbook(output, data_only=True)
    summary = _summary_map(workbook)
    assertions = case.get("assertions", {}) or {}

    for sheet_name in assertions.get("required_sheets", []):
        assert sheet_name in workbook.sheetnames, f"Required sheet missing: {sheet_name}"

    for metric, expected in assertions.get("expected_summary", {}).items():
        assert summary.get(metric) == expected, f"Summary metric {metric!r} expected {expected!r}, got {summary.get(metric)!r}"

    for metric, maximum in assertions.get("max_summary", {}).items():
        value = summary.get(metric)
        assert value is not None and value <= maximum, f"Summary metric {metric!r} expected <= {maximum}, got {value!r}"

    for metric, minimum in assertions.get("min_summary", {}).items():
        value = summary.get(metric)
        assert value is not None and value >= minimum, f"Summary metric {metric!r} expected >= {minimum}, got {value!r}"

    workbook_text = "\n".join(_sheet_text(workbook, sheet) for sheet in workbook.sheetnames)
    for snippet in assertions.get("required_text", []):
        assert snippet in workbook_text, f"Expected workbook text not found: {snippet}"

    for snippet in assertions.get("forbidden_text", []):
        assert snippet not in workbook_text, f"Forbidden workbook text found: {snippet}"

    for sheet_assertion in assertions.get("required_sheet_text", []):
        sheet_text = _sheet_text(workbook, sheet_assertion["sheet"])
        assert sheet_assertion["contains"] in sheet_text, f"Expected text not found in {sheet_assertion['sheet']}: {sheet_assertion['contains']}"

    for sheet_assertion in assertions.get("forbidden_sheet_text", []):
        sheet_text = _sheet_text(workbook, sheet_assertion["sheet"])
        assert sheet_assertion["contains"] not in sheet_text, f"Forbidden text found in {sheet_assertion['sheet']}: {sheet_assertion['contains']}"


def test_golden_regression_cases_exist():
    assert _case_files(), "Expected at least one synthetic golden regression case."

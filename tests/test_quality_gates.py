from io import BytesIO

import openpyxl

import reviewer
from models import Finding, PdfDocument, PdfPage
from report_exports import build_excel_export


def test_low_confidence_ai_policy_prompt_is_visible_but_not_elevated():
    document = PdfDocument([PdfPage(1, "Notes to the financial statements\n1. Significant accounting policies", [])])
    weak_prompt = Finding(
        category="AI policy judgement",
        severity="Low",
        location="Document-wide",
        issue="Generic consolidation wording may need review.",
        evidence="AI saw related party wording but no subsidiary or consolidation balance evidence.",
        recommendation="Review only if group structure indicators exist.",
        metadata={"match_confidence": "Low", "check_type": "AI policy judgement"},
    )

    result = reviewer._build_result(document, [weak_prompt], ["Synthetic policy judgement check"], [])

    assert result.findings == []
    assert result.metrics["findings"] == 0
    assert result.metrics["review_prompts_not_elevated_count"] == 1
    assert "Review prompts not elevated" in result.metrics["checks_skipped"]

    workbook = openpyxl.load_workbook(BytesIO(build_excel_export(result)), data_only=True)
    assert "Review prompts not elevated" in workbook.sheetnames
    prompts = list(workbook["Review prompts not elevated"].iter_rows(min_row=2, values_only=True))
    assert any("Generic consolidation" in str(cell or "") for row in prompts for cell in row)
    exceptions = list(workbook["Exception register"].iter_rows(min_row=2, values_only=True))
    assert any("No automated findings" in str(cell or "") for row in exceptions for cell in row)


def test_strong_page_backed_finding_stays_in_exception_register():
    document = PdfDocument([PdfPage(4, "Statement of profit or loss\nRevenue 100\nTotal 90", [])])
    finding = Finding(
        category="Totals and rounding",
        severity="Medium",
        location="Page 4 | Table 1",
        issue="Visible subtotal does not agree to components.",
        evidence="Page 4: reported 90, visible sum 100, difference 10.",
        recommendation="Recast the table on Page 4.",
    )

    result = reviewer._build_result(document, [finding], ["Synthetic casting check"], [])

    assert len(result.findings) == 1
    assert result.metrics["findings"] == 1
    assert result.metrics["review_prompts_not_elevated_count"] == 0


def test_note_low_confidence_prompt_is_not_hidden_or_counted_as_exception():
    document = PdfDocument([PdfPage(8, "Statement of financial position\nTrade receivables Note 5 100", [])])
    prompt = Finding(
        category="Notes agreement",
        severity="Low",
        location="Page 8 | Note 5",
        issue="Amount not located in referenced note.",
        evidence="Page 8: statement line references Note 5, but note extraction confidence was low.",
        recommendation="Review Note 5 manually.",
        metadata={"referenced_note": "5", "match_confidence": "Low", "page_reference": "Page 8"},
    )

    result = reviewer._build_result(document, [prompt], ["Cautious note agreement check"], [])

    assert result.findings == []
    assert result.metrics["review_prompts_not_elevated_count"] == 1
    row = result.metrics["review_prompts_not_elevated"][0]
    assert row["Note reference"] == "5"
    assert "Low-confidence note agreement" in row["Reason not elevated"]
